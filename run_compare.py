"""
EDC ↔ Collibra Data Comparator
A professional-grade GUI tool for comparing EDC and Collibra metadata files.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import customtkinter as ctk
import pandas as pd
import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import threading

# ─── Theme ────────────────────────────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

DARK_BG      = "#0D1117"
PANEL_BG     = "#161B22"
CARD_BG      = "#1C2128"
BORDER       = "#30363D"
ACCENT_BLUE  = "#2F81F7"
ACCENT_TEAL  = "#3DC9B0"
ACCENT_GREEN = "#3FB950"
ACCENT_RED   = "#F85149"
ACCENT_GOLD  = "#D29922"
TEXT_PRIMARY = "#E6EDF3"
TEXT_MUTED   = "#8B949E"
TEXT_DIM     = "#484F58"

# ─── Full Name Parser ─────────────────────────────────────────────────────────
def parse_full_name(full_name: str) -> dict:
    """
    Parse a Collibra Full Name string like:
    ADCX_OOO_DBV_SXC>hive_metastone>auto>tdxc_cfhb_uuu_delp>cra_cred_pau_his(column)
    Returns dict with azure_server, data_store, mal_code, table_name, column_name.
    """
    if not isinstance(full_name, str) or not full_name.strip():
        return {k: "" for k in ["azure_server","data_store","mal_code","table_name","column_name"]}
    parts = [p.strip() for p in full_name.split(">")]
    col_name = ""
    if parts:
        last = parts[-1]
        m = re.match(r'^(.*?)\s*\(column\)\s*$', last, re.IGNORECASE)
        if m:
            col_name = m.group(1).strip()
            parts[-1] = col_name
    keys = ["azure_server","data_store","mal_code","table_name","column_name"]
    result = {k: "" for k in keys}
    for i, k in enumerate(keys):
        result[k] = parts[i] if i < len(parts) else ""
    if col_name:
        result["column_name"] = col_name
    return result


# ─── Comparison Engine ────────────────────────────────────────────────────────
def normalise(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return str(val).strip().lower()


def run_comparison(edc_df: pd.DataFrame,
                   col_df: pd.DataFrame,
                   edc_fields: list,
                   col_fields: list) -> dict:
    """
    Core comparison logic.
    Returns a rich results dict with matches, mismatches, only-in-edc, only-in-collibra,
    and a full merged detail dataframe.
    """
    # build normalised keys for each side
    edc_norm = edc_df[edc_fields].apply(
        lambda row: tuple(normalise(row[c]) for c in edc_fields), axis=1
    )
    col_norm = col_df[col_fields].apply(
        lambda row: tuple(normalise(row[c]) for c in col_fields), axis=1
    )

    edc_key_set = set(edc_norm)
    col_key_set = set(col_norm)

    matches        = edc_key_set & col_key_set
    only_in_edc    = edc_key_set - col_key_set
    only_in_col    = col_key_set - edc_key_set

    edc_df = edc_df.copy()
    col_df = col_df.copy()
    edc_df["__key__"] = edc_norm
    col_df["__key__"] = col_norm
    edc_df["__status__"] = edc_df["__key__"].apply(
        lambda k: "✅ Match" if k in col_key_set else "❌ Only in EDC"
    )
    col_df["__status__"] = col_df["__key__"].apply(
        lambda k: "✅ Match" if k in edc_key_set else "❌ Only in Collibra"
    )

    return {
        "edc_detail":     edc_df,
        "col_detail":     col_df,
        "matches":        matches,
        "only_in_edc":    only_in_edc,
        "only_in_col":    only_in_col,
        "edc_total":      len(edc_df),
        "col_total":      len(col_df),
        "match_count":    len(matches),
        "only_edc_count": len(only_in_edc),
        "only_col_count": len(only_in_col),
    }


# ─── Excel Export ─────────────────────────────────────────────────────────────
def export_to_excel(results: dict, edc_path: str, col_path: str, out_path: str):
    wb = Workbook()

    HDR_FILL   = PatternFill("solid", fgColor="1C3A5E")
    MATCH_FILL = PatternFill("solid", fgColor="1A3A2A")
    FAIL_FILL  = PatternFill("solid", fgColor="3A1A1A")
    ONLY_FILL  = PatternFill("solid", fgColor="2A2A1A")
    HDR_FONT   = Font(bold=True, color="E6EDF3", name="Calibri", size=11)
    BODY_FONT  = Font(color="E6EDF3", name="Calibri", size=10)
    THIN       = Border(
        left=Side(style="thin", color="30363D"),
        right=Side(style="thin", color="30363D"),
        top=Side(style="thin", color="30363D"),
        bottom=Side(style="thin", color="30363D"),
    )

    def style_header(ws, row_num, col_count):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill   = HDR_FILL
            cell.font   = HDR_FONT
            cell.border = THIN
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def style_data_row(ws, row_num, col_count, fill):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill   = fill
            cell.font   = BODY_FONT
            cell.border = THIN
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    def autofit(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

    # ── Summary Sheet ──
    ws_sum = wb.active
    ws_sum.title = "📊 Summary"
    ws_sum.sheet_view.showGridLines = False

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary_rows = [
        ("EDC vs Collibra — Comparison Report", ""),
        ("Generated", now),
        ("EDC File",  os.path.basename(edc_path)),
        ("Collibra File", os.path.basename(col_path)),
        ("", ""),
        ("METRIC", "VALUE"),
        ("EDC Total Rows",         results["edc_total"]),
        ("Collibra Total Rows",    results["col_total"]),
        ("✅ Matches",              results["match_count"]),
        ("❌ Only in EDC",          results["only_edc_count"]),
        ("❌ Only in Collibra",     results["only_col_count"]),
        ("Match Rate",
         f'{results["match_count"]/max(results["edc_total"],1)*100:.1f}%'),
        ("", ""),
        ("EDC File Size",
         f'{os.path.getsize(edc_path)/1024:.1f} KB' if os.path.exists(edc_path) else "N/A"),
        ("Collibra File Size",
         f'{os.path.getsize(col_path)/1024:.1f} KB' if os.path.exists(col_path) else "N/A"),
    ]

    # add DQ stats
    edc_df = results["edc_detail"]
    col_df = results["col_detail"]
    edc_nulls = int(edc_df.isnull().sum().sum())
    col_nulls = int(col_df.isnull().sum().sum())
    edc_dupes = int(edc_df.duplicated().sum())
    col_dupes = int(col_df.duplicated().sum())
    summary_rows += [
        ("", ""),
        ("── DATA QUALITY ──", ""),
        ("EDC Null/Empty Cells",     edc_nulls),
        ("Collibra Null/Empty Cells", col_nulls),
        ("EDC Duplicate Rows",       edc_dupes),
        ("Collibra Duplicate Rows",  col_dupes),
        ("EDC Unique Columns",       len(edc_df.columns) - 2),   # excl __key__, __status__
        ("Collibra Unique Columns",  len(col_df.columns) - 2),
    ]

    for r_idx, (label, val) in enumerate(summary_rows, start=1):
        ws_sum.cell(r_idx, 1, label)
        ws_sum.cell(r_idx, 2, val)
        if r_idx == 1:
            ws_sum.cell(r_idx, 1).font = Font(bold=True, size=14, color="2F81F7", name="Calibri")
        elif label in ("METRIC", "── DATA QUALITY ──"):
            for c in (1, 2):
                ws_sum.cell(r_idx, c).fill = HDR_FILL
                ws_sum.cell(r_idx, c).font = HDR_FONT
        else:
            for c in (1, 2):
                ws_sum.cell(r_idx, c).font = BODY_FONT
                fill = PatternFill("solid", fgColor="161B22")
                ws_sum.cell(r_idx, c).fill = fill

    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 40

    def write_detail_sheet(wb, title, df, status_col="__status__"):
        ws = wb.create_sheet(title)
        ws.sheet_view.showGridLines = False
        display_cols = [c for c in df.columns if c not in ("__key__", "__status__")]
        headers = display_cols + ["Status"]
        for c_idx, h in enumerate(headers, start=1):
            ws.cell(1, c_idx, h)
        style_header(ws, 1, len(headers))
        ws.row_dimensions[1].height = 22

        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            status = row.get(status_col, "")
            if "Match" in str(status):
                fill = MATCH_FILL
            elif "EDC" in str(status):
                fill = FAIL_FILL
            else:
                fill = ONLY_FILL
            for c_idx, col in enumerate(display_cols, start=1):
                ws.cell(r_idx, c_idx, row[col])
            ws.cell(r_idx, len(headers), status)
            style_data_row(ws, r_idx, len(headers), fill)

        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        ws.freeze_panes = "A2"
        autofit(ws)

    write_detail_sheet(wb, "📄 EDC Detail", edc_df)
    write_detail_sheet(wb, "📄 Collibra Detail", col_df)

    # ── Matches only ──
    match_keys = results["matches"]
    matched_edc = edc_df[edc_df["__key__"].isin(match_keys)]
    write_detail_sheet(wb, "✅ Matches", matched_edc)

    # ── Mismatches ──
    only_edc_df = edc_df[edc_df["__key__"].isin(results["only_in_edc"])]
    only_col_df = col_df[col_df["__key__"].isin(results["only_in_col"])]
    write_detail_sheet(wb, "❌ Only in EDC",      only_edc_df)
    write_detail_sheet(wb, "❌ Only in Collibra", only_col_df)

    wb.save(out_path)


# ─── Main App ─────────────────────────────────────────────────────────────────
class DataComparatorApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("EDC ↔ Collibra Data Comparator")
        self.geometry("1400x900")
        self.minsize(1100, 750)
        self.configure(fg_color=DARK_BG)

        self.edc_df:  pd.DataFrame | None = None
        self.col_df:  pd.DataFrame | None = None
        self.edc_path = ""
        self.col_path = ""
        self.results: dict | None = None

        # Preferred default columns
        self.EDC_PRIORITY  = ["Name", "Business Name", "Context", "Description"]
        self.COL_PRIORITY  = ["Name", "Business Name", "Asset Type", "Description"]

        self._build_ui()

    # ── UI Layout ──────────────────────────────────────────────────────────────
    def _build_ui(self):
        # ── Top title bar ──
        title_bar = ctk.CTkFrame(self, fg_color=PANEL_BG, height=56, corner_radius=0)
        title_bar.pack(fill="x", side="top")
        title_bar.pack_propagate(False)

        ctk.CTkLabel(
            title_bar,
            text="⬡  EDC ↔ Collibra  Data Comparator",
            font=ctk.CTkFont("Courier New", 20, "bold"),
            text_color=ACCENT_BLUE
        ).pack(side="left", padx=24, pady=14)

        self.status_label = ctk.CTkLabel(
            title_bar, text="Ready", font=ctk.CTkFont(size=12),
            text_color=TEXT_MUTED
        )
        self.status_label.pack(side="right", padx=24)

        # ── Main paned layout ──
        main = ctk.CTkFrame(self, fg_color=DARK_BG)
        main.pack(fill="both", expand=True, padx=12, pady=(8, 12))

        left = ctk.CTkFrame(main, fg_color=DARK_BG, width=360)
        left.pack(side="left", fill="y", padx=(0, 8))
        left.pack_propagate(False)

        right = ctk.CTkFrame(main, fg_color=DARK_BG)
        right.pack(side="left", fill="both", expand=True)

        self._build_left_panel(left)
        self._build_right_panel(right)

    # ── Left Panel ─────────────────────────────────────────────────────────────
    def _build_left_panel(self, parent):
        scroll = ctk.CTkScrollableFrame(parent, fg_color=DARK_BG, scrollbar_button_color=BORDER)
        scroll.pack(fill="both", expand=True)

        self._section(scroll, "① LOAD FILES")
        self._file_card(scroll, "EDC File", "edc")
        self._file_card(scroll, "Collibra File", "col")

        self._section(scroll, "② FILTER (Collibra)")
        self._filter_card(scroll)

        self._section(scroll, "③ SELECT FIELDS TO COMPARE")
        self._fields_card(scroll)

        self._section(scroll, "④ PARSE FULL NAME")
        self._parse_card(scroll)

        self._section(scroll, "⑤ RUN")
        self._run_card(scroll)

    def _section(self, parent, text):
        ctk.CTkLabel(
            parent, text=text,
            font=ctk.CTkFont("Courier New", 11, "bold"),
            text_color=ACCENT_TEAL
        ).pack(anchor="w", padx=6, pady=(14, 2))

    def _card(self, parent) -> ctk.CTkFrame:
        f = ctk.CTkFrame(parent, fg_color=CARD_BG, corner_radius=8,
                         border_width=1, border_color=BORDER)
        f.pack(fill="x", padx=4, pady=4)
        return f

    def _file_card(self, parent, label, key):
        card = self._card(parent)
        ctk.CTkLabel(card, text=label, font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=TEXT_PRIMARY).pack(anchor="w", padx=12, pady=(10, 2))
        path_var = tk.StringVar(value="No file loaded")
        setattr(self, f"{key}_path_var", path_var)
        ctk.CTkLabel(card, textvariable=path_var, font=ctk.CTkFont(size=10),
                     text_color=TEXT_MUTED, wraplength=300).pack(anchor="w", padx=12)
        btn = ctk.CTkButton(
            card, text=f"Browse {label}…", height=32,
            fg_color=ACCENT_BLUE, hover_color="#1A6FD6",
            font=ctk.CTkFont(size=12),
            command=lambda k=key, lbl=label: self._load_file(k, lbl)
        )
        btn.pack(fill="x", padx=12, pady=(6, 10))

    def _filter_card(self, parent):
        card = self._card(parent)
        ctk.CTkLabel(card, text="Filter by Asset Type:", font=ctk.CTkFont(size=11),
                     text_color=TEXT_MUTED).pack(anchor="w", padx=12, pady=(8, 2))
        self.asset_type_var = tk.StringVar(value="(All)")
        self.asset_type_menu = ctk.CTkOptionMenu(
            card, variable=self.asset_type_var, values=["(All)"],
            fg_color=PANEL_BG, button_color=ACCENT_BLUE,
            font=ctk.CTkFont(size=11), width=200,
            command=self._apply_filter
        )
        self.asset_type_menu.pack(fill="x", padx=12, pady=(0, 10))

    def _fields_card(self, parent):
        card = self._card(parent)
        # EDC fields
        ctk.CTkLabel(card, text="EDC columns to compare:",
                     font=ctk.CTkFont(size=11), text_color=TEXT_MUTED).pack(anchor="w", padx=12, pady=(8,2))
        self.edc_col_frame = ctk.CTkScrollableFrame(card, fg_color=CARD_BG, height=90,
                                                     scrollbar_button_color=BORDER)
        self.edc_col_frame.pack(fill="x", padx=12, pady=(0,6))
        self.edc_col_checks: dict[str, tk.BooleanVar] = {}

        # Collibra fields
        ctk.CTkLabel(card, text="Collibra columns to compare:",
                     font=ctk.CTkFont(size=11), text_color=TEXT_MUTED).pack(anchor="w", padx=12, pady=(4,2))
        self.col_col_frame = ctk.CTkScrollableFrame(card, fg_color=CARD_BG, height=90,
                                                     scrollbar_button_color=BORDER)
        self.col_col_frame.pack(fill="x", padx=12, pady=(0, 10))
        self.col_col_checks: dict[str, tk.BooleanVar] = {}

    def _parse_card(self, parent):
        card = self._card(parent)
        self.parse_var = tk.BooleanVar(value=True)
        ctk.CTkCheckBox(
            card, text="Parse 'Full Name' column (Azure path)",
            variable=self.parse_var,
            font=ctk.CTkFont(size=11),
            text_color=TEXT_PRIMARY,
            fg_color=ACCENT_TEAL, hover_color=ACCENT_BLUE
        ).pack(anchor="w", padx=12, pady=10)

    def _run_card(self, parent):
        card = self._card(parent)
        self.run_btn = ctk.CTkButton(
            card, text="▶  RUN COMPARISON", height=40,
            fg_color=ACCENT_GREEN, hover_color="#2E8B40",
            font=ctk.CTkFont("Courier New", 13, "bold"),
            text_color="#0D1117",
            command=self._run_comparison
        )
        self.run_btn.pack(fill="x", padx=12, pady=(10, 6))
        self.export_btn = ctk.CTkButton(
            card, text="⬇  EXPORT TO EXCEL", height=36,
            fg_color=PANEL_BG, hover_color=BORDER,
            border_width=1, border_color=ACCENT_GOLD,
            font=ctk.CTkFont(size=12), text_color=ACCENT_GOLD,
            command=self._export,
            state="disabled"
        )
        self.export_btn.pack(fill="x", padx=12, pady=(0, 10))

        self.progress = ctk.CTkProgressBar(card, mode="indeterminate",
                                            fg_color=PANEL_BG, progress_color=ACCENT_BLUE)
        self.progress.pack(fill="x", padx=12, pady=(0, 10))
        self.progress.set(0)

    # ── Right Panel ────────────────────────────────────────────────────────────
    def _build_right_panel(self, parent):
        # Stats bar
        self.stats_frame = ctk.CTkFrame(parent, fg_color=PANEL_BG,
                                         corner_radius=8, height=88)
        self.stats_frame.pack(fill="x", pady=(0, 8))
        self.stats_frame.pack_propagate(False)
        self._stats_labels = {}
        stats = [
            ("EDC Rows",    "edc_rows",    ACCENT_BLUE),
            ("Collibra Rows","col_rows",   ACCENT_TEAL),
            ("✅ Matches",  "matches",     ACCENT_GREEN),
            ("Only EDC",    "only_edc",    ACCENT_RED),
            ("Only Collibra","only_col",   ACCENT_GOLD),
            ("Match Rate",  "match_rate",  TEXT_PRIMARY),
        ]
        for i, (lbl, key, clr) in enumerate(stats):
            f = ctk.CTkFrame(self.stats_frame, fg_color=CARD_BG, corner_radius=6)
            f.grid(row=0, column=i, padx=6, pady=8, sticky="nsew")
            self.stats_frame.grid_columnconfigure(i, weight=1)
            ctk.CTkLabel(f, text=lbl, font=ctk.CTkFont(size=10),
                         text_color=TEXT_MUTED).pack(pady=(6,0))
            val_lbl = ctk.CTkLabel(f, text="—", font=ctk.CTkFont("Courier New", 18, "bold"),
                                    text_color=clr)
            val_lbl.pack(pady=(0, 6))
            self._stats_labels[key] = val_lbl

        # Tabs
        self.tabs = ctk.CTkTabview(parent, fg_color=PANEL_BG,
                                    segmented_button_fg_color=CARD_BG,
                                    segmented_button_selected_color=ACCENT_BLUE,
                                    segmented_button_unselected_color=CARD_BG,
                                    segmented_button_selected_hover_color="#1A6FD6",
                                    text_color=TEXT_PRIMARY)
        self.tabs.pack(fill="both", expand=True)

        for tab_name in ["All Results", "✅ Matches", "❌ Only EDC",
                         "❌ Only Collibra", "DQ Report", "Full Name Parsed"]:
            self.tabs.add(tab_name)

        self._result_trees = {}
        for tab_name in ["All Results", "✅ Matches", "❌ Only EDC", "❌ Only Collibra"]:
            tree = self._make_treeview(self.tabs.tab(tab_name))
            self._result_trees[tab_name] = tree

        self.dq_text = ctk.CTkTextbox(
            self.tabs.tab("DQ Report"),
            fg_color=CARD_BG, text_color=TEXT_PRIMARY,
            font=ctk.CTkFont("Courier New", 12), wrap="word"
        )
        self.dq_text.pack(fill="both", expand=True, padx=4, pady=4)

        self.parse_tree = self._make_treeview(self.tabs.tab("Full Name Parsed"))

    def _make_treeview(self, parent) -> ttk.Treeview:
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Dark.Treeview",
                         background=CARD_BG, foreground=TEXT_PRIMARY,
                         fieldbackground=CARD_BG, bordercolor=BORDER,
                         rowheight=26, font=("Courier New", 10))
        style.configure("Dark.Treeview.Heading",
                         background=PANEL_BG, foreground=ACCENT_BLUE,
                         font=("Courier New", 10, "bold"),
                         relief="flat", borderwidth=0)
        style.map("Dark.Treeview",
                  background=[("selected", ACCENT_BLUE)],
                  foreground=[("selected", "#FFFFFF")])

        frame = ctk.CTkFrame(parent, fg_color=DARK_BG, corner_radius=0)
        frame.pack(fill="both", expand=True, padx=4, pady=4)

        # search bar
        search_var = tk.StringVar()
        search_entry = ctk.CTkEntry(
            frame, placeholder_text="🔍 Search…",
            textvariable=search_var,
            fg_color=CARD_BG, border_color=BORDER,
            text_color=TEXT_PRIMARY, height=30
        )
        search_entry.pack(fill="x", pady=(0, 4))

        vsb = ttk.Scrollbar(frame, orient="vertical")
        hsb = ttk.Scrollbar(frame, orient="horizontal")
        tree = ttk.Treeview(frame, style="Dark.Treeview",
                             yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.config(command=tree.yview)
        hsb.config(command=tree.xview)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)

        def _search(*_):
            q = search_var.get().lower()
            for item in tree.get_children():
                vals = [str(v).lower() for v in tree.item(item, "values")]
                if any(q in v for v in vals) or not q:
                    tree.reattach(item, "", "end")
                else:
                    tree.detach(item)

        search_var.trace_add("write", _search)
        return tree

    # ── File Loading ───────────────────────────────────────────────────────────
    def _load_file(self, key, label):
        path = filedialog.askopenfilename(
            title=f"Select {label}",
            filetypes=[("Excel files","*.xlsx *.xls"),
                       ("CSV files","*.csv"),
                       ("All files","*.*")]
        )
        if not path:
            return
        self._set_status(f"Loading {os.path.basename(path)}…")
        try:
            if path.lower().endswith(".csv"):
                df = pd.read_csv(path)
            else:
                df = pd.read_excel(path, sheet_name=0)
            df.columns = df.columns.str.strip()
        except Exception as e:
            messagebox.showerror("Load Error", str(e))
            return

        if key == "edc":
            self.edc_df   = df
            self.edc_path = path
            self.edc_path_var.set(f"✔ {os.path.basename(path)}  ({len(df):,} rows)")
            self._populate_col_checkboxes("edc", df.columns.tolist())
        else:
            self.col_df   = df
            self.col_path = path
            self.col_path_var.set(f"✔ {os.path.basename(path)}  ({len(df):,} rows)")
            self._populate_col_checkboxes("col", df.columns.tolist())
            self._populate_asset_type_filter(df)

        self._set_status(f"Loaded {os.path.basename(path)}")

    def _populate_col_checkboxes(self, key, columns):
        frame = self.edc_col_frame if key == "edc" else self.col_col_frame
        checks = self.edc_col_checks if key == "edc" else self.col_col_checks
        priority = self.EDC_PRIORITY if key == "edc" else self.COL_PRIORITY

        for w in frame.winfo_children():
            w.destroy()
        checks.clear()

        for col in columns:
            var = tk.BooleanVar(value=(col in priority))
            checks[col] = var
            ctk.CTkCheckBox(
                frame, text=col, variable=var,
                font=ctk.CTkFont(size=11), text_color=TEXT_PRIMARY,
                fg_color=ACCENT_BLUE, hover_color=ACCENT_TEAL,
                checkbox_width=16, checkbox_height=16
            ).pack(anchor="w", pady=1)

    def _populate_asset_type_filter(self, df):
        if "Asset Type" in df.columns:
            vals = ["(All)"] + sorted(df["Asset Type"].dropna().unique().tolist())
            self.asset_type_menu.configure(values=vals)
            self.asset_type_var.set("(All)")

    def _apply_filter(self, _=None):
        if self.col_df is None:
            return
        val = self.asset_type_var.get()
        if val == "(All)":
            filtered = self.col_df
        else:
            filtered = self.col_df[self.col_df.get("Asset Type", pd.Series()) == val]
        self.col_path_var.set(
            f"✔ {os.path.basename(self.col_path)}  ({len(filtered):,} rows  [filtered])"
        )
        self._set_status(f"Filter applied: Asset Type = '{val}'  ({len(filtered):,} rows)")

    # ── Run Comparison ─────────────────────────────────────────────────────────
    def _run_comparison(self):
        if self.edc_df is None or self.col_df is None:
            messagebox.showwarning("Missing Files", "Please load both files first.")
            return

        edc_fields = [c for c, v in self.edc_col_checks.items() if v.get()]
        col_fields  = [c for c, v in self.col_col_checks.items() if v.get()]

        if not edc_fields or not col_fields:
            messagebox.showwarning("No Fields", "Select at least one field to compare from each file.")
            return

        if len(edc_fields) != len(col_fields):
            messagebox.showwarning(
                "Field Mismatch",
                f"You selected {len(edc_fields)} EDC field(s) and {len(col_fields)} Collibra field(s).\n"
                "Please select the same number of fields for a paired comparison."
            )
            return

        val = self.asset_type_var.get()
        col_df = self.col_df.copy()
        if val != "(All)" and "Asset Type" in col_df.columns:
            col_df = col_df[col_df["Asset Type"] == val].reset_index(drop=True)

        self.run_btn.configure(state="disabled")
        self.progress.configure(mode="indeterminate")
        self.progress.start()
        self._set_status("Running comparison…")

        def _worker():
            try:
                results = run_comparison(
                    self.edc_df.copy(), col_df,
                    edc_fields, col_fields
                )
                self.results = results
                self.after(0, lambda: self._display_results(results, edc_fields, col_fields))
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("Error", str(e)))
            finally:
                self.after(0, self._stop_progress)

        threading.Thread(target=_worker, daemon=True).start()

    def _stop_progress(self):
        self.progress.stop()
        self.progress.set(0)
        self.run_btn.configure(state="normal")
        self.export_btn.configure(state="normal")

    def _display_results(self, results: dict, edc_fields, col_fields):
        self._set_status("Comparison complete ✓")

        # Stats
        self._stats_labels["edc_rows"].configure(text=f'{results["edc_total"]:,}')
        self._stats_labels["col_rows"].configure(text=f'{results["col_total"]:,}')
        self._stats_labels["matches"].configure(text=f'{results["match_count"]:,}')
        self._stats_labels["only_edc"].configure(text=f'{results["only_edc_count"]:,}')
        self._stats_labels["only_col"].configure(text=f'{results["only_col_count"]:,}')
        rate = results["match_count"] / max(results["edc_total"], 1) * 100
        self._stats_labels["match_rate"].configure(text=f'{rate:.1f}%')

        edc_df = results["edc_detail"]
        col_df = results["col_detail"]

        # Helper: fill a treeview
        def fill_tree(tree: ttk.Treeview, df: pd.DataFrame):
            tree.delete(*tree.get_children())
            disp_cols = [c for c in df.columns if c not in ("__key__",)]
            tree["columns"] = disp_cols
            tree["show"] = "headings"
            for col in disp_cols:
                tree.heading(col, text=col)
                tree.column(col, width=120, minwidth=60)
            tree.tag_configure("match",    background="#1A3A2A", foreground="#3FB950")
            tree.tag_configure("only_edc", background="#3A1A1A", foreground="#F85149")
            tree.tag_configure("only_col", background="#2A2A1A", foreground="#D29922")
            tree.tag_configure("alt",      background="#1C2128")

            for i, (_, row) in enumerate(df.iterrows()):
                status = str(row.get("__status__", ""))
                tag = "match" if "Match" in status else \
                      "only_edc" if "EDC" in status else \
                      "only_col" if "Collibra" in status else \
                      ("alt" if i % 2 else "")
                vals = [row[c] for c in disp_cols]
                tree.insert("", "end", values=vals, tags=(tag,))

        fill_tree(self._result_trees["All Results"],     edc_df)
        fill_tree(self._result_trees["✅ Matches"],
                  edc_df[edc_df["__key__"].isin(results["matches"])])
        fill_tree(self._result_trees["❌ Only EDC"],
                  edc_df[edc_df["__key__"].isin(results["only_in_edc"])])
        fill_tree(self._result_trees["❌ Only Collibra"],
                  col_df[col_df["__key__"].isin(results["only_in_col"])])

        # DQ report
        self._build_dq_report(results, edc_fields, col_fields)

        # Full Name parse tab
        if self.parse_var.get() and "Full Name" in col_df.columns:
            self._build_parse_tab(col_df)
        else:
            self.parse_tree.delete(*self.parse_tree.get_children())

    def _build_dq_report(self, results, edc_fields, col_fields):
        edc_df = results["edc_detail"]
        col_df = results["col_detail"]
        self.dq_text.configure(state="normal")
        self.dq_text.delete("0.0", "end")
        lines = [
            "═" * 60,
            "  DATA QUALITY & STATISTICS REPORT",
            f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "═" * 60,
            "",
            "── FILE OVERVIEW ────────────────────────────────────────",
            f"  EDC File         : {os.path.basename(self.edc_path)}",
            f"  Collibra File    : {os.path.basename(self.col_path)}",
            f"  EDC File Size    : {os.path.getsize(self.edc_path)/1024:.1f} KB" if os.path.exists(self.edc_path) else "",
            f"  Collibra Size    : {os.path.getsize(self.col_path)/1024:.1f} KB" if os.path.exists(self.col_path) else "",
            "",
            "── ROW / COLUMN COUNTS ──────────────────────────────────",
            f"  EDC Total Rows   : {results['edc_total']:,}",
            f"  Collibra Rows    : {results['col_total']:,}",
            f"  EDC Columns      : {len(edc_df.columns) - 2}",
            f"  Collibra Columns : {len(col_df.columns) - 2}",
            "",
            "── COMPARISON RESULTS ───────────────────────────────────",
            f"  Fields compared  : EDC {edc_fields} ↔ Collibra {col_fields}",
            f"  ✅ Matches        : {results['match_count']:,}",
            f"  ❌ Only in EDC   : {results['only_edc_count']:,}",
            f"  ❌ Only Collibra : {results['only_col_count']:,}",
            f"  Match Rate       : {results['match_count']/max(results['edc_total'],1)*100:.2f}%",
            "",
            "── DATA QUALITY CHECKS ──────────────────────────────────",
        ]
        for name, df in [("EDC", edc_df), ("Collibra", col_df)]:
            real_cols = [c for c in df.columns if c not in ("__key__","__status__")]
            sub = df[real_cols]
            nulls = int(sub.isnull().sum().sum())
            dupes = int(df.duplicated(subset=real_cols).sum())
            lines += [
                f"  {name}:",
                f"    Null/Empty cells  : {nulls:,}",
                f"    Duplicate rows    : {dupes:,}",
            ]
            for col in real_cols[:8]:
                n = int(sub[col].isnull().sum())
                if n:
                    lines.append(f"    '{col}' nulls      : {n:,}")
            lines.append("")

        lines += [
            "── ASSET TYPE DISTRIBUTION (Collibra) ───────────────────",
        ]
        if "Asset Type" in col_df.columns:
            counts = col_df["Asset Type"].value_counts()
            for at, cnt in counts.items():
                lines.append(f"    {str(at):<40} {cnt:>6,}")
        lines.append("")

        if "Full Name" in col_df.columns and self.parse_var.get():
            parsed = col_df["Full Name"].dropna().apply(parse_full_name)
            parsed_df = pd.DataFrame(list(parsed))
            lines += [
                "── FULL NAME PARSE STATS ─────────────────────────────────",
                f"  Total parsed       : {len(parsed_df):,}",
                f"  Unique Azure Srvr  : {parsed_df['azure_server'].nunique():,}",
                f"  Unique Data Stores : {parsed_df['data_store'].nunique():,}",
                f"  Unique Mal Codes   : {parsed_df['mal_code'].nunique():,}",
                f"  Unique Tables      : {parsed_df['table_name'].nunique():,}",
                f"  Unique Columns     : {parsed_df['column_name'].nunique():,}",
            ]

        self.dq_text.insert("0.0", "\n".join(lines))
        self.dq_text.configure(state="disabled")

    def _build_parse_tab(self, col_df: pd.DataFrame):
        tree = self.parse_tree
        tree.delete(*tree.get_children())
        parse_cols = ["Full Name", "azure_server", "data_store",
                      "mal_code", "table_name", "column_name"]
        tree["columns"] = parse_cols
        tree["show"] = "headings"
        for c in parse_cols:
            tree.heading(c, text=c)
            tree.column(c, width=160, minwidth=80)

        for _, row in col_df.iterrows():
            fn = row.get("Full Name", "")
            parsed = parse_full_name(fn)
            vals = [fn, parsed["azure_server"], parsed["data_store"],
                    parsed["mal_code"], parsed["table_name"], parsed["column_name"]]
            tree.insert("", "end", values=vals)

    # ── Export ─────────────────────────────────────────────────────────────────
    def _export(self):
        if not self.results:
            return
        out = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")],
            initialfile=f"comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not out:
            return
        self._set_status("Exporting…")
        self.progress.configure(mode="indeterminate")
        self.progress.start()

        def _worker():
            try:
                export_to_excel(self.results, self.edc_path, self.col_path, out)
                self.after(0, lambda: messagebox.showinfo(
                    "Export Complete", f"Report saved to:\n{out}"))
                self.after(0, lambda: self._set_status(f"Exported → {os.path.basename(out)}"))
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("Export Error", str(e)))
            finally:
                self.after(0, self._stop_progress)

        threading.Thread(target=_worker, daemon=True).start()

    def _set_status(self, msg):
        self.status_label.configure(text=msg)
        self.update_idletasks()


# ── Entry ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = DataComparatorApp()
    app.mainloop()
