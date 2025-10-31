import pandas as pd
import openpyxl
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import sys
import os

# --- Configuration Constants ---
DQC_EVAL_TAB = 'DQ Controls Evaluation Sheet'
DQC_SCOPING_TAB = 'DQ Controls Scoping'
EVAL_HEADER_COLUMNS = ['MAL/ EUC Code', 'Database Name', 'DQ Evaluation Unique Identifier in scope of DQ Control Evaluation(MAL/EUC.Database Name.Schema.Table Name.Column Name)', 'DMC Requirement Mapping']
SCOPING_START_ROW = 10 # Starting row for data in Scoping tab (as per Note 3)
SCOPING_COLUMNS_FORMULA_CHECK = ['AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV']
COMPONENTS = ['MALCODE', 'Database Name', 'Schema', 'Table Name', 'Column Name']
SCOPING_COMPARE_COLS = ['Z', 'AA', 'AB', 'AC', 'AD'] # System, Database, Schema, Table, Column

class DQ_Analysis_Tool:
    """Core logic for performing DQC checks on Excel workbooks."""
    def __init__(self):
        self.file_path = None
        self.log_results_df = pd.DataFrame() # DataFrame for the main log
        self.scoping_formula_results_df = pd.DataFrame() # DataFrame for the formula check
        self.wb = None # openpyxl workbook object

    # --- UTILITY FUNCTIONS ---
    def _find_data_start_row(self, ws, target_headers):
        """Finds the header row and returns the index of the next row (data start)."""
        # Note 2: Robust header search logic
        for i, row in enumerate(ws.iter_rows(), 1):
            row_values = [str(cell.value).strip().lower() for cell in row if cell.value is not None]
            
            # Check if all target headers are present in the current row (case-insensitive)
            # This is a simplified check assuming the headers are *in* the row, not necessarily in the first few columns
            header_match = 0
            for header in target_headers:
                if header.lower() in row_values:
                    header_match += 1
            
            if header_match == len(target_headers):
                return i + 1 # Return the row *after* the header row
        return None # Headers not found

    # --- CORE ANALYSIS FUNCTIONS ---
    def load_file(self):
        """Opens a file dialog, checks naming, and loads the workbook."""
        self.file_path = None
        self.wb = None
        
        path = filedialog.askopenfilename(
            title="Select EDMO Data Quality Control Evaluation Workbook",
            filetypes=[("Excel files", "*.xlsx *.xlsm")]
        )
        
        if not path:
            return "File Load Canceled.", 'info'

        # Requirement 1: Check Naming Convention
        filename = os.path.basename(path)
        if not re.match(r"EDMO Data Quality Control Evaluation Workbook.*", filename):
            return f"FAIL: File name '{filename}' does not match convention.", 'error'

        self.file_path = path
        
        try:
            # Load with openpyxl for formula checks
            self.wb = openpyxl.load_workbook(self.file_path, data_only=False)
            
            # Requirement 2: Check for required tabs
            if DQC_EVAL_TAB not in self.wb.sheetnames or DQC_SCOPING_TAB not in self.wb.sheetnames:
                 return f"FAIL: Workbook must contain tabs '{DQC_EVAL_TAB}' and '{DQC_SCOPING_TAB}'.", 'error'
            
            return f"SUCCESS: File loaded: {filename}", 'success'

        except Exception as e:
            self.wb = None
            return f"ERROR: Could not load file. Details: {e}", 'error'

    def run_analysis(self):
        """Runs all DQC checks and populates result DataFrames."""
        if not self.wb:
            return "ERROR: No file loaded to analyze. Load a file first.", 'error'

        ws_eval = self.wb[DQC_EVAL_TAB]
        ws_scoping = self.wb[DQC_SCOPING_TAB]
        
        log_results = []
        
        # 4. Get Scoping Data (Columns Z to AD) for comparison
        scoping_comparison_data = {} # {row_num: [System, Database, ...]}
        for i in range(SCOPING_START_ROW, ws_scoping.max_row + 1):
            scoping_comparison_data[i] = [
                str(ws_scoping[f'{col}{i}'].value or '').strip().lower()
                for col in SCOPING_COMPARE_COLS
            ]
        
        # Find Evaluation Sheet Data Start (Note 2)
        data_start_row = self._find_data_start_row(ws_eval, EVAL_HEADER_COLUMNS)
        if data_start_row is None:
            return f"FAIL: Could not locate header row in '{DQC_EVAL_TAB}'. Analysis aborted.", 'error'
        
        # --- Evaluate DQ Controls Evaluation Sheet (Requirements 3 & 4) ---
        for i in range(data_start_row, ws_eval.max_row + 1):
            cell_D = ws_eval[f'D{i}']
            full_string = str(cell_D.value or '').strip()
            
            # 3i. Formula Check
            formula_check_status = "PASS" if cell_D.data_type == 'f' else "FAIL"

            # 3ii. & 4. Component Check & Tracing
            components = full_string.split('.')
            components_check_status = "FAIL"
            row_log = {
                'Row': i,
                'DQ Unique ID': full_string,
                'Formula Check': formula_check_status,
                'Component Check': 'N/A' # Default, updated below
            }
            
            component_match_results = {} # For detailed component logging

            if len(components) == 5:
                all_match = True
                scoping_row_index = i # ASSUMPTION: Evaluation row 'i' corresponds to Scoping row 'i'
                
                if scoping_row_index in scoping_comparison_data:
                    scoping_comps = scoping_comparison_data[scoping_row_index]
                    
                    for k in range(5):
                        eval_comp = components[k].strip().lower()
                        scoping_comp = scoping_comps[k]
                        
                        # Note 1: Ignore Case
                        match_status = "PASS" if eval_comp == scoping_comp else "FAIL"
                        component_match_results[f'{COMPONENTS[k]} Match'] = match_status
                        
                        if match_status == "FAIL":
                            all_match = False
                            
                    components_check_status = "PASS" if all_match else "FAIL"
                else:
                    components_check_status = "FAIL (No Scoping Data at Row)"
            else:
                component_match_results['Component_Length_Error'] = f"FAIL (Expected 5, Found {len(components)})"
                components_check_status = "FAIL (Component Count)"
            
            row_log['Component Check'] = components_check_status
            log_results.append({**row_log, **component_match_results})
            
        self.log_results_df = pd.DataFrame(log_results)
        self.log_results_df['Total Status'] = self.log_results_df.apply(
            lambda x: 'PASS' if x['Formula Check'] == 'PASS' and x['Component Check'] == 'PASS' else 'FAIL', axis=1
        )
        
        # --- Evaluate Scoping Tab Formulae (Requirement 5) ---
        scoping_formula_results = []
        for i in range(SCOPING_START_ROW, ws_scoping.max_row + 1):
            row_result = {'Row': i}
            all_pass = True
            
            for col_letter in SCOPING_COLUMNS_FORMULA_CHECK:
                cell = ws_scoping[f'{col_letter}{i}']
                
                # Check if the cell contains a formula
                is_formula = cell.data_type == 'f'
                check_status = "PASS" if is_formula else "FAIL"
                
                if not is_formula:
                    all_pass = False
                
                row_result[f'Col {col_letter} Formula Check'] = check_status
                
            row_result['Formula Check - Pass/Fail'] = "PASS" if all_pass else "FAIL"
            scoping_formula_results.append(row_result)
        
        self.scoping_formula_results_df = pd.DataFrame(scoping_formula_results)

        return "SUCCESS: All analysis checks completed.", 'success'


class DQC_App_UI:
    """The Tkinter GUI interface for the DQ_Analysis_Tool."""
    def __init__(self, master):
        self.master = master
        master.title("EDMO Data Quality Control Analyst Tool 📊")
        master.geometry("1200x800")
        
        self.tool = DQ_Analysis_Tool()
        self.style = ttk.Style()
        self.style.theme_use('clam')
        
        # --- Variables ---
        self.file_status_var = tk.StringVar(value="Status: Ready to load file...")
        self.file_path_var = tk.StringVar(value="No file selected.")

        # --- Setup Layout ---
        self._create_widgets()
        self.master.grid_columnconfigure(0, weight=1)
        self.master.grid_rowconfigure(1, weight=1)

    def _create_widgets(self):
        # 1. Control Panel Frame (Top)
        control_frame = ttk.Frame(self.master, padding="10")
        control_frame.grid(row=0, column=0, sticky="ew")
        control_frame.columnconfigure(1, weight=1)

        # File Label
        ttk.Label(control_frame, text="Selected File:", font=('Arial', 10, 'bold')).grid(row=0, column=0, padx=5, pady=5, sticky='w')
        
        # File Path Display
        ttk.Label(control_frame, textvariable=self.file_path_var, wraplength=700).grid(row=0, column=1, padx=5, pady=5, sticky='ew')

        # Load Button
        self.load_button = ttk.Button(control_frame, text="📂 Load Workbook", command=self.load_workbook_action)
        self.load_button.grid(row=0, column=2, padx=10, pady=5)
        
        # Analysis Button
        self.analyze_button = ttk.Button(control_frame, text="▶️ Start Analysis", command=self.run_analysis_action, state=tk.DISABLED)
        self.analyze_button.grid(row=0, column=3, padx=10, pady=5)
        
        # Status Bar
        self.status_label = ttk.Label(control_frame, textvariable=self.file_status_var, relief=tk.SUNKEN, anchor='w')
        self.status_label.grid(row=1, column=0, columnspan=4, sticky='ew', pady=5)

        # 2. Results Viewer (Main Content Area)
        self.notebook = ttk.Notebook(self.master)
        self.notebook.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)

        # Tabs
        self.log_tab = ttk.Frame(self.notebook)
        self.scoping_tab = ttk.Frame(self.notebook)
        
        self.notebook.add(self.log_tab, text="✅ Analysis Summary Log")
        self.notebook.add(self.scoping_tab, text="📝 Scoping Formula Check")
        
        self.log_tree = self._setup_treeview(self.log_tab, ['Row', 'DQ Unique ID', 'Formula Check', 'Component Check', 'Total Status'])
        self.scoping_tree = self._setup_treeview(self.scoping_tab, ['Row', 'Formula Check - Pass/Fail'])
        
    def _setup_treeview(self, parent_frame, columns):
        """Helper to create and configure a Treeview for tabular data."""
        parent_frame.grid_columnconfigure(0, weight=1)
        parent_frame.grid_rowconfigure(0, weight=1)
        
        tree = ttk.Treeview(parent_frame, columns=columns, show='headings')
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, anchor='center', width=100)
        
        # Scrollbars
        v_scroll = ttk.Scrollbar(parent_frame, orient="vertical", command=tree.yview)
        h_scroll = ttk.Scrollbar(parent_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
        
        tree.grid(row=0, column=0, sticky='nsew')
        v_scroll.grid(row=0, column=1, sticky='ns')
        h_scroll.grid(row=1, column=0, sticky='ew')
        
        return tree

    def _update_treeview(self, tree, df):
        """Clears and populates a Treeview with DataFrame content."""
        tree.delete(*tree.get_children())
        if df.empty:
            return
            
        # Dynamically set columns if the DataFrame changes (e.g., adding component details)
        new_cols = list(df.columns)
        tree['columns'] = new_cols
        
        # Update headings and column widths
        for col in new_cols:
            tree.heading(col, text=col)
            tree.column(col, anchor='center', width=100)
            
        # Insert data
        for index, row in df.iterrows():
            tree.insert('', 'end', values=list(row), tags=('pass' if row['Total Status'] == 'PASS' or row['Formula Check - Pass/Fail'] == 'PASS' else 'fail'))
            
        # Apply row colors for clarity
        tree.tag_configure('pass', background='#e0ffe0', foreground='black') # Light green for PASS
        tree.tag_configure('fail', background='#ffdddd', foreground='red') # Light red for FAIL
        
        # Auto-adjust column widths (a basic estimate)
        for col in new_cols:
            max_width = max(len(str(df[col].max())) * 8 + 10, len(col) * 10 + 10)
            tree.column(col, width=max_width, anchor='center')

    # --- ACTION HANDLERS ---
    def load_workbook_action(self):
        """Handles the 'Load Workbook' button click."""
        self.file_status_var.set("Status: Loading file...")
        status_msg, status_type = self.tool.load_file()
        
        self.file_status_var.set(f"Status: {status_msg}")
        
        if status_type == 'success':
            self.file_path_var.set(self.tool.file_path)
            self.analyze_button.config(state=tk.NORMAL)
            messagebox.showinfo("File Loaded", status_msg)
        else:
            self.file_path_var.set("No file selected.")
            self.analyze_button.config(state=tk.DISABLED)
            messagebox.showerror("Error", status_msg)

    def run_analysis_action(self):
        """Handles the 'Start Analysis' button click."""
        self.analyze_button.config(state=tk.DISABLED)
        self.file_status_var.set("Status: Running analysis...")
        self.master.update() # Update UI to show status change
        
        try:
            status_msg, status_type = self.tool.run_analysis()
            
            # 1. Update Analysis Summary Log Tab
            # Select only the key columns for the summary log
            summary_cols = ['Row', 'DQ Unique ID', 'Formula Check', 'Component Check', 'Total Status']
            log_summary_df = self.tool.log_results_df[summary_cols]
            self._update_treeview(self.log_tree, log_summary_df)

            # 2. Update Scoping Formula Check Tab
            self._update_treeview(self.scoping_tree, self.tool.scoping_formula_results_df)

            self.file_status_var.set(f"Status: {status_msg} Checks complete. Results displayed.")
            if status_type == 'success':
                messagebox.showinfo("Analysis Complete", status_msg)
            else:
                messagebox.showerror("Analysis Error", status_msg)
                
        except Exception as e:
            error_msg = f"An unexpected error occurred during analysis: {e}"
            self.file_status_var.set(f"Status: ERROR. Check log.")
            messagebox.showerror("Critical Error", error_msg)

        finally:
            self.analyze_button.config(state=tk.NORMAL)


if __name__ == "__main__":
    # Ensure a basic environment for Tkinter is available
    root = tk.Tk()
    app = DQC_App_UI(root)
    root.mainloop()
