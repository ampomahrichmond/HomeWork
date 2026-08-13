"""
DSCIT Tier-1 Report Consolidator
---------------------------------
Scans a folder for DSCIT_* Excel files, locates the DSCIT tab (any version,
e.g. "DSCIT", "DSCIT v1", "DSCIT_v2"), auto-detects the header row (the row
whose first populated cell is "Object ID"), fuzzy-matches the 14 target
columns regardless of where they sit in each file, and consolidates
everything into a single output workbook with the source file name as the
first column.

Run:  python dscit_consolidator.py
Requires:  pip install openpyxl pandas
"""

import os
import re
import sys
import glob
import queue
import threading
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

# ----------------------------------------------------------------------------
# Extraction engine
# ----------------------------------------------------------------------------

TARGET_COLUMNS = [
    "Output Name",
    "Output Owner",
    "Data Element",
    "Data Element Description",
    "Source Name",
    "Source Type",
    "Source Application EUC / MAL Code",
    "Database Type",
    "Database Name",
    "Database",
    "Business Segment / Corporate Function",
    "Asset Owner Name",
    "Technology Asset Owner Name",
    "Data Owner Name",
]

# Aliases: normalized (lowercase, alphanumeric-only) variants seen in the wild.
COLUMN_ALIASES = {
    "Output Name": ["outputname", "reportname", "outputreportname"],
    "Output Owner": ["outputowner", "reportowner"],
    "Data Element": ["dataelement", "dataelementname", "criticaldataelement", "cde"],
    "Data Element Description": [
        "dataelementdescription", "dataelementdiscription",  # common typo
        "dataelementdesc", "elementdescription",
    ],
    "Source Name": ["sourcename", "sourcesystemname", "datasourcename"],
    "Source Type": ["sourcetype", "sourcesystemtype"],
    "Source Application EUC / MAL Code": [
        "sourceapplicationeucmalcode", "sourceapplicationeucormalcode",
        "applicationeucmalcode", "eucmalcode", "sourceappeucmalcode",
        "sourceapplicationmalcode", "sourceapplicationeuccode", "malcode",
    ],
    "Database Type": ["databasetype", "dbtype"],
    "Database Name": ["databasename", "dbname"],
    "Database": ["database", "db"],
    "Business Segment / Corporate Function": [
        "businesssegmentcorporatefunction", "businesssegmentorcorporatefunction",
        "businesssegment", "corporatefunction", "segmentfunction",
    ],
    "Asset Owner Name": ["assetownername", "assetowner", "dataassetownername"],
    "Technology Asset Owner Name": [
        "technologyassetownername", "technologyassetowner",
        "techassetownername", "techassetowner", "taoname",
    ],
    "Data Owner Name": ["dataownername", "dataowner"],
}

HEADER_SCAN_ROWS = 40          # how deep to look for the "Object ID" header row
HEADER_SCAN_COLS = 5           # "Object ID" should be in one of the first few cols
SHEET_KEY = "dscit"


def _norm(value) -> str:
    """Normalize a cell/header for matching: lowercase alphanumerics only."""
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(value).lower())


def find_dscit_sheet(sheetnames):
    """Return best-matching sheet name containing 'DSCIT' (exact name wins)."""
    candidates = [s for s in sheetnames if SHEET_KEY in _norm(s)]
    if not candidates:
        return None
    # Prefer exact "DSCIT", then shortest (e.g. "DSCIT v2" over "DSCIT backup old")
    candidates.sort(key=lambda s: (_norm(s) != SHEET_KEY, len(s)))
    return candidates[0]


def find_header_row(rows):
    """rows: list of row tuples. Return (index, row) whose first populated
    cells contain 'Object ID'; else (None, None)."""
    for idx, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        for cell in row[:HEADER_SCAN_COLS]:
            n = _norm(cell)
            if n == "objectid":
                return idx, row
            if n:      # first populated cell wasn't Object ID -> next row
                break
    return None, None


def map_columns(header_row):
    """Map each target column -> column index in this sheet.
    Two passes: exact alias match, then contains-match for leftovers.
    Longer / more specific targets are resolved first so 'Technology Asset
    Owner Name' never gets stolen by 'Asset Owner Name'."""
    headers = {i: _norm(h) for i, h in enumerate(header_row) if _norm(h)}
    mapping, claimed = {}, set()

    order = sorted(TARGET_COLUMNS, key=lambda t: -len(_norm(t)))

    # Pass 1: exact alias match
    for target in order:
        aliases = set(COLUMN_ALIASES[target]) | {_norm(target)}
        for idx, h in headers.items():
            if idx not in claimed and h in aliases:
                mapping[target] = idx
                claimed.add(idx)
                break

    # Pass 2: header contains alias (handles suffixes like "sourcename1",
    # "dataelementdescriptionifapplicable")
    for target in order:
        if target in mapping:
            continue
        aliases = sorted(set(COLUMN_ALIASES[target]) | {_norm(target)},
                         key=len, reverse=True)
        for alias in aliases:
            if len(alias) < 6:          # too short to contains-match safely
                continue
            hit = None
            for idx, h in headers.items():
                if idx not in claimed and alias in h:
                    hit = idx
                    break
            if hit is not None:
                mapping[target] = hit
                claimed.add(hit)
                break
    return mapping


@dataclass
class FileResult:
    file: str
    status: str = "OK"            # OK / WARN / ERROR
    sheet: str = ""
    header_row: int | None = None
    rows: int = 0
    matched: int = 0
    missing: list = field(default_factory=list)
    message: str = ""
    data: pd.DataFrame | None = None


def process_file(path: str) -> FileResult:
    fname = os.path.basename(path)
    res = FileResult(file=fname)
    try:
        if fname.lower().endswith(".xls"):
            return _process_via_pandas(path, res)

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet_name = find_dscit_sheet(wb.sheetnames)
            if not sheet_name:
                res.status, res.message = "ERROR", "No DSCIT sheet found"
                return res
            res.sheet = sheet_name
            ws = wb[sheet_name]
            rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
        return _extract(rows, res)
    except Exception as e:
        res.status = "ERROR"
        res.message = f"{type(e).__name__}: {e}"
        return res


def _process_via_pandas(path, res: FileResult) -> FileResult:
    """Legacy .xls fallback (needs xlrd installed)."""
    xls = pd.ExcelFile(path)
    sheet_name = find_dscit_sheet(xls.sheet_names)
    if not sheet_name:
        res.status, res.message = "ERROR", "No DSCIT sheet found"
        return res
    res.sheet = sheet_name
    df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
    rows = [tuple(r) for r in df.itertuples(index=False, name=None)]
    return _extract(rows, res)


def _extract(rows, res: FileResult) -> FileResult:
    hdr_idx, header = find_header_row(rows)
    if hdr_idx is None:
        res.status, res.message = "ERROR", "Header row with 'Object ID' not found"
        return res
    res.header_row = hdr_idx + 1  # 1-based for display

    mapping = map_columns(header)
    res.matched = len(mapping)
    res.missing = [t for t in TARGET_COLUMNS if t not in mapping]
    if not mapping:
        res.status, res.message = "ERROR", "No target columns matched"
        return res
    if res.missing:
        res.status = "WARN"
        res.message = f"{len(res.missing)} column(s) not found"

    records = []
    for row in rows[hdr_idx + 1:]:
        rec = {}
        blank = True
        for target in TARGET_COLUMNS:
            idx = mapping.get(target)
            val = row[idx] if (idx is not None and idx < len(row)) else None
            if val is not None and str(val).strip() != "":
                blank = False
                rec[target] = str(val).strip() if isinstance(val, str) else val
            else:
                rec[target] = None
        if not blank:
            records.append(rec)

    df = pd.DataFrame(records, columns=TARGET_COLUMNS)
    df.insert(0, "DSCIT File Name", res.file)
    res.rows = len(df)
    res.data = df
    return res


def discover_files(folder: str):
    pats = ["DSCIT*.xlsx", "DSCIT*.xlsm", "DSCIT*.xls",
            "dscit*.xlsx", "dscit*.xlsm", "dscit*.xls"]
    files = set()
    for p in pats:
        files.update(glob.glob(os.path.join(folder, p)))
    return sorted(f for f in files if not os.path.basename(f).startswith("~$"))


def run_consolidation(folder, progress_cb=None, workers=8):
    """Process all files in parallel. Returns (results, combined_df)."""
    files = discover_files(folder)
    results = []
    with ThreadPoolExecutor(max_workers=min(workers, max(1, len(files)))) as ex:
        futures = {ex.submit(process_file, f): f for f in files}
        done = 0
        for fut in as_completed(futures):
            results.append(fut.result())
            done += 1
            if progress_cb:
                progress_cb(done, len(files), results[-1])
    results.sort(key=lambda r: r.file.lower())
    frames = [r.data for r in results if r.data is not None and not r.data.empty]
    combined = pd.concat(frames, ignore_index=True) if frames else \
        pd.DataFrame(columns=["DSCIT File Name"] + TARGET_COLUMNS)
    return results, combined


def write_output(folder, results, combined) -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(folder, f"DSCIT_Consolidated_{stamp}.xlsx")
    summary = pd.DataFrame([{
        "File": r.file, "Status": r.status, "Sheet": r.sheet,
        "Header Row": r.header_row, "Rows Extracted": r.rows,
        "Columns Matched": f"{r.matched}/{len(TARGET_COLUMNS)}",
        "Missing Columns": "; ".join(r.missing), "Notes": r.message,
    } for r in results])
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        combined.to_excel(xw, sheet_name="Consolidated Data", index=False)
        summary.to_excel(xw, sheet_name="Run Summary", index=False)
        for ws in xw.book.worksheets:   # simple column autofit
            for col in ws.columns:
                width = max((len(str(c.value)) for c in col[:200] if c.value), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(width + 2, 45)
    return out_path


# ----------------------------------------------------------------------------
# TD-themed Tkinter UI
# ----------------------------------------------------------------------------

TD_GREEN = "#008A00"
TD_GREEN_DARK = "#006A00"
TD_GREEN_LIGHT = "#54B848"
TD_MIST = "#F3F7F3"
CHARCOAL = "#2E2E2E"
GREY = "#6B7280"
WHITE = "#FFFFFF"
AMBER = "#B45309"
RED = "#B91C1C"


def launch_ui():
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox

    root = tk.Tk()
    root.title("TD  |  DSCIT Tier-1 Report Consolidator")
    root.geometry("1120x760")
    root.configure(bg=TD_MIST)
    root.minsize(960, 640)

    ui_queue = queue.Queue()
    state = {"running": False, "folder": tk.StringVar(value=""), "out": None}

    # ---------- styles ----------
    style = ttk.Style(root)
    style.theme_use("clam")
    style.configure("TFrame", background=TD_MIST)
    style.configure("Card.TFrame", background=WHITE)
    style.configure("Header.TFrame", background=TD_GREEN)
    style.configure("TLabel", background=TD_MIST, foreground=CHARCOAL,
                    font=("Segoe UI", 10))
    style.configure("Card.TLabel", background=WHITE, foreground=CHARCOAL,
                    font=("Segoe UI", 10))
    style.configure("StatValue.TLabel", background=WHITE, foreground=TD_GREEN,
                    font=("Segoe UI Semibold", 22))
    style.configure("StatCap.TLabel", background=WHITE, foreground=GREY,
                    font=("Segoe UI", 9))
    style.configure("HeaderTitle.TLabel", background=TD_GREEN, foreground=WHITE,
                    font=("Segoe UI Semibold", 17))
    style.configure("HeaderSub.TLabel", background=TD_GREEN, foreground="#D8F0D8",
                    font=("Segoe UI", 10))
    style.configure("TD.TButton", background=TD_GREEN, foreground=WHITE,
                    font=("Segoe UI Semibold", 10), padding=(16, 8), borderwidth=0)
    style.map("TD.TButton",
              background=[("active", TD_GREEN_DARK), ("disabled", "#9CC79C")])
    style.configure("Ghost.TButton", background=WHITE, foreground=TD_GREEN,
                    font=("Segoe UI", 10), padding=(12, 7), borderwidth=1)
    style.map("Ghost.TButton", background=[("active", TD_MIST)])
    style.configure("TD.Horizontal.TProgressbar", troughcolor="#E3EAE3",
                    background=TD_GREEN_LIGHT, thickness=14, borderwidth=0)
    style.configure("Treeview", font=("Segoe UI", 9), rowheight=26,
                    background=WHITE, fieldbackground=WHITE, foreground=CHARCOAL)
    style.configure("Treeview.Heading", font=("Segoe UI Semibold", 9),
                    background=TD_MIST, foreground=CHARCOAL)

    # ---------- header ----------
    header = ttk.Frame(root, style="Header.TFrame")
    header.pack(fill="x")
    hin = ttk.Frame(header, style="Header.TFrame")
    hin.pack(fill="x", padx=24, pady=16)
    logo = tk.Canvas(hin, width=46, height=46, bg=TD_GREEN, highlightthickness=0)
    logo.create_rectangle(2, 2, 44, 44, fill=WHITE, outline=WHITE)
    logo.create_text(23, 23, text="TD", fill=TD_GREEN,
                     font=("Segoe UI Black", 16, "bold"))
    logo.pack(side="left", padx=(0, 14))
    tbox = ttk.Frame(hin, style="Header.TFrame")
    tbox.pack(side="left")
    ttk.Label(tbox, text="DSCIT Tier-1 Report Consolidator",
              style="HeaderTitle.TLabel").pack(anchor="w")
    ttk.Label(tbox, text="Enterprise Data Management Office  ·  Tier-1 Sources",
              style="HeaderSub.TLabel").pack(anchor="w")

    # ---------- folder card ----------
    body = ttk.Frame(root)
    body.pack(fill="both", expand=True, padx=24, pady=18)

    fcard = tk.Frame(body, bg=WHITE, highlightbackground="#E2E8E2",
                     highlightthickness=1)
    fcard.pack(fill="x")
    frow = tk.Frame(fcard, bg=WHITE)
    frow.pack(fill="x", padx=18, pady=14)
    tk.Label(frow, text="Report Folder", bg=WHITE, fg=GREY,
             font=("Segoe UI Semibold", 9)).pack(anchor="w")
    inner = tk.Frame(frow, bg=WHITE)
    inner.pack(fill="x", pady=(6, 0))
    entry = tk.Entry(inner, textvariable=state["folder"], font=("Segoe UI", 10),
                     bg=TD_MIST, fg=CHARCOAL, relief="flat",
                     highlightthickness=1, highlightbackground="#D5DDD5")
    entry.pack(side="left", fill="x", expand=True, ipady=7, padx=(0, 10))

    def browse():
        d = filedialog.askdirectory(title="Select the Tier-1 Report Analysis folder")
        if d:
            state["folder"].set(d)
            refresh_file_count()

    def refresh_file_count():
        folder = state["folder"].get()
        n = len(discover_files(folder)) if os.path.isdir(folder) else 0
        stat_vars["found"].set(str(n))
        status_var.set(f"{n} DSCIT file(s) detected — ready to run." if n
                       else "No DSCIT files detected in this folder.")

    ttk.Button(inner, text="Browse…", style="Ghost.TButton",
               command=browse).pack(side="left", padx=(0, 8))
    run_btn = ttk.Button(inner, text="▶  Run Consolidation", style="TD.TButton")
    run_btn.pack(side="left")

    # ---------- stats row ----------
    stats = tk.Frame(body, bg=TD_MIST)
    stats.pack(fill="x", pady=(16, 0))
    stat_vars = {}

    def stat_card(parent, key, caption):
        c = tk.Frame(parent, bg=WHITE, highlightbackground="#E2E8E2",
                     highlightthickness=1)
        c.pack(side="left", fill="x", expand=True, padx=(0, 12))
        v = tk.StringVar(value="0")
        stat_vars[key] = v
        ttk.Label(c, textvariable=v, style="StatValue.TLabel").pack(
            anchor="w", padx=16, pady=(12, 0))
        ttk.Label(c, text=caption, style="StatCap.TLabel").pack(
            anchor="w", padx=16, pady=(0, 12))
        return c

    stat_card(stats, "found", "FILES FOUND")
    stat_card(stats, "processed", "FILES PROCESSED")
    stat_card(stats, "rows", "ROWS EXTRACTED")
    stat_card(stats, "warn", "WARNINGS")
    last = stat_card(stats, "err", "ERRORS")
    last.pack_configure(padx=(0, 0))

    # ---------- progress ----------
    pcard = tk.Frame(body, bg=TD_MIST)
    pcard.pack(fill="x", pady=(16, 0))
    pbar = ttk.Progressbar(pcard, style="TD.Horizontal.TProgressbar",
                           mode="determinate")
    pbar.pack(fill="x")
    status_var = tk.StringVar(value="Select your Tier-1 Report Analysis folder to begin.")
    tk.Label(pcard, textvariable=status_var, bg=TD_MIST, fg=GREY,
             font=("Segoe UI", 9)).pack(anchor="w", pady=(6, 0))

    # ---------- results table ----------
    tcard = tk.Frame(body, bg=WHITE, highlightbackground="#E2E8E2",
                     highlightthickness=1)
    tcard.pack(fill="both", expand=True, pady=(16, 0))
    cols = ("file", "status", "sheet", "hdr", "rows", "matched", "notes")
    tree = ttk.Treeview(tcard, columns=cols, show="headings")
    heads = {"file": ("File", 300), "status": ("Status", 70),
             "sheet": ("Sheet", 110), "hdr": ("Header Row", 85),
             "rows": ("Rows", 70), "matched": ("Cols Matched", 95),
             "notes": ("Notes", 300)}
    for c, (txt, w) in heads.items():
        tree.heading(c, text=txt)
        tree.column(c, width=w, anchor="w")
    vsb = ttk.Scrollbar(tcard, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)
    tree.pack(side="left", fill="both", expand=True, padx=(1, 0), pady=1)
    vsb.pack(side="right", fill="y")
    tree.tag_configure("OK", foreground=TD_GREEN_DARK)
    tree.tag_configure("WARN", foreground=AMBER)
    tree.tag_configure("ERROR", foreground=RED)

    # ---------- footer ----------
    foot = tk.Frame(body, bg=TD_MIST)
    foot.pack(fill="x", pady=(12, 0))
    open_btn = ttk.Button(foot, text="Open Consolidated Workbook",
                          style="Ghost.TButton", state="disabled",
                          command=lambda: state["out"] and os.startfile(state["out"]))
    open_btn.pack(side="right")
    open_folder_btn = ttk.Button(
        foot, text="Open Folder", style="Ghost.TButton", state="disabled",
        command=lambda: state["out"] and os.startfile(os.path.dirname(state["out"])))
    open_folder_btn.pack(side="right", padx=(0, 8))

    # ---------- run logic ----------
    def worker(folder):
        try:
            def cb(done, total, res):
                ui_queue.put(("progress", done, total, res))
            results, combined = run_consolidation(folder, progress_cb=cb)
            out = write_output(folder, results, combined)
            ui_queue.put(("done", results, combined, out))
        except Exception:
            ui_queue.put(("fatal", traceback.format_exc()))

    def start():
        folder = state["folder"].get().strip()
        if not os.path.isdir(folder):
            messagebox.showwarning("Folder required",
                                   "Please browse to a valid folder first.")
            return
        files = discover_files(folder)
        if not files:
            messagebox.showinfo("No files",
                                "No DSCIT*.xlsx files were found in that folder.")
            return
        for i in tree.get_children():
            tree.delete(i)
        for k in ("processed", "rows", "warn", "err"):
            stat_vars[k].set("0")
        stat_vars["found"].set(str(len(files)))
        pbar["value"] = 0
        pbar["maximum"] = len(files)
        state["running"] = True
        state["out"] = None
        run_btn.state(["disabled"])
        open_btn.state(["disabled"])
        open_folder_btn.state(["disabled"])
        status_var.set("Running…")
        threading.Thread(target=worker, args=(folder,), daemon=True).start()

    run_btn.configure(command=start)

    def poll():
        try:
            while True:
                msg = ui_queue.get_nowait()
                if msg[0] == "progress":
                    _, done, total, res = msg
                    pbar["value"] = done
                    stat_vars["processed"].set(str(done))
                    stat_vars["rows"].set(
                        str(int(stat_vars["rows"].get()) + res.rows))
                    if res.status == "WARN":
                        stat_vars["warn"].set(str(int(stat_vars["warn"].get()) + 1))
                    if res.status == "ERROR":
                        stat_vars["err"].set(str(int(stat_vars["err"].get()) + 1))
                    tree.insert("", "end", tags=(res.status,), values=(
                        res.file, res.status, res.sheet,
                        res.header_row or "—", res.rows,
                        f"{res.matched}/{len(TARGET_COLUMNS)}",
                        (res.message + ("  |  Missing: " + ", ".join(res.missing)
                                        if res.missing else ""))[:180]))
                    status_var.set(f"Processed {done} of {total}…")
                elif msg[0] == "done":
                    _, results, combined, out = msg
                    state["running"] = False
                    state["out"] = out
                    run_btn.state(["!disabled"])
                    open_btn.state(["!disabled"])
                    open_folder_btn.state(["!disabled"])
                    status_var.set(
                        f"Complete — {len(combined):,} rows from "
                        f"{sum(1 for r in results if r.status != 'ERROR')} file(s). "
                        f"Saved: {os.path.basename(out)}")
                elif msg[0] == "fatal":
                    state["running"] = False
                    run_btn.state(["!disabled"])
                    status_var.set("Run failed — see details.")
                    messagebox.showerror("Run failed", msg[1][-1500:])
        except queue.Empty:
            pass
        root.after(100, poll)

    poll()
    root.mainloop()


if __name__ == "__main__":
    if "--headless" in sys.argv:
        # optional CLI mode: python dscit_consolidator.py --headless <folder>
        folder = sys.argv[sys.argv.index("--headless") + 1]
        results, combined = run_consolidation(
            folder, progress_cb=lambda d, t, r: print(
                f"[{d}/{t}] {r.file}: {r.status} rows={r.rows} {r.message}"))
        out = write_output(folder, results, combined)
        print(f"\nSaved: {out}  ({len(combined):,} rows)")
    else:
        launch_ui()
