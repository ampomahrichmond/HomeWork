"""
DSCIT Tier-1 Report Consolidator + Executive Analytics
-------------------------------------------------------
v2 enhancements:
  * Database Type normalization (Oracle/ORACLE/Oracle 19c/Essbase -> Oracle, etc.)
  * Analytics engine (EUC/MAL, outputs, sources, segments, owners, DB landscape,
    gap analysis, completeness scoring)
  * Ultra-premium TD-themed executive HTML dashboard (self-contained, no
    internet needed - safe to email to VP/SVP)
  * Analytics tabs added to the Excel output
  * New "Generate Executive Dashboard" button in the UI

Run:  python dscit_consolidator.py
Requires:  pip install openpyxl pandas
"""

import os
import re
import sys
import glob
import json
import hashlib
import html
import queue
import threading
import traceback
import webbrowser
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

# ============================================================================
# 1. EXTRACTION ENGINE
# ============================================================================

TARGET_COLUMNS = [
    "Output Name",
    "Output Owner",
    "Data Element",
    "Data Element Description",
    "Data Element Indicator",
    "Source Name",
    "Source Type",
    "Source Application EUC / MAL Code",
    "Database Type",
    "Database Name",
    "Schema Name / File Path / API",
    "Database",
    "Business Segment / Corporate Function",
    "Asset Owner Name",
    "Technology Asset Owner Name",
    "Data Owner Name",
    "Lineage System",
    "Lineage Database",
    "Lineage Schema",
    "Lineage Physical Table / API",
    "Lineage Physical Column / API",
]

# Columns that live on the "Data Flow Linkage Information" (lineage) side of
# the DSCIT sheet. Their headers (System, Database, Schema...) collide with
# source-side names, so they are mapped with section awareness.
LINEAGE_COLUMNS = [
    "Lineage System",
    "Lineage Database",
    "Lineage Schema",
    "Lineage Physical Table / API",
    "Lineage Physical Column / API",
]

COLUMN_ALIASES = {
    "Output Name": ["outputname", "reportname", "outputreportname"],
    "Output Owner": ["outputowner", "reportowner"],
    "Data Element": ["dataelement", "dataelementname", "criticaldataelement", "cde"],
    "Data Element Description": [
        "dataelementdescription", "dataelementdiscription",
        "dataelementdesc", "elementdescription",
    ],
    "Source Name": ["sourcename", "sourcesystemname", "datasourcename"],
    "Source Type": ["sourcetype", "sourcesystemtype"],
    "Source Application EUC / MAL Code": [
        "sourceapplicationeucmalcode", "sourceapplicationeucormalcode",
        "applicationeucmalcode", "eucmalcode", "sourceappeucmalcode",
        "sourceapplicationmalcode", "sourceapplicationeuccode", "malcode",
    ],
    "Database Type": ["databasetype", "dbtype"],
    "Database Name": ["databasename", "dbname"],
    "Database": ["database", "db"],
    "Business Segment / Corporate Function": [
        "businesssegmentcorporatefunction", "businesssegmentorcorporatefunction",
        "businesssegment", "corporatefunction", "segmentfunction",
    ],
    "Asset Owner Name": ["assetownername", "assetowner", "dataassetownername"],
    "Technology Asset Owner Name": [
        "technologyassetownername", "technologyassetowner",
        "techassetownername", "techassetowner", "taoname",
    ],
    "Data Owner Name": ["dataownername", "dataowner"],
    "Data Element Indicator": [
        "dataelementindicator", "dataelementindicatorcdebdemetric",
        "elementindicator", "cdebdemetricindicator", "deindicator",
    ],
    "Schema Name / File Path / API": [
        "schemanamefilepathapi", "schemanamefilepathorapi",
        "schemanamefilepath", "schemanameorfilepath", "schemaname",
        "sourceschemaname",
    ],
    "Lineage System": ["system", "lineagesystem", "targetsystem",
                       "dataflowsystem", "destinationsystem"],
    "Lineage Database": ["database", "lineagedatabase", "targetdatabase",
                         "destinationdatabase"],
    "Lineage Schema": ["schema", "lineageschema", "targetschema",
                       "destinationschema"],
    "Lineage Physical Table / API": [
        "physicaltablenameapidatastructure", "physicaltablename",
        "physicaltableapidatastructure", "physicaltablenameapi",
        "lineagephysicaltable", "targetphysicaltable", "physicaltable",
    ],
    "Lineage Physical Column / API": [
        "physicalcolumnnameapidataelement", "physicalcolumnname",
        "physicalcolumnapidataelement", "physicalcolumnnameapi",
        "lineagephysicalcolumn", "targetphysicalcolumn", "physicalcolumn",
    ],
}

HEADER_SCAN_ROWS = 40
HEADER_SCAN_COLS = 5
SHEET_KEY = "dscit"


def _norm(value) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(value).lower())


def find_dscit_sheet(sheetnames):
    candidates = [s for s in sheetnames if SHEET_KEY in _norm(s)]
    if not candidates:
        return None
    candidates.sort(key=lambda s: (_norm(s) != SHEET_KEY, len(s)))
    return candidates[0]


def find_header_row(rows):
    for idx, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        for cell in row[:HEADER_SCAN_COLS]:
            n = _norm(cell)
            if n == "objectid":
                return idx, row
            if n:
                break
    return None, None


_LINEAGE_BANNER_KEYS = ("dataflowlinkage", "dataflowlineage",
                        "datalineage", "lineageinformation",
                        "linkageinformation")


def find_lineage_boundary(rows, hdr_idx):
    """Locate the column where the 'Data Flow Linkage Information' section
    begins by scanning the banner rows just above the header row. Merged
    banner cells surface their value at the section's first column."""
    for row in rows[max(0, hdr_idx - 6):hdr_idx]:
        for i, cell in enumerate(row):
            n = _norm(cell)
            if n and any(k in n for k in _LINEAGE_BANNER_KEYS):
                return i
    # No banner found: infer from the header row itself. A bare "System"
    # header followed by bare "Database"/"Schema"/"Physical Table..." is the
    # signature of the Data Flow Linkage block (source-side headers are
    # always qualified, e.g. "Database Name", "Source Physical Table...").
    header = rows[hdr_idx]
    followers = ("database", "schema")
    for i in range(len(header) - 1, -1, -1):
        if _norm(header[i]) != "system":
            continue
        rest = [_norm(h) for h in header[i + 1:] if _norm(h)]
        if any(h in followers or h.startswith("physicaltable")
               or h.startswith("physicalcolumn") for h in rest):
            return i
    return None


def map_columns(header_row, lineage_start=None):
    headers = {i: _norm(h) for i, h in enumerate(header_row) if _norm(h)}
    mapping, claimed = {}, set()

    src_targets = [t for t in TARGET_COLUMNS if t not in LINEAGE_COLUMNS]

    def src_items():
        return sorted((i, h) for i, h in headers.items()
                      if i not in claimed
                      and (lineage_start is None or i < lineage_start))

    order = sorted(src_targets, key=lambda t: -len(_norm(t)))
    for target in order:                                    # exact pass
        aliases = set(COLUMN_ALIASES[target]) | {_norm(target)}
        for idx, h in src_items():
            if h in aliases:
                mapping[target] = idx
                claimed.add(idx)
                break
    for target in order:                                    # contains pass
        if target in mapping:
            continue
        aliases = sorted(set(COLUMN_ALIASES[target]) | {_norm(target)},
                         key=len, reverse=True)
        for alias in aliases:
            if len(alias) < 6:
                continue
            hit = next((i for i, h in src_items() if alias in h), None)
            if hit is not None:
                mapping[target] = hit
                claimed.add(hit)
                break

    # ---- lineage side ----
    def lin_items():
        return sorted((i, h) for i, h in headers.items()
                      if i not in claimed
                      and (lineage_start is None or i >= lineage_start))

    lorder = sorted(LINEAGE_COLUMNS, key=lambda t: -len(_norm(t)))
    for target in lorder:                                   # exact pass
        aliases = set(COLUMN_ALIASES[target]) | {_norm(target)}
        cand = [i for i, h in lin_items() if h in aliases]
        if cand:
            # With a known boundary take the first match inside the lineage
            # region; without one, take the RIGHTMOST occurrence, since the
            # lineage block sits to the right of Source Information.
            idx = cand[0] if lineage_start is not None else cand[-1]
            mapping[target] = idx
            claimed.add(idx)
    if lineage_start is not None:                           # contains pass
        for target in lorder:
            if target in mapping:
                continue
            aliases = sorted(set(COLUMN_ALIASES[target]) | {_norm(target)},
                             key=len, reverse=True)
            for alias in aliases:
                if len(alias) < 6:
                    continue
                hit = next((i for i, h in lin_items() if alias in h), None)
                if hit is not None:
                    mapping[target] = hit
                    claimed.add(hit)
                    break
    return mapping


# ============================================================================
# 2. DATABASE TYPE NORMALIZATION
# ============================================================================
# Rules are checked top-to-bottom against the normalized value. First hit wins.

DB_TYPE_RULES = [
    (r"essbase|oracle|exadata|oradb", "Oracle"),
    (r"excel|xls|spreadsheet|workbook", "Excel"),
    (r"sqlserver|mssql|microsoftsql|sqlsrv", "SQL Server"),
    (r"db2|udb", "DB2"),
    (r"teradata", "Teradata"),
    (r"snowflake", "Snowflake"),
    (r"databricks|deltalake|unitycatalog", "Databricks"),
    (r"azure|synapse|adls|blob|awss3|amazons3|gcs|bigquery|redshift|cloud",
     "Cloud"),
    (r"hive|hadoop|impala|hdfs", "Hadoop"),
    (r"postgres", "PostgreSQL"),
    (r"mysql|mariadb", "MySQL"),
    (r"mongo", "MongoDB"),
    (r"msaccess|accessdb|access", "MS Access"),
    (r"sharepoint", "SharePoint"),
    (r"csvfile|csv|flatfile|textfile|txtfile|delimited", "Flat File"),
    (r"sasdataset|sas", "SAS"),
    (r"mainframe|vsam|imsdb|ims|cobol", "Mainframe"),
    (r"sybase", "Sybase"),
    (r"informix", "Informix"),
    (r"netezza", "Netezza"),
]

_DB_RULES_COMPILED = [(re.compile(p), canon) for p, canon in DB_TYPE_RULES]


def normalize_db_type(raw):
    """Map any raw Database Type variant to a canonical name."""
    if raw is None or (isinstance(raw, float) and raw != raw) \
            or str(raw).strip() == "":
        return "Not Specified"
    n = _norm(raw)
    if not n or n in ("na", "nan", "none", "null", "tbd", "unknown"):
        return "Not Specified"
    for rx, canon in _DB_RULES_COMPILED:
        if rx.search(n):
            return canon
    return str(raw).strip().title()


# ============================================================================
# 3. FILE PROCESSING
# ============================================================================

@dataclass
class FileResult:
    file: str
    size: int = 0
    status: str = "OK"
    sheet: str = ""
    header_row: int | None = None
    rows: int = 0
    matched: int = 0
    missing: list = field(default_factory=list)
    message: str = ""
    data: pd.DataFrame | None = None


def human_size(n):
    for u in ("B", "KB", "MB", "GB"):
        if n < 1024 or u == "GB":
            return f"{n:,.0f} {u}" if u == "B" else f"{n / 1:,.1f} {u}"
        n /= 1024
    return f"{n:,.1f} GB"


def process_file(path: str) -> FileResult:
    fname = os.path.basename(path)
    res = FileResult(file=fname)
    try:
        res.size = os.path.getsize(path)
    except OSError:
        pass
    try:
        if fname.lower().endswith(".xls"):
            xls = pd.ExcelFile(path)
            sheet_name = find_dscit_sheet(xls.sheet_names)
            if not sheet_name:
                res.status, res.message = "ERROR", "No DSCIT sheet found"
                return res
            res.sheet = sheet_name
            df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
            rows = [tuple(r) for r in df.itertuples(index=False, name=None)]
            return _extract(rows, res)

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet_name = find_dscit_sheet(wb.sheetnames)
            if not sheet_name:
                res.status, res.message = "ERROR", "No DSCIT sheet found"
                return res
            res.sheet = sheet_name
            ws = wb[sheet_name]
            rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
        return _extract(rows, res)
    except Exception as e:
        res.status = "ERROR"
        res.message = f"{type(e).__name__}: {e}"
        return res


def _extract(rows, res: FileResult) -> FileResult:
    hdr_idx, header = find_header_row(rows)
    if hdr_idx is None:
        res.status, res.message = "ERROR", "Header row with 'Object ID' not found"
        return res
    res.header_row = hdr_idx + 1

    lineage_start = find_lineage_boundary(rows, hdr_idx)
    mapping = map_columns(header, lineage_start)
    res.matched = len(mapping)
    res.missing = [t for t in TARGET_COLUMNS if t not in mapping]
    if not mapping:
        res.status, res.message = "ERROR", "No target columns matched"
        return res
    if res.missing:
        res.status = "WARN"
        res.message = f"{len(res.missing)} column(s) not found"

    records = []
    consecutive_blank = 0
    for row in rows[hdr_idx + 1:]:
        rec, blank = {}, True
        for target in TARGET_COLUMNS:
            idx = mapping.get(target)
            val = row[idx] if (idx is not None and idx < len(row)) else None
            if val is not None and str(val).strip() != "":
                blank = False
                rec[target] = str(val).strip() if isinstance(val, str) else val
            else:
                rec[target] = None
        if not blank:
            records.append(rec)
            consecutive_blank = 0
        else:
            consecutive_blank += 1
            if consecutive_blank >= 300:      # stray-formatting ghost rows
                break

    df = pd.DataFrame(records, columns=TARGET_COLUMNS)
    # --- Database Type normalization (keep original for audit) ---
    df["Database Type (Original)"] = df["Database Type"]
    df["Database Type"] = df["Database Type"].map(normalize_db_type)
    df.insert(0, "DSCIT File Name", res.file)
    res.rows = len(df)
    res.data = df
    return res


def discover_files(folder: str):
    pats = ["DSCIT*.xlsx", "DSCIT*.xlsm", "DSCIT*.xls",
            "dscit*.xlsx", "dscit*.xlsm", "dscit*.xls"]
    files = set()
    for p in pats:
        files.update(glob.glob(os.path.join(folder, p)))
    return sorted(f for f in files
                  if not os.path.basename(f).startswith("~$")
                  and "consolidated" not in os.path.basename(f).lower())


def _file_md5(path, chunk=1 << 20):
    h = hashlib.md5()
    with open(path, "rb") as f:
        while True:
            b = f.read(chunk)
            if not b:
                break
            h.update(b)
    return h.hexdigest()


_ID_RX = re.compile(r"(?i)^DSCIT[\s_\-]*([A-Za-z0-9]+)")
_DATE_RX = re.compile(r"(20\d{6})")


def _dscit_id(fname):
    m = _ID_RX.match(fname)
    return m.group(1).upper() if m else fname.upper()


def _file_version_key(path):
    """Sort key for 'newest' file: date in the name, else modified time."""
    m = _DATE_RX.search(os.path.basename(path))
    if m:
        return (1, int(m.group(1)))
    try:
        return (0, os.path.getmtime(path))
    except OSError:
        return (0, 0)


def dedupe_files(files, keep_latest_per_id=False):
    """Return (files_to_process, skipped) where skipped is a list of
    (path, reason).

    Default behaviour is CONTENT-BASED only: a file is skipped solely when
    its bytes are identical (MD5) to a file already kept. Similar-looking
    file names are NOT treated as duplicates — files sharing a DSCIT ID
    can legitimately carry different content (different Output Names), so
    every distinct file is processed and duplicated DATA is removed at the
    row level instead.

    keep_latest_per_id=True enables the optional filename-versioning mode
    (keep only the newest file per DSCIT ID). Use only if IDs are known to
    be unique per report."""
    skipped = []

    # -- identical content --
    by_hash = {}
    survivors = []
    for f in files:
        try:
            h = _file_md5(f)
        except OSError as e:
            skipped.append((f, f"Unreadable: {e}"))
            continue
        if h in by_hash:
            skipped.append(
                (f, "Duplicate file — identical content to "
                    f"{os.path.basename(by_hash[h])}"))
        else:
            by_hash[h] = f
            survivors.append(f)

    # -- superseded versions of the same DSCIT ID --
    if keep_latest_per_id:
        by_id = {}
        for f in survivors:
            by_id.setdefault(_dscit_id(os.path.basename(f)), []).append(f)
        final = []
        for fid, group in by_id.items():
            if len(group) == 1:
                final.extend(group)
                continue
            group.sort(key=_file_version_key, reverse=True)
            final.append(group[0])
            for g in group[1:]:
                skipped.append(
                    (g, f"Superseded — newer file for DSCIT {fid}: "
                        f"{os.path.basename(group[0])}"))
        survivors = sorted(final)

    return survivors, skipped


def run_consolidation(folder, progress_cb=None, workers=8,
                      keep_latest_per_id=False):
    all_files = discover_files(folder)
    files, skipped = dedupe_files(all_files, keep_latest_per_id)
    total = len(all_files)
    results, done = [], 0

    for path, reason in skipped:
        r = FileResult(file=os.path.basename(path), status="SKIP",
                       message=reason)
        try:
            r.size = os.path.getsize(path)
        except OSError:
            pass
        results.append(r)
        done += 1
        if progress_cb:
            progress_cb(done, total, r)

    if files:
        with ThreadPoolExecutor(
                max_workers=min(workers, max(1, len(files)))) as ex:
            futures = {ex.submit(process_file, f): f for f in files}
            for fut in as_completed(futures):
                results.append(fut.result())
                done += 1
                if progress_cb:
                    progress_cb(done, total, results[-1])

    results.sort(key=lambda r: r.file.lower())
    frames = [r.data for r in results if r.data is not None and not r.data.empty]
    cols = ["DSCIT File Name"] + TARGET_COLUMNS + ["Database Type (Original)"]
    combined = pd.concat(frames, ignore_index=True) if frames else \
        pd.DataFrame(columns=cols)

    # -- row-level dedup: identical records inflate every count --
    # Content-based: a row is a duplicate when ALL 14 extracted fields
    # match an earlier row, regardless of which file it came from.
    rows_before = len(combined)
    per_file = pd.DataFrame(columns=["File", "Rows Extracted",
                                     "Duplicate Rows Removed",
                                     "Unique Rows Kept"])
    overlap = pd.DataFrame(columns=["File", "Duplicates Rows First Seen In",
                                    "Overlapping Rows"])
    if rows_before:
        dup_mask = combined.duplicated(subset=TARGET_COLUMNS, keep="first")
        key_hash = pd.util.hash_pandas_object(
            combined[TARGET_COLUMNS], index=False)
        first_file = combined["DSCIT File Name"].groupby(
            key_hash.values).transform("first")
        per_file = (pd.DataFrame({"File": combined["DSCIT File Name"],
                                  "dup": dup_mask})
                    .groupby("File")["dup"]
                    .agg([("Rows Extracted", "size"),
                          ("Duplicate Rows Removed", "sum")])
                    .reset_index())
        per_file["Duplicate Rows Removed"] = \
            per_file["Duplicate Rows Removed"].astype(int)
        per_file["Unique Rows Kept"] = (per_file["Rows Extracted"]
                                        - per_file["Duplicate Rows Removed"])
        if dup_mask.any():
            overlap = (pd.DataFrame({
                "File": combined.loc[dup_mask, "DSCIT File Name"],
                "Duplicates Rows First Seen In": first_file[dup_mask]})
                .groupby(["File", "Duplicates Rows First Seen In"])
                .size().reset_index(name="Overlapping Rows")
                .sort_values("Overlapping Rows", ascending=False))
        combined = combined[~dup_mask].reset_index(drop=True)
    dedup = {
        "files_found": total,
        "files_processed": len(files),
        "identical_files_skipped": sum(
            1 for _, r in skipped if r.startswith("Duplicate file")),
        "superseded_files_skipped": sum(
            1 for _, r in skipped if r.startswith("Superseded")),
        "rows_before_dedup": rows_before,
        "duplicate_rows_removed": rows_before - len(combined),
        "rows_after_dedup": len(combined),
        "skipped_detail": [(os.path.basename(p), r) for p, r in skipped],
        "per_file": per_file,
        "overlap": overlap,
    }
    return results, combined, dedup


# ============================================================================
# 4. ANALYTICS ENGINE
# ============================================================================

def classify_indicator(v):
    """Canonicalize Data Element Indicator values to CDE / BDE / Metric."""
    if v is None or (isinstance(v, float) and v != v):
        return None
    n = _norm(v)
    if not n or n in ("na", "nan", "none", "null", "tbd", "unknown"):
        return None
    if "cde" in n or "critical" in n:
        return "CDE"
    if "bde" in n or "basicdata" in n or "businessdata" in n:
        return "BDE"
    if "metric" in n or "kpi" in n:
        return "Metric"
    return str(v).strip().title()


def lineage_status_series(df):
    """Full Lineage = all five lineage fields populated; Partial = some;
    No Lineage = none."""
    present = pd.DataFrame({c: df[c].notna() for c in LINEAGE_COLUMNS})
    cnt = present.sum(axis=1)
    return cnt.map(lambda c: "Full Lineage" if c == len(LINEAGE_COLUMNS)
                   else ("Partial Lineage" if c > 0 else "No Lineage"))


def _clean(series):
    """Series of stripped strings with blanks/placeholder values as NA."""
    s = series.astype("string").str.strip()
    s = s.mask(s.str.lower().isin(["", "na", "n/a", "none", "null", "tbd", "-"]))
    return s


def _nunique(df, col):
    return int(_clean(df[col]).dropna().nunique()) if col in df else 0


def build_analytics(combined: pd.DataFrame, results=None, dedup=None):
    df = combined.copy()
    for c in TARGET_COLUMNS:
        if c in df:
            df[c] = _clean(df[c])

    a = {}

    # ---- KPIs ----
    kpi = {
        "Files Consolidated": df["DSCIT File Name"].nunique(),
        "Total Data Element Rows": len(df),
        "Distinct Data Elements": _nunique(df, "Data Element"),
        "Output Names": _nunique(df, "Output Name"),
        "Output Owners": _nunique(df, "Output Owner"),
        "Source Systems (Source Names)": _nunique(df, "Source Name"),
        "Source Types": _nunique(df, "Source Type"),
        "EUC / MAL Codes": _nunique(df, "Source Application EUC / MAL Code"),
        "Database Names": _nunique(df, "Database Name"),
        "Database Types": _nunique(df, "Database Type"),
        "Business Segments / Corp Functions":
            _nunique(df, "Business Segment / Corporate Function"),
        "Asset Owners": _nunique(df, "Asset Owner Name"),
        "Technology Asset Owners": _nunique(df, "Technology Asset Owner Name"),
        "Data Owners": _nunique(df, "Data Owner Name"),
    }
    df["_ind"] = df["Data Element Indicator"].map(classify_indicator)
    df["_lstat"] = lineage_status_series(df)

    def _elems(mask):
        return int(df.loc[mask, "Data Element"].dropna().nunique())
    kpi["Schemas (Schema Name / File Path / API)"] = \
        _nunique(df, "Schema Name / File Path / API")
    kpi["CDEs"] = _elems(df["_ind"] == "CDE")
    kpi["BDEs"] = _elems(df["_ind"] == "BDE")
    kpi["Metrics"] = _elems(df["_ind"] == "Metric")
    kpi["Rows with Full Lineage"] = int((df["_lstat"] == "Full Lineage").sum())
    kpi["Rows with Partial Lineage"] = \
        int((df["_lstat"] == "Partial Lineage").sum())
    kpi["Rows with No Lineage"] = int((df["_lstat"] == "No Lineage").sum())
    kpi["Lineage Coverage %"] = round(
        100 * (len(df) - int((df["_lstat"] == "No Lineage").sum()))
        / len(df), 1) if len(df) else 0.0
    if dedup:
        kpi["Duplicate Rows Removed"] = dedup.get("duplicate_rows_removed", 0)
        kpi["Duplicate Files Skipped"] = (
            dedup.get("identical_files_skipped", 0)
            + dedup.get("superseded_files_skipped", 0))
    a["kpis"] = kpi

    # ---- By Database Type (normalized) ----
    g = (df.groupby("Database Type", dropna=False)
           .agg(Databases=("Database Name", lambda s: s.dropna().nunique()),
                Data_Element_Rows=("Database Type", "size"),
                Outputs=("Output Name", lambda s: s.dropna().nunique()),
                Sources=("Source Name", lambda s: s.dropna().nunique()))
           .reset_index()
           .sort_values("Data_Element_Rows", ascending=False))
    g.columns = ["Database Type", "Distinct Databases", "Data Element Rows",
                 "Outputs", "Sources"]
    a["db_type"] = g

    # ---- Database Name x Type inventory ----
    inv = (df.dropna(subset=["Database Name"])
             .groupby(["Database Type", "Database Name"])
             .agg(Rows=("Database Name", "size"),
                  Outputs=("Output Name", lambda s: s.dropna().nunique()))
             .reset_index()
             .sort_values(["Database Type", "Rows"],
                          ascending=[True, False]))
    a["db_inventory"] = inv

    # ---- Outputs by Business Segment ----
    seg = (df.groupby("Business Segment / Corporate Function", dropna=False)
             .agg(Outputs=("Output Name", lambda s: s.dropna().nunique()),
                  Output_Owners=("Output Owner", lambda s: s.dropna().nunique()),
                  Data_Elements=("Data Element", lambda s: s.dropna().nunique()),
                  Rows=("Output Name", "size"))
             .reset_index()
             .sort_values("Outputs", ascending=False))
    seg["Business Segment / Corporate Function"] = \
        seg["Business Segment / Corporate Function"].fillna("Not Specified")
    seg.columns = ["Business Segment / Corporate Function", "Outputs",
                   "Output Owners", "Data Elements", "Rows"]
    a["segment"] = seg

    # ---- Databases by Asset Owner ----
    ao = (df.groupby("Asset Owner Name", dropna=False)
            .agg(Databases=("Database Name", lambda s: s.dropna().nunique()),
                 DB_Types=("Database Type", lambda s: s.dropna().nunique()),
                 Outputs=("Output Name", lambda s: s.dropna().nunique()))
            .reset_index()
            .sort_values("Databases", ascending=False))
    ao["Asset Owner Name"] = ao["Asset Owner Name"].fillna("Not Specified")
    ao.columns = ["Asset Owner Name", "Databases", "Database Types", "Outputs"]
    a["asset_owner"] = ao

    # ---- Databases by Output Name / Output Owner / Source Type ----
    def dbs_by(col):
        t = (df.groupby(col, dropna=False)
               .agg(Databases=("Database Name", lambda s: s.dropna().nunique()),
                    DB_Types=("Database Type",
                              lambda s: ", ".join(sorted(s.dropna().unique())[:6])))
               .reset_index()
               .sort_values("Databases", ascending=False))
        t[col] = t[col].fillna("Not Specified")
        t.columns = [col, "Databases", "Database Types"]
        return t
    a["dbs_by_output"] = dbs_by("Output Name")
    a["dbs_by_owner"] = dbs_by("Output Owner")
    a["dbs_by_srctype"] = dbs_by("Source Type")

    # ---- Source inventory + gap flags ----
    src = (df.groupby(["Source Name", "Source Application EUC / MAL Code",
                       "Source Type"], dropna=False)
             .agg(Databases=("Database Name", lambda s: s.dropna().nunique()),
                  Rows=("Source Name", "size"),
                  Missing_DB_Name=("Database Name",
                                   lambda s: int(s.isna().sum())),
                  Missing_Data_Owner=("Data Owner Name",
                                      lambda s: int(s.isna().sum())),
                  Missing_Asset_Owner=("Asset Owner Name",
                                       lambda s: int(s.isna().sum())))
             .reset_index())
    src["Source Name"] = src["Source Name"].fillna("«Source Missing»")
    src["Source Application EUC / MAL Code"] = \
        src["Source Application EUC / MAL Code"].fillna("«Code Missing»")
    src["Source Type"] = src["Source Type"].fillna("«Type Missing»")
    src["Gap Flags"] = src.apply(
        lambda r: "; ".join(f for f, c in [
            ("Source missing", r["Source Name"] == "«Source Missing»"),
            ("EUC/MAL missing", r["Source Application EUC / MAL Code"] == "«Code Missing»"),
            ("DB name gaps", r["Missing_DB_Name"] > 0),
            ("Data owner gaps", r["Missing_Data_Owner"] > 0),
            ("Asset owner gaps", r["Missing_Asset_Owner"] > 0),
        ] if c) or "Complete", axis=1)
    src = src.sort_values("Rows", ascending=False)
    a["source_inventory"] = src

    # ---- Data quality / completeness ----
    dq_rows = []
    for c in TARGET_COLUMNS:
        if c not in df:
            continue
        missing = int(df[c].isna().sum())
        pct = round(100 * (1 - missing / len(df)), 1) if len(df) else 0.0
        dq_rows.append({"Field": c, "Populated %": pct,
                        "Missing Rows": missing, "Total Rows": len(df)})
    dq = pd.DataFrame(dq_rows).sort_values("Populated %")
    a["dq"] = dq
    key_fields = ["Output Name", "Source Name", "Database Name",
                  "Data Owner Name", "Asset Owner Name",
                  "Source Application EUC / MAL Code"]
    kdq = dq[dq["Field"].isin(key_fields)]
    a["completeness_score"] = round(float(kdq["Populated %"].mean()), 1) \
        if len(kdq) else 0.0

    # ---- Top outputs by data-element footprint ----
    top = (df.groupby("Output Name", dropna=False)
             .agg(Data_Elements=("Data Element", lambda s: s.dropna().nunique()),
                  Sources=("Source Name", lambda s: s.dropna().nunique()),
                  Databases=("Database Name", lambda s: s.dropna().nunique()),
                  Owner=("Output Owner",
                         lambda s: s.dropna().iloc[0] if s.dropna().size else "—"))
             .reset_index()
             .sort_values("Data_Elements", ascending=False))
    top["Output Name"] = top["Output Name"].fillna("Not Specified")
    top.columns = ["Output Name", "Data Elements", "Sources", "Databases",
                   "Output Owner"]
    a["top_outputs"] = top

    # ---- Schemas by Database (source side) ----
    sdb = (df.dropna(subset=["Database Name"])
             .groupby(["Database Type", "Database Name"], dropna=False)
             .agg(Schemas=("Schema Name / File Path / API",
                           lambda s: s.dropna().nunique()),
                  Elements=("Data Element", lambda s: s.dropna().nunique()),
                  Outputs=("Output Name", lambda s: s.dropna().nunique()),
                  Rows=("Database Name", "size"))
             .reset_index()
             .sort_values("Schemas", ascending=False))
    a["schemas_by_db"] = sdb

    # ---- CDE / BDE / Metric inventory ----
    ind = (df.assign(_i=df["_ind"].fillna("Not Specified"))
             .groupby("_i")
             .agg(Elements=("Data Element", lambda s: s.dropna().nunique()),
                  Outputs=("Output Name", lambda s: s.dropna().nunique()),
                  Databases=("Database Name",
                             lambda s: s.dropna().nunique()),
                  Rows=("_i", "size"))
             .reset_index()
             .rename(columns={"_i": "Data Element Indicator"})
             .sort_values("Elements", ascending=False))
    a["cde_inventory"] = ind
    seg_ind = (df.assign(_i=df["_ind"].fillna("Not Specified"),
                         _s=df["Business Segment / Corporate Function"]
                         .fillna("Not Specified"))
               .groupby(["_s", "_i"])["Data Element"]
               .apply(lambda s: s.dropna().nunique())
               .unstack(fill_value=0).reset_index()
               .rename(columns={"_s":
                                "Business Segment / Corporate Function"}))
    a["cde_by_segment"] = seg_ind

    # ---- Lineage coverage ----
    lstat = (df.groupby("_lstat")
               .agg(Rows=("_lstat", "size"),
                    Elements=("Data Element",
                              lambda s: s.dropna().nunique()),
                    Outputs=("Output Name",
                             lambda s: s.dropna().nunique()))
               .reset_index()
               .rename(columns={"_lstat": "Lineage Status"}))
    a["lineage_status"] = lstat
    lsys = (df.dropna(subset=["Lineage System"])
              .groupby("Lineage System")
              .agg(Rows=("Lineage System", "size"),
                   Elements=("Data Element",
                             lambda s: s.dropna().nunique()),
                   Lineage_DBs=("Lineage Database",
                                lambda s: s.dropna().nunique()),
                   Source_DBs=("Database Name",
                               lambda s: s.dropna().nunique()))
              .reset_index()
              .sort_values("Rows", ascending=False))
    lsys.columns = ["Lineage System", "Rows", "Elements",
                    "Lineage Databases", "Source Databases"]
    a["lineage_systems"] = lsys

    # ---- Harvest Priority: elements/CDEs lacking lineage, by source DB ----
    hv_rows = []
    for (dbt_, dbn_), g in df.groupby(["Database Type", "Database Name"],
                                      dropna=False):
        if dbn_ is None or (isinstance(dbn_, float) and dbn_ != dbn_):
            dbn_ = "«Name Missing»"
        elems = g["Data Element"].dropna().nunique()
        with_lin = g.loc[g["_lstat"] != "No Lineage",
                         "Data Element"].dropna().nunique()
        cde_g = g[g["_ind"] == "CDE"]
        cdes = cde_g["Data Element"].dropna().nunique()
        cde_with = cde_g.loc[cde_g["_lstat"] != "No Lineage",
                             "Data Element"].dropna().nunique()
        hv_rows.append({
            "Database Type": dbt_ if isinstance(dbt_, str) else
            "Not Specified",
            "Database Name": dbn_,
            "Elements": elems,
            "CDEs": cdes,
            "Elements w/o Lineage": elems - with_lin,
            "CDEs w/o Lineage": cdes - cde_with,
            "Schemas": g["Schema Name / File Path / API"]
            .dropna().nunique(),
            "Outputs": g["Output Name"].dropna().nunique(),
            "Rows": len(g),
            "Lineage Coverage %": round(100 * with_lin / elems, 1)
            if elems else 0.0,
        })
    hv = pd.DataFrame(hv_rows).sort_values(
        ["CDEs w/o Lineage", "Elements w/o Lineage"],
        ascending=False).reset_index(drop=True)
    hv.insert(0, "Priority Rank", range(1, len(hv) + 1))
    a["harvest"] = hv

    # ---- Lineage target inventory (system > db > schema > table > column) --
    ldet = (df.dropna(subset=["Lineage System"], how="all")
              .groupby(LINEAGE_COLUMNS, dropna=False)
              .agg(Rows=("Lineage System", "size"),
                   Elements=("Data Element",
                             lambda s: s.dropna().nunique()))
              .reset_index()
              .sort_values("Rows", ascending=False))
    ldet = ldet[ldet[LINEAGE_COLUMNS].notna().any(axis=1)]
    a["lineage_detail"] = ldet

    df.drop(columns=["_ind", "_lstat"], inplace=True, errors="ignore")

    # ---- DB type normalization audit map ----
    if "Database Type (Original)" in df:
        m = (combined.groupby(["Database Type (Original)", "Database Type"],
                              dropna=False).size().reset_index(name="Rows")
             .sort_values(["Database Type", "Rows"], ascending=[True, False]))
        m.columns = ["Raw Value", "Normalized To", "Rows"]
        a["norm_map"] = m

    return a


# ============================================================================
# 5. EXCEL OUTPUT (data + analytics tabs)
# ============================================================================

_ILLEGAL_XML = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _excel_safe(v):
    """Make any cell value safe for openpyxl: strip illegal XML control
    characters from strings, drop timezones from datetimes, stringify
    exotic objects."""
    if v is None:
        return v
    if isinstance(v, str):
        return _ILLEGAL_XML.sub("", v)
    if isinstance(v, datetime):
        return v.replace(tzinfo=None) if v.tzinfo else v
    tz = getattr(v, "tzinfo", None)          # pandas Timestamp etc.
    if tz is not None:
        try:
            return v.tz_localize(None)
        except Exception:
            return str(v)
    if isinstance(v, (int, float, bool)):
        return v
    if hasattr(v, "isoformat") or isinstance(v, (bytes, complex)):
        return _ILLEGAL_XML.sub("", str(v))
    return v


def sanitize_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for c in out.columns:
        if out[c].dtype == object or str(out[c].dtype) in ("string", "str"):
            out[c] = out[c].map(_excel_safe)
        elif "datetime" in str(out[c].dtype) and getattr(
                out[c].dtype, "tz", None) is not None:
            out[c] = out[c].dt.tz_localize(None)
    return out


def write_output(folder, results, combined, analytics=None,
                 dedup=None) -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(folder, f"DSCIT_Consolidated_{stamp}.xlsx")
    summary = pd.DataFrame([{
        "File": r.file, "File Size": human_size(r.size),
        "Status": r.status, "Sheet": r.sheet,
        "Header Row": r.header_row, "Rows Extracted": r.rows,
        "Columns Matched": f"{r.matched}/{len(TARGET_COLUMNS)}",
        "Missing Columns": "; ".join(r.missing), "Notes": r.message,
    } for r in results])
    if dedup is not None and len(dedup.get("per_file", [])):
        pf = dedup["per_file"]
        dmap = dict(zip(pf["File"], pf["Duplicate Rows Removed"]))
        kmap = dict(zip(pf["File"], pf["Unique Rows Kept"]))
        summary.insert(6, "Duplicate Rows Removed",
                       summary["File"].map(dmap).fillna(0).astype(int))
        summary.insert(7, "Unique Rows Kept",
                       summary["File"].map(kmap).fillna(0).astype(int))

    XL_MAX_ROWS = 1_000_000            # headroom under Excel's 1,048,576
    xw = pd.ExcelWriter(out_path, engine="openpyxl")
    try:
        clean = sanitize_df(combined)
        if len(clean) <= XL_MAX_ROWS:
            clean.to_excel(xw, sheet_name="Consolidated Data", index=False)
        else:
            for i, start in enumerate(range(0, len(clean), XL_MAX_ROWS)):
                name = ("Consolidated Data" if i == 0
                        else f"Consolidated Data ({i + 1})")
                clean.iloc[start:start + XL_MAX_ROWS].to_excel(
                    xw, sheet_name=name, index=False)
        sanitize_df(summary).to_excel(
            xw, sheet_name="Run Summary", index=False)
        if analytics:
            pd.DataFrame(list(analytics["kpis"].items()),
                         columns=["Metric", "Value"]).to_excel(
                xw, sheet_name="Executive KPIs", index=False)
            sanitize_df(analytics["db_type"]).to_excel(xw, sheet_name="By Database Type", index=False)
            sanitize_df(analytics["db_inventory"]).to_excel(xw, sheet_name="DB Inventory", index=False)
            sanitize_df(analytics["segment"]).to_excel(xw, sheet_name="By Business Segment", index=False)
            sanitize_df(analytics["asset_owner"]).to_excel(xw, sheet_name="By Asset Owner", index=False)
            sanitize_df(analytics["dbs_by_output"]).to_excel(xw, sheet_name="DBs by Output", index=False)
            sanitize_df(analytics["dbs_by_owner"]).to_excel(xw, sheet_name="DBs by Output Owner",
                                               index=False)
            sanitize_df(analytics["dbs_by_srctype"]).to_excel(xw, sheet_name="DBs by Source Type",
                                                 index=False)
            sanitize_df(analytics["source_inventory"]).to_excel(xw, sheet_name="Source Inventory & Gaps",
                                                   index=False)
            sanitize_df(analytics["dq"]).to_excel(xw, sheet_name="Data Quality", index=False)
            if "norm_map" in analytics:
                sanitize_df(analytics["norm_map"]).to_excel(xw, sheet_name="DB Type Mapping",
                                               index=False)
            sanitize_df(analytics["schemas_by_db"]).to_excel(
                xw, sheet_name="Schemas by Database", index=False)
            sanitize_df(analytics["cde_inventory"]).to_excel(
                xw, sheet_name="CDE Inventory", index=False)
            sanitize_df(analytics["cde_by_segment"]).to_excel(
                xw, sheet_name="CDE Inventory", index=False,
                startrow=len(analytics["cde_inventory"]) + 3)
            sanitize_df(analytics["lineage_status"]).to_excel(
                xw, sheet_name="Lineage Coverage", index=False)
            sanitize_df(analytics["lineage_systems"]).to_excel(
                xw, sheet_name="Lineage Coverage", index=False,
                startrow=len(analytics["lineage_status"]) + 3)
            sanitize_df(analytics["harvest"]).to_excel(
                xw, sheet_name="Harvest Priority", index=False)
            sanitize_df(analytics["lineage_detail"].head(100000)).to_excel(
                xw, sheet_name="Lineage Detail", index=False)
        if dedup:
            metrics = pd.DataFrame([
                ("Files Found", dedup["files_found"]),
                ("Files Processed", dedup["files_processed"]),
                ("Identical Files Skipped (byte-for-byte)",
                 dedup["identical_files_skipped"]),
                ("Superseded Files Skipped (filename mode, optional)",
                 dedup["superseded_files_skipped"]),
                ("Rows Before Row-Level Dedup", dedup["rows_before_dedup"]),
                ("Duplicate Rows Removed", dedup["duplicate_rows_removed"]),
                ("Rows After Dedup", dedup["rows_after_dedup"]),
            ], columns=["Metric", "Value"])
            skipped = pd.DataFrame(
                dedup["skipped_detail"] or [("(none)", "")],
                columns=["Skipped File", "Reason"])
            metrics.to_excel(xw, sheet_name="Dedup Summary", index=False)
            skipped.to_excel(xw, sheet_name="Dedup Summary", index=False,
                             startrow=len(metrics) + 3)
            if len(dedup.get("per_file", [])):
                sanitize_df(dedup["per_file"]).to_excel(
                    xw, sheet_name="Content Overlap", index=False)
                if len(dedup.get("overlap", [])):
                    sanitize_df(dedup["overlap"]).to_excel(
                        xw, sheet_name="Content Overlap", index=False,
                        startrow=len(dedup["per_file"]) + 3)
        for ws in xw.book.worksheets:
            widths = {}
            for row in ws.iter_rows(min_row=1, max_row=200):
                for c in row:
                    if c.value is not None:
                        widths[c.column_letter] = max(
                            widths.get(c.column_letter, 10),
                            len(str(c.value)))
            for letter, wdt in widths.items():
                ws.column_dimensions[letter].width = min(wdt + 2, 48)
        xw.close()
    except Exception:
        # Ensure the save-on-close can't mask the real error with
        # "At least one sheet must be visible".
        try:
            if not xw.book.worksheets:
                xw.book.create_sheet("Sheet1")
            xw.close()
        except Exception:
            pass
        raise
    return out_path


# ============================================================================
# 6. EXECUTIVE HTML DASHBOARD (self-contained, SVG charts, TD theme)
# ============================================================================

def _dash_payload(combined: pd.DataFrame):
    """Compress row-level data for the dashboard.

    Rows are grouped into weighted combinations of 17 dimension fields
    (dictionary-encoded as integers), each carrying a row count, a count of
    rows with no Data Element, and the set of distinct Data Element ids.
    Dimension 12 is the canonicalized Data Element Indicator (CDE/BDE/
    Metric) and dimension 16 is the derived Lineage Status (Full/Partial/
    No Lineage across the five Data Flow Linkage fields)."""
    base = ["Output Name", "Output Owner", "Source Name", "Source Type",
            "Source Application EUC / MAL Code", "Database Type",
            "Database Name", "Database",
            "Business Segment / Corporate Function", "Asset Owner Name",
            "Technology Asset Owner Name", "Data Owner Name"]
    cols = base + ["Data Element Indicator", "Schema Name / File Path / API"] \
        + LINEAGE_COLUMNS + ["Data Element"]

    def cv(v):
        if v is None or (isinstance(v, float) and v != v):
            return None
        s = str(v).strip()
        return s if s and s.lower() not in (
            "na", "n/a", "none", "null", "nan", "tbd", "-") else None

    n_dims = 17
    lookups = [{} for _ in range(n_dims)]
    tables = [[] for _ in range(n_dims)]
    elem_map, elems = {}, []
    combos = {}

    def enc(j, v):
        if v is None:
            return -1
        d = lookups[j]
        idx = d.get(v)
        if idx is None:
            idx = d[v] = len(tables[j])
            tables[j].append(v)
        return idx

    df = combined.reindex(columns=cols)
    for tup in df.itertuples(index=False, name=None):
        key = [enc(j, cv(tup[j])) for j in range(12)]
        key.append(enc(12, classify_indicator(tup[12])))       # ind
        key.append(enc(13, cv(tup[13])))                       # src schema
        lvals = [cv(x) for x in tup[14:19]]                    # 5 lineage
        key.append(enc(14, lvals[0]))                          # lin system
        key.append(enc(15, lvals[1]))                          # lin database
        present = sum(x is not None for x in lvals)
        lstat = ("Full Lineage" if present == 5 else
                 "Partial Lineage" if present else "No Lineage")
        key.append(enc(16, lstat))                             # lin status
        e = cv(tup[19])
        k = tuple(key)
        rec = combos.get(k)
        if rec is None:
            rec = combos[k] = [0, 0, set()]
        rec[0] += 1
        if e is None:
            rec[1] += 1
        else:
            ei = elem_map.get(e)
            if ei is None:
                ei = elem_map[e] = len(elems)
                elems.append(e)
            rec[2].add(ei)

    packed = [list(k) + [c[0], c[1], sorted(c[2])]
              for k, c in combos.items()]
    return {"fields": tables, "elems": elems, "combos": packed}


DASH_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>DSCIT Tier-1 Executive Dashboard</title>
<style>
:root{
 --td:#008A00; --td-d:#006A00; --td-xd:#0B3D2E; --td-l:#54B848;
 --mist:#F3F7F3; --ink:#243024; --grey:#6b7f6b; --line:#E2EBE2;
 --amber:#B45309; --card:#ffffff;
}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Arial,sans-serif;background:var(--mist);
 color:var(--ink);overflow:hidden;height:100vh;display:flex;
 flex-direction:column}
.topbar{background:linear-gradient(120deg,var(--td-xd),var(--td) 65%,
 var(--td-l) 130%);color:#fff;padding:14px 26px;display:flex;
 align-items:center;gap:16px;flex-shrink:0;z-index:5;
 box-shadow:0 4px 16px rgba(11,61,46,.25)}
.tdmark{background:#fff;color:var(--td);font-weight:900;font-size:17px;
 width:40px;height:40px;display:flex;align-items:center;
 justify-content:center;border-radius:8px;flex-shrink:0}
.topbar h1{font-size:18px;font-weight:650;line-height:1.15}
.topbar .sub{font-size:11.5px;color:#CBE9CB}
.topbar .meta{margin-left:auto;text-align:right;font-size:11px;
 color:#B7E0B7;line-height:1.5}
.filters{background:#fff;border-bottom:1px solid var(--line);
 padding:10px 26px;display:flex;gap:10px;align-items:center;
 flex-wrap:wrap;flex-shrink:0}
.filters label{font-size:10px;font-weight:700;color:var(--grey);
 text-transform:uppercase;letter-spacing:.6px;display:block;
 margin-bottom:3px}
.fgroup{min-width:150px;flex:1;max-width:220px}
.filters select,.filters input[type=text]{width:100%;padding:7px 9px;
 border:1px solid #D5DDD5;border-radius:8px;font-size:12.5px;
 background:var(--mist);color:var(--ink);outline:none}
.filters select:focus,.filters input:focus{border-color:var(--td)}
.fbtn{background:var(--td);color:#fff;border:none;border-radius:8px;
 padding:9px 16px;font-size:12px;font-weight:650;cursor:pointer;
 align-self:flex-end}
.fbtn.ghost{background:#fff;color:var(--td);border:1px solid var(--td)}
.fbtn:hover{filter:brightness(.95)}
.chip{background:#EAF5EA;color:var(--td-d);border-radius:99px;
 padding:4px 12px;font-size:11px;font-weight:650;align-self:flex-end;
 margin-bottom:2px}
.main{display:flex;flex:1;min-height:0}
.side{width:230px;background:#fff;border-right:1px solid var(--line);
 padding:16px 12px;flex-shrink:0;overflow-y:auto}
.side .nav{display:flex;flex-direction:column;gap:4px}
.navbtn{display:flex;align-items:center;gap:10px;padding:11px 13px;
 border:none;background:none;border-radius:10px;font-size:13px;
 font-weight:600;color:#415441;cursor:pointer;text-align:left;width:100%;
 font-family:inherit}
.navbtn:hover{background:var(--mist)}
.navbtn.active{background:linear-gradient(120deg,var(--td),var(--td-l));
 color:#fff;box-shadow:0 4px 12px rgba(0,138,0,.28)}
.navbtn .ic{width:20px;text-align:center;font-size:14px}
.side .foot{margin-top:18px;font-size:10px;color:#9ab09a;
 padding:0 10px;line-height:1.6}
.content{flex:1;overflow-y:auto;padding:22px 28px 48px}
.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(158px,1fr));
 gap:12px;margin-bottom:18px}
.kpi{background:var(--card);border-radius:13px;padding:14px 16px;
 box-shadow:0 4px 14px rgba(11,61,46,.08);border-top:4px solid var(--td)}
.kpi b{display:block;font-size:24px;color:var(--td);font-weight:700}
.kpi span{font-size:10px;color:var(--grey);text-transform:uppercase;
 letter-spacing:.6px;font-weight:650}
.sec{background:var(--card);border-radius:15px;padding:20px 24px;
 margin-bottom:18px;box-shadow:0 4px 14px rgba(11,61,46,.07)}
.sec h2{font-size:15.5px;color:var(--td-xd);font-weight:650;display:flex;
 align-items:center;gap:9px;margin-bottom:4px}
.sec h2:before{content:'';width:7px;height:19px;background:var(--td);
 border-radius:4px}
.sec .note{font-size:12px;color:var(--grey);margin-bottom:14px}
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:22px}
@media(max-width:1100px){.grid2{grid-template-columns:1fr}}
.tbl{width:100%;border-collapse:collapse;font-size:12px}
.tbl th{background:#F0F6F0;color:var(--td-xd);text-align:left;
 padding:8px 9px;font-size:10px;text-transform:uppercase;
 letter-spacing:.5px;border-bottom:2px solid #DCE8DC;white-space:nowrap;
 position:sticky;top:0}
.tbl td{padding:7px 9px;border-bottom:1px solid #EDF3ED;vertical-align:top}
.tbl tr:hover td{background:#F7FBF7}
.tbl .num{text-align:right;font-weight:650;color:var(--td-xd)}
.tbl .gap{color:var(--amber);font-weight:650}
.tbl .ok{color:var(--td);font-weight:650}
.scrolltbl{max-height:430px;overflow-y:auto;border:1px solid var(--line);
 border-radius:10px}
.story{background:linear-gradient(135deg,#F0F8F0,#fff);
 border-left:5px solid var(--td);border-radius:12px;padding:16px 20px;
 margin-bottom:18px}
.story h3{color:var(--td-xd);font-size:13.5px;margin-bottom:7px}
.story li{font-size:13px;margin:5px 0 5px 18px;color:#37473a}
.donutwrap{display:flex;align-items:center;gap:20px;flex-wrap:wrap}
.legend{display:flex;flex-direction:column;gap:5px;font-size:12px}
.legend .lg{cursor:pointer}
.legend .lg:hover{text-decoration:underline}
.legend .lg i{display:inline-block;width:10px;height:10px;border-radius:3px;
 margin-right:6px}
.meter{background:#E7EFE7;border-radius:99px;height:11px;overflow:hidden;
 margin:5px 0 2px}
.meter i{display:block;height:100%;border-radius:99px;
 background:linear-gradient(90deg,var(--td),var(--td-l))}
.meter i.warn{background:linear-gradient(90deg,#D97706,#F59E0B)}
.meter i.bad{background:linear-gradient(90deg,#B91C1C,#EF4444)}
.dqrow{margin-bottom:11px;font-size:12px}
.dqrow .lbl{display:flex;justify-content:space-between;color:#37473a}
.empty{color:#9ab09a;font-size:13px;padding:22px;text-align:center}
.bar-click{cursor:pointer}
.hint{font-size:10.5px;color:#9ab09a;margin-top:6px}
</style></head><body>

<div class="topbar">
  <div class="tdmark">TD</div>
  <div><h1>DSCIT Tier-1 Source Landscape</h1>
    <div class="sub">Interactive Executive Analytics · Enterprise Data
    Management Office</div></div>
  <div class="meta">Generated __GENERATED__<br>
    Companion workbook: __EXCEL__<br>__DEDUP__</div>
</div>

<div class="filters">
  <div class="fgroup"><label>Business Segment</label>
    <select id="f_seg"></select></div>
  <div class="fgroup"><label>Database Type</label>
    <select id="f_dbt"></select></div>
  <div class="fgroup"><label>Output Owner</label>
    <select id="f_own"></select></div>
  <div class="fgroup"><label>Source Type</label>
    <select id="f_sty"></select></div>
  <div class="fgroup"><label>Asset Owner</label>
    <select id="f_ao"></select></div>
  <div class="fgroup"><label>Element Type</label>
    <select id="f_ind"></select></div>
  <div class="fgroup"><label>Search (output / source / element / MAL)</label>
    <input type="text" id="f_q" placeholder="Type to search…"></div>
  <button class="fbtn ghost" id="f_reset">Reset</button>
  <span class="chip" id="f_chip"></span>
</div>

<div class="main">
  <div class="side">
    <div class="nav" id="nav"></div>
    <div class="foot">TD Bank Group · EDMO<br>Tier-1 Report Analysis<br>
      Internal use only · Self-contained</div>
  </div>
  <div class="content" id="content"></div>
</div>

<script>
/* ==================== DATA (dictionary-encoded weighted combos) ========
   Each combo: [12 field indexes (-1 = missing), rowCount,
                rowsWithNoElement, [distinct element ids]]              */
const P = __PAYLOAD__;
const LOOK = P.fields, ELEMS = P.elems, COMBOS = P.combos;
const ELEMS_LC = ELEMS.map(s => s.toLowerCase());
const F = {out:0,owner:1,src:2,sty:3,mal:4,dbt:5,dbn:6,db:7,seg:8,ao:9,
           tao:10,downer:11,ind:12,sch:13,lsys:14,ldb:15,lstat:16};
const CNT = 17, NUL = 18, IDS = 19;
const FIELD_LABELS = [
 ["out","Output Name"],["owner","Output Owner"],["de","Data Element"],
 ["ind","Data Element Indicator (CDE/BDE/Metric)"],
 ["src","Source Name"],["sty","Source Type"],
 ["mal","Source Application EUC / MAL Code"],["dbt","Database Type"],
 ["dbn","Database Name"],["sch","Schema Name / File Path / API"],
 ["db","Database"],
 ["seg","Business Segment / Corporate Function"],
 ["ao","Asset Owner Name"],["tao","Technology Asset Owner Name"],
 ["downer","Data Owner Name"],
 ["lsys","Lineage System"],["ldb","Lineage Database"]];
const PALETTE = ["#008A00","#54B848","#1E5E3A","#8CC63F","#2E7D52",
 "#A7D28D","#0B3D2E","#66A182","#C7E5B5","#94b894"];

/* ==================== STATE ==================== */
const state = {tab:"overview", seg:"ALL", dbt:"ALL", own:"ALL",
               sty:"ALL", ao:"ALL", ind:"ALL", q:""};
const TABS = [
 ["overview","⌂","Executive Overview"],
 ["segment","▦","Business Segment"],
 ["mal","⌗","MAL Code Analysis"],
 ["owners","👤","Asset Owners"],
 ["tech","🗄","Technology Landscape"],
 ["lineage","⇄","Lineage & Harvest"],
 ["quality","✓","Data Quality"]];

/* ==================== CORE HELPERS ==================== */
const esc = s => s==null ? "—" :
  String(s).replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;");
const fmt = n => (n==null?0:n).toLocaleString("en-US");
function val(r, f){ const i = r[F[f]]; return i < 0 ? null : LOOK[F[f]][i]; }
function sumW(rows){ let s=0; for(const r of rows) s += r[CNT]; return s; }
function uniqF(rows, f){
  const idx = F[f], s = new Set();
  for(const r of rows) if(r[idx] >= 0) s.add(r[idx]);
  return s.size;
}
function uniqDE(rows){
  const s = new Set();
  for(const r of rows) for(const id of r[IDS]) s.add(id);
  return s.size;
}
function missing(rows, f){
  if(f === "de"){ let s=0; for(const r of rows) s += r[NUL]; return s; }
  const idx = F[f]; let s=0;
  for(const r of rows) if(r[idx] < 0) s += r[CNT];
  return s;
}
function firstVal(rows, f){
  for(const r of rows){ const v = val(r, f); if(v != null) return v; }
  return "—";
}
function groupBy(rows, f, blank){
  const m = new Map();
  for(const r of rows){
    const k = val(r, f) ?? (blank || "Not Specified");
    let a = m.get(k); if(!a){ a = []; m.set(k, a); }
    a.push(r);
  }
  return m;
}
const topN = (m, fn, n) => [...m.entries()]
  .map(([k,v]) => [k, fn(v), v]).sort((a,b)=>b[1]-a[1]).slice(0, n||9999);
const TOTAL_ROWS = sumW(COMBOS);

function filtered(){
  const q = state.q.toLowerCase();
  let matchIds = null;
  if(q){
    matchIds = new Set();
    for(let i=0;i<ELEMS_LC.length;i++)
      if(ELEMS_LC[i].includes(q)) matchIds.add(i);
  }
  const pass = (r,f,sel) => sel==="ALL" ||
    (val(r,f) ?? "Not Specified") === sel;
  return COMBOS.filter(r =>
    pass(r,"seg",state.seg) && pass(r,"dbt",state.dbt) &&
    pass(r,"owner",state.own) && pass(r,"sty",state.sty) &&
    pass(r,"ao",state.ao) && pass(r,"ind",state.ind) &&
    (!q ||
      ["out","src","mal","dbn"].some(f =>
        (val(r,f)||"").toLowerCase().includes(q)) ||
      r[IDS].some(id => matchIds.has(id))));
}

/* ==================== CHARTS (SVG) ==================== */
function hbar(pairs, opts){
  opts = opts||{};
  pairs = pairs.filter(p=>p[1]>0).slice(0, opts.max||12);
  if(!pairs.length) return "<div class='empty'>No data for current filters</div>";
  const mx = Math.max(...pairs.map(p=>p[1]));
  const W = 560, LW = 195, RH = 28, GAP = 9;
  const H = pairs.length*(RH+GAP);
  let s = `<svg viewBox="0 0 ${W} ${H}" width="100%"
    xmlns="http://www.w3.org/2000/svg">`;
  pairs.forEach((p,i)=>{
    const y = i*(RH+GAP), bw = Math.max(4,(W-LW-70)*p[1]/mx);
    const c = i===0 ? "#0B3D2E" : "#008A00";
    const click = opts.click ?
      ` class="bar-click" onclick="${opts.click}('${String(p[0])
        .replace(/'/g,"\\'")}')"` : "";
    s += `<g${click}><text x="${LW-8}" y="${y+RH/2+4}" text-anchor="end"
      font-size="11.5" fill="#3c4a3c"
      font-family="Segoe UI,Arial">${esc(String(p[0]).slice(0,36))}</text>
      <rect x="${LW}" y="${y+4}" rx="6" width="${bw.toFixed(1)}"
      height="${RH-8}" fill="${c}" opacity=".92"/>
      <text x="${(LW+bw+8).toFixed(1)}" y="${y+RH/2+4}" font-size="11.5"
      font-weight="650" fill="#0B3D2E"
      font-family="Segoe UI,Arial">${fmt(p[1])}</text></g>`;
  });
  return s + "</svg>" +
    (opts.click ? "<div class='hint'>Tip: click a bar to filter the whole dashboard.</div>" : "");
}
function donut(pairs, unitLabel, clickFn){
  pairs = pairs.filter(p=>p[1]>0);
  if(!pairs.length) return "<div class='empty'>No data for current filters</div>";
  if(pairs.length>8){
    const head = pairs.slice(0,7);
    head.push(["Other", pairs.slice(7).reduce((a,p)=>a+p[1],0)]);
    pairs = head;
  }
  const total = pairs.reduce((a,p)=>a+p[1],0);
  const SZ=210, R=72, SW=30, CIRC=2*Math.PI*R;
  let off = CIRC*0.25;
  let s = `<svg viewBox="0 0 ${SZ} ${SZ}" width="${SZ}"
    xmlns="http://www.w3.org/2000/svg">`;
  pairs.forEach((p,i)=>{
    const dash = p[1]/total*CIRC;
    s += `<circle cx="${SZ/2}" cy="${SZ/2}" r="${R}" fill="none"
      stroke="${PALETTE[i%PALETTE.length]}" stroke-width="${SW}"
      stroke-dasharray="${dash.toFixed(2)} ${(CIRC-dash).toFixed(2)}"
      stroke-dashoffset="${off.toFixed(2)}"/>`;
    off -= dash;
  });
  s += `<text x="${SZ/2}" y="${SZ/2-3}" text-anchor="middle" font-size="24"
    font-weight="700" fill="#0B3D2E"
    font-family="Segoe UI,Arial">${fmt(total)}</text>
    <text x="${SZ/2}" y="${SZ/2+17}" text-anchor="middle" font-size="10.5"
    fill="#6b7f6b" font-family="Segoe UI,Arial">${unitLabel}</text></svg>`;
  const legend = pairs.map((p,i)=>{
    const click = clickFn && p[0]!=="Other" ?
      ` onclick="${clickFn}('${String(p[0]).replace(/'/g,"\\'")}')"` : "";
    return `<span class="lg"${click}><i style="background:${
      PALETTE[i%PALETTE.length]}"></i>${esc(p[0])} <b>${fmt(p[1])}</b></span>`;
  }).join("");
  return `<div class="donutwrap">${s}<div class="legend">${legend}</div></div>`;
}
function table(headers, rows, opts){
  opts = opts||{};
  if(!rows.length) return "<div class='empty'>No data for current filters</div>";
  const th = headers.map(h=>`<th>${esc(h)}</th>`).join("");
  const body = rows.slice(0, opts.max||15).map(r=>"<tr>"+r.map((v,i)=>{
    if (typeof v==="number") return `<td class="num">${fmt(v)}</td>`;
    if (opts.flagCol===i){
      const cls = v==="Complete" ? "ok" : "gap";
      return `<td class="${cls}">${esc(v)}</td>`;
    }
    return `<td>${esc(v)}</td>`;
  }).join("")+"</tr>").join("");
  const t = `<table class="tbl"><thead><tr>${th}</tr></thead>
    <tbody>${body}</tbody></table>`;
  return opts.scroll ? `<div class="scrolltbl">${t}</div>` : t;
}
const kpiCards = items => `<div class="kpis">${items.map(([l,v])=>
  `<div class="kpi"><b>${fmt(v)}</b><span>${esc(l)}</span></div>`)
  .join("")}</div>`;

function setSeg(v){ state.seg=v; syncSelects(); render(); }
function setDbt(v){ state.dbt=v; syncSelects(); render(); }
function setAo(v){ state.ao=v; syncSelects(); render(); }

/* ==================== TAB RENDERERS ==================== */
function tabOverview(rows){
  const segTop = topN(groupBy(rows,"seg"), v=>uniqF(v,"out"));
  const dbtTop = topN(groupBy(rows,"dbt"), v=>sumW(v));
  const malMissing = missing(rows,"mal");
  const story = [];
  story.push(`The filtered view spans <b>${fmt(uniqF(rows,"out"))} outputs</b>
    owned by <b>${fmt(uniqF(rows,"owner"))} output owners</b>, drawing on
    <b>${fmt(uniqF(rows,"src"))} source systems</b> across
    <b>${fmt(uniqF(rows,"mal"))} EUC/MAL-coded applications</b>.`);
  if(dbtTop.length) story.push(`<b>${esc(dbtTop[0][0])}</b> is the dominant
    platform with ${fmt(uniqF(dbtTop[0][2],"dbn"))} databases behind
    ${fmt(uniqF(dbtTop[0][2],"out"))} outputs.`);
  if(segTop.length) story.push(`<b>${esc(segTop[0][0])}</b> carries the largest
    reporting footprint (${fmt(segTop[0][1])} outputs).`);
  if(malMissing) story.push(`<b>${fmt(malMissing)} rows</b> are missing an
    EUC/MAL code — see the Data Quality tab for the remediation list.`);
  const noLin = sumW(rows.filter(r=>val(r,"lstat")==="No Lineage"));
  if(noLin) story.push(`<b>${fmt(noLin)} rows</b> carry no Data Flow Linkage
    information — the Lineage &amp; Harvest tab ranks which databases to
    harvest first to close the gap.`);
  return `
  ${kpiCards([["Outputs",uniqF(rows,"out")],
    ["Output Owners",uniqF(rows,"owner")],
    ["Source Systems",uniqF(rows,"src")],
    ["Source Types",uniqF(rows,"sty")],
    ["EUC / MAL Codes",uniqF(rows,"mal")],
    ["Databases",uniqF(rows,"dbn")],
    ["Database Types",uniqF(rows,"dbt")],
    ["Data Elements",uniqDE(rows)],
    ["Segments / Functions",uniqF(rows,"seg")],
    ["Asset Owners",uniqF(rows,"ao")],
    ["Data Owners",uniqF(rows,"downer")],
    ["CDEs",uniqDE(rows.filter(r=>val(r,"ind")==="CDE"))],
    ["Schemas",uniqF(rows,"sch")],
    ["Lineage Coverage %", sumW(rows) ? Math.round(1000*
      (sumW(rows)-sumW(rows.filter(r=>val(r,"lstat")==="No Lineage")))
      /sumW(rows))/10 : 0],
    ["Data Element Rows",sumW(rows)]])}
  <div class="story"><h3>Executive Summary</h3>
    <ul>${story.map(s=>`<li>${s}</li>`).join("")}</ul></div>
  <div class="grid2">
    <div class="sec"><h2>Rows by Database Type</h2>
      <div class="note">Normalized platform view — click a legend item to
      filter.</div>${donut(dbtTop, "rows", "setDbt")}</div>
    <div class="sec"><h2>Outputs by Business Segment</h2>
      <div class="note">Distinct outputs per segment / corporate
      function.</div>${hbar(segTop, {click:"setSeg"})}</div>
  </div>
  <div class="sec"><h2>Top Outputs by Data-Element Footprint</h2>
    <div class="note">Concentration of Tier-1 data elements.</div>
    ${table(["Output Name","Output Owner","Data Elements","Sources","Databases"],
      topN(groupBy(rows,"out","Not Specified"), v=>uniqDE(v), 12)
      .map(([k,n,v])=>[k, firstVal(v,"owner"), n, uniqF(v,"src"),
        uniqF(v,"dbn")]))}</div>`;
}

function tabSegment(rows){
  const m = groupBy(rows,"seg");
  const t = topN(m, v=>uniqF(v,"out"));
  return `
  ${kpiCards([["Segments / Functions",m.size],
    ["Outputs",uniqF(rows,"out")],
    ["Output Owners",uniqF(rows,"owner")],
    ["Data Elements",uniqDE(rows)]])}
  <div class="grid2">
    <div class="sec"><h2>Outputs by Segment</h2>
      <div class="note">Click a bar to focus the dashboard on one
      segment.</div>${hbar(t,{click:"setSeg"})}</div>
    <div class="sec"><h2>Segment Detail</h2>
      <div class="note">Owners, elements, sources and databases per
      segment.</div>
      ${table(["Segment / Function","Outputs","Owners","Elements","Sources",
        "Databases"], t.map(([k,n,v])=>[k,n,uniqF(v,"owner"),
        uniqDE(v),uniqF(v,"src"),uniqF(v,"dbn")]),
        {scroll:true,max:100})}</div>
  </div>
  <div class="sec"><h2>Segment × Database Type Mix</h2>
    <div class="note">Which platforms each segment depends on.</div>
    ${table(["Segment / Function","Database Types Used","Top Platform",
      "Rows"], t.map(([k,n,v])=>{
        const dm = topN(groupBy(v,"dbt"), x=>sumW(x));
        return [k, dm.length, dm.length?dm[0][0]+" ("+fmt(dm[0][1])+")":"—",
          sumW(v)];}), {scroll:true,max:100})}</div>`;
}

function tabMal(rows){
  const m = groupBy(rows,"mal","«Code Missing»");
  const total = sumW(rows);
  const noCode = missing(rows,"mal");
  const pct = total ? Math.round(100*(total-noCode)/total) : 0;
  const t = topN(m, v=>sumW(v));
  return `
  ${kpiCards([["Distinct EUC / MAL Codes",uniqF(rows,"mal")],
    ["Rows with a Code",total-noCode],
    ["Rows Missing a Code",noCode],
    ["Code Coverage %",pct]])}
  <div class="sec"><h2>MAL / EUC Code Inventory</h2>
    <div class="note">Every application code with its sources, outputs and
    database footprint. «Code Missing» groups the uncoded rows —
    a key remediation queue for application-inventory alignment.</div>
    ${table(["EUC / MAL Code","Sources","Source Types","Outputs","Databases",
      "Rows"], t.map(([k,n,v])=>[k,uniqF(v,"src"),uniqF(v,"sty"),
      uniqF(v,"out"),uniqF(v,"dbn"),n]), {scroll:true,max:300})}</div>
  <div class="grid2">
    <div class="sec"><h2>Rows per Code (Top 12)</h2>
      ${hbar(t.slice(0,12))}</div>
    <div class="sec"><h2>Sources Missing a Code</h2>
      <div class="note">Source systems appearing without any EUC/MAL
      code.</div>
      ${table(["Source Name","Source Type","Rows"],
        topN(groupBy(rows.filter(r=>r[F.mal]<0), "src",
        "«Source Missing»"), v=>sumW(v))
        .map(([k,n,v])=>[k, firstVal(v,"sty"), n]),
        {scroll:true,max:100})}</div>
  </div>`;
}

function tabOwners(rows){
  const ao = topN(groupBy(rows,"ao"), v=>uniqF(v,"dbn"));
  const dow = topN(groupBy(rows,"downer"), v=>sumW(v));
  return `
  ${kpiCards([["Asset Owners",uniqF(rows,"ao")],
    ["Technology Asset Owners",uniqF(rows,"tao")],
    ["Data Owners",uniqF(rows,"downer")],
    ["Rows Missing Asset Owner",missing(rows,"ao")],
    ["Rows Missing Data Owner",missing(rows,"downer")]])}
  <div class="grid2">
    <div class="sec"><h2>Databases by Asset Owner</h2>
      <div class="note">Custody concentration — click to filter.</div>
      ${hbar(ao,{click:"setAo"})}</div>
    <div class="sec"><h2>Asset Owner Detail</h2>
      ${table(["Asset Owner","Databases","DB Types","Outputs","Tech Owners"],
      ao.map(([k,n,v])=>[k,n,uniqF(v,"dbt"),uniqF(v,"out"),
      uniqF(v,"tao")]), {scroll:true,max:200})}</div>
  </div>
  <div class="grid2">
    <div class="sec"><h2>Data Owner Coverage</h2>
      <div class="note">Rows accountable to each data owner.</div>
      ${table(["Data Owner","Rows","Outputs","Elements"],
        dow.map(([k,n,v])=>[k,n,uniqF(v,"out"),uniqDE(v)]),
        {scroll:true,max:200})}</div>
    <div class="sec"><h2>Technology Asset Owners</h2>
      ${table(["Technology Asset Owner","Databases","DB Types","Rows"],
        topN(groupBy(rows,"tao"), v=>uniqF(v,"dbn"))
        .map(([k,n,v])=>[k,n,uniqF(v,"dbt"),sumW(v)]),
        {scroll:true,max:200})}</div>
  </div>`;
}

function tabTech(rows){
  const m = groupBy(rows,"dbt");
  const t = topN(m, v=>sumW(v));
  const inv = [];
  for (const [dt, v] of m){
    for (const [dbn, vv] of groupBy(v,"dbn","«Name Missing»"))
      inv.push([dt, dbn, uniqF(vv,"out"), uniqF(vv,"src"), sumW(vv)]);
  }
  inv.sort((a,b)=> a[0]===b[0] ? b[4]-a[4] : (a[0]<b[0]?-1:1));
  const sty = topN(groupBy(rows,"sty"), v=>uniqF(v,"dbn"));
  return `
  ${kpiCards([["Database Types",m.size],
    ["Databases",uniqF(rows,"dbn")],
    ["Schemas",uniqF(rows,"sch")],
    ["Source Systems",uniqF(rows,"src")],
    ["Source Types",uniqF(rows,"sty")],
    ["Rows Missing DB Name",missing(rows,"dbn")]])}
  <div class="grid2">
    <div class="sec"><h2>Platform Mix</h2>
      <div class="note">Normalized Database Types (ORACLE / Oracle 19c /
      Essbase → Oracle, etc.). Click a legend item to filter.</div>
      ${donut(t,"rows","setDbt")}</div>
    <div class="sec"><h2>Distinct Databases per Platform</h2>
      ${hbar(topN(m, v=>uniqF(v,"dbn")),{click:"setDbt"})}</div>
  </div>
  <div class="sec"><h2>Database Inventory (Name × Type)</h2>
    <div class="note">Every database with its platform, outputs and
    sources.</div>
    ${table(["Database Type","Database Name","Outputs","Sources","Rows"],
      inv, {scroll:true,max:500})}</div>
  <div class="sec"><h2>Schemas per Database</h2>
    <div class="note">Distinct Schema Name / File Path / API values behind
    each source database.</div>
    ${table(["Database Name","Schemas","Elements","Outputs","Rows"],
      topN(groupBy(rows,"dbn","«Name Missing»"), v=>uniqF(v,"sch"))
      .map(([k,n,v])=>[k,n,uniqDE(v),uniqF(v,"out"),sumW(v)]),
      {scroll:true,max:300})}</div>
  <div class="sec"><h2>Databases by Source Type</h2>
    ${table(["Source Type","Databases","Sources","Rows"],
      sty.map(([k,n,v])=>[k,n,uniqF(v,"src"),sumW(v)]),
      {scroll:true,max:100})}</div>`;
}

function tabQuality(rows){
  const total = sumW(rows);
  const meters = FIELD_LABELS.map(([key,label])=>{
    const miss = missing(rows, key);
    const pct = total ? Math.round(1000*(1-miss/total))/10 : 0;
    const cls = pct>=95?"":(pct>=75?"warn":"bad");
    return `<div class="dqrow"><div class="lbl"><span>${esc(label)}</span>
      <b>${pct}%</b></div><div class="meter"><i class="${cls}"
      style="width:${pct}%"></i></div>
      <div style="font-size:10.5px;color:#9ab09a">${fmt(miss)} missing
      rows</div></div>`;
  }).join("");
  const keyF = ["out","src","dbn","downer","ao","mal"];
  const score = total ? Math.round(10*keyF.reduce((a,k)=>
    a+100*(1-missing(rows,k)/total),0)/keyF.length)/10 : 0;
  const srcMap = new Map();
  for (const r of rows){
    const k = [val(r,"src")||"«Source Missing»",
               val(r,"mal")||"«Code Missing»",
               val(r,"sty")||"«Type Missing»"].join("|");
    let a = srcMap.get(k); if(!a){ a=[]; srcMap.set(k,a); }
    a.push(r);
  }
  const gapRows = [...srcMap.entries()].map(([k,v])=>{
    const [src,mal,sty] = k.split("|");
    const flags = [];
    if(src==="«Source Missing»") flags.push("Source missing");
    if(mal==="«Code Missing»") flags.push("EUC/MAL missing");
    if(v.some(r=>r[F.dbn]<0)) flags.push("DB name gaps");
    if(v.some(r=>r[F.downer]<0)) flags.push("Data owner gaps");
    if(v.some(r=>r[F.ao]<0)) flags.push("Asset owner gaps");
    return [src,mal,sty,uniqF(v,"dbn"),sumW(v),
            flags.length?flags.join("; "):"Complete"];
  }).sort((a,b)=> (a[5]==="Complete") - (b[5]==="Complete") || b[4]-a[4]);
  const gapCt = gapRows.filter(r=>r[5]!=="Complete").length;
  return `
  ${kpiCards([["Key-Field Completeness %",score],
    ["Source Lines with Gaps",gapCt],
    ["Rows Missing DB Name",missing(rows,"dbn")],
    ["Rows Missing Data Owner",missing(rows,"downer")],
    ["Rows Missing MAL Code",missing(rows,"mal")]])}
  <div class="grid2">
    <div class="sec"><h2>Field Completeness</h2>
      <div class="note">Populated % per extracted field, under current
      filters. Green ≥95 · amber ≥75 · red &lt;75.</div>${meters}</div>
    <div class="sec"><h2>Source Gap Register</h2>
      <div class="note">Each source line with its flags — amber rows are the
      remediation queue.</div>
      ${table(["Source Name","EUC / MAL Code","Source Type","Databases",
        "Rows","Gap Flags"], gapRows, {scroll:true,max:500,flagCol:5})}</div>
  </div>`;
}

function tabLineage(rows){
  const total = sumW(rows);
  const st = groupBy(rows, "lstat", "No Lineage");
  const full = sumW(st.get("Full Lineage")||[]);
  const part = sumW(st.get("Partial Lineage")||[]);
  const none = sumW(st.get("No Lineage")||[]);
  const cov = total ? Math.round(1000*(full+part)/total)/10 : 0;
  const allE = uniqDE(rows);
  const withL = uniqDE(rows.filter(r=>val(r,"lstat")!=="No Lineage"));
  const sys = topN(groupBy(rows,"lsys","«No System»"), v=>sumW(v));

  const harvest = topN(groupBy(rows,"dbn","«Name Missing»"), v=>{
      return uniqDE(v) - uniqDE(v.filter(r=>val(r,"lstat")!=="No Lineage"));
    })
    .map(([k,noL,v])=>{
      const el = uniqDE(v);
      const cde = uniqDE(v.filter(r=>val(r,"ind")==="CDE"));
      const cdeNoL = cde - uniqDE(v.filter(r=>val(r,"ind")==="CDE"
        && val(r,"lstat")!=="No Lineage"));
      const covd = el ? Math.round(100*(el-noL)/el) : 0;
      return [k, el, cde, noL, cdeNoL, uniqF(v,"sch"), covd, sumW(v)];
    })
    .sort((a,b)=> b[4]-a[4] || b[3]-a[3]);

  const cross = [];
  for (const [s,v] of groupBy(rows.filter(r=>r[F.lsys]>=0), "lsys"))
    for (const [d,vv] of groupBy(v, "dbn", "«Name Missing»"))
      cross.push([s, d, uniqDE(vv), sumW(vv)]);
  cross.sort((a,b)=>b[3]-a[3]);

  return `
  ${kpiCards([["Lineage Coverage %",cov],
    ["Rows — Full Lineage",full],
    ["Rows — Partial Lineage",part],
    ["Rows — No Lineage",none],
    ["Elements with Lineage",withL],
    ["Elements without Lineage",allE-withL],
    ["Lineage Systems",uniqF(rows,"lsys")],
    ["Lineage Databases",uniqF(rows,"ldb")]])}
  <div class="grid2">
    <div class="sec"><h2>Lineage Population Status</h2>
      <div class="note">Full = all five Data Flow Linkage fields populated
      (System, Database, Schema, Physical Table, Physical Column);
      Partial = some; No Lineage = none.</div>
      ${donut([["Full Lineage",full],["Partial Lineage",part],
        ["No Lineage",none]],"rows")}</div>
    <div class="sec"><h2>Rows by Lineage System</h2>
      <div class="note">Where lineage says the data lands (System field on
      the Data Flow Linkage side).</div>${hbar(sys)}</div>
  </div>
  <div class="sec"><h2>Metadata Harvest Priority</h2>
    <div class="note">Ranked by CDEs without lineage, then elements without
    lineage — the databases where harvesting metadata closes the biggest
    gaps first. An element counts as having lineage if any of its rows for
    that database carries lineage information.</div>
    ${table(["Source Database","Elements","CDEs","Elements w/o Lineage",
      "CDEs w/o Lineage","Schemas","Lineage Coverage %","Rows"],
      harvest, {scroll:true,max:300})}</div>
  <div class="sec"><h2>Source Database → Lineage System Map</h2>
    <div class="note">How source-side databases connect to lineage-side
    landing systems.</div>
    ${table(["Lineage System","Source Database","Elements","Rows"],
      cross, {scroll:true,max:300})}</div>`;
}

/* ==================== SHELL ==================== */
const RENDERERS = {overview:tabOverview, segment:tabSegment, mal:tabMal,
                   owners:tabOwners, tech:tabTech, lineage:tabLineage,
                   quality:tabQuality};

function buildNav(){
  document.getElementById("nav").innerHTML = TABS.map(([id,ic,label])=>
    `<button class="navbtn${state.tab===id?" active":""}"
     onclick="state.tab='${id}';buildNav();render()">
     <span class="ic">${ic}</span>${label}</button>`).join("");
}
function fillSelect(id, f, label){
  const vals = new Set(LOOK[F[f]]);
  if (COMBOS.some(r=>r[F[f]]<0)) vals.add("Not Specified");
  const arr = [...vals].sort((a,b)=>a.localeCompare(b));
  document.getElementById(id).innerHTML =
    `<option value="ALL">All ${label} (${arr.length})</option>` +
    arr.map(v=>`<option>${esc(v)}</option>`).join("");
}
function syncSelects(){
  f_seg.value=state.seg; f_dbt.value=state.dbt; f_own.value=state.own;
  f_sty.value=state.sty; f_ao.value=state.ao; f_ind.value=state.ind;
  f_q.value=state.q;
}
function render(){
  const rows = filtered();
  const active = ["seg","dbt","own","sty","ao","ind"]
    .filter(k=>state[k]!=="ALL").length + (state.q?1:0);
  document.getElementById("f_chip").textContent =
    `${fmt(sumW(rows))} of ${fmt(TOTAL_ROWS)} rows` +
    (active? ` · ${active} filter${active>1?"s":""}` : "");
  document.getElementById("content").innerHTML =
    RENDERERS[state.tab](rows);
  document.querySelector(".content").scrollTop = 0;
}
window.addEventListener("DOMContentLoaded", ()=>{
  fillSelect("f_seg","seg","Segments");
  fillSelect("f_dbt","dbt","Types");
  fillSelect("f_own","owner","Owners");
  fillSelect("f_sty","sty","Types");
  fillSelect("f_ao","ao","Owners");
  fillSelect("f_ind","ind","Types");
  const bind = (id,key)=>document.getElementById(id)
    .addEventListener("change", e=>{state[key]=e.target.value; render();});
  bind("f_seg","seg"); bind("f_dbt","dbt"); bind("f_own","own");
  bind("f_sty","sty"); bind("f_ao","ao"); bind("f_ind","ind");
  let deb;
  document.getElementById("f_q").addEventListener("input", e=>{
    clearTimeout(deb);
    deb = setTimeout(()=>{state.q=e.target.value; render();}, 220);
  });
  document.getElementById("f_reset").addEventListener("click", ()=>{
    Object.assign(state,{seg:"ALL",dbt:"ALL",own:"ALL",sty:"ALL",
      ao:"ALL",ind:"ALL",q:""});
    syncSelects(); render();
  });
  buildNav(); render();
});
</script></body></html>"""


def generate_dashboard(folder, analytics, combined, excel_path,
                       dedup=None) -> str:
    stamp = datetime.now()
    out_path = os.path.join(
        folder,
        f"DSCIT_Executive_Dashboard_{stamp.strftime('%Y%m%d_%H%M%S')}.html")
    payload = _dash_payload(combined)
    data_json = json.dumps(payload, ensure_ascii=False,
                           separators=(",", ":")).replace("</", "<\\/")
    if dedup:
        dd = (f"De-duplicated: {dedup['duplicate_rows_removed']:,} duplicate "
              f"rows removed · "
              f"{dedup['identical_files_skipped'] + dedup['superseded_files_skipped']}"
              f" duplicate/superseded file(s) skipped")
    else:
        dd = "De-duplication: not applied"
    doc = (DASH_TEMPLATE
           .replace("__GENERATED__",
                    stamp.strftime("%B %d, %Y at %H:%M"))
           .replace("__EXCEL__",
                    html.escape(os.path.basename(excel_path or "—")))
           .replace("__DEDUP__", html.escape(dd))
           .replace("__PAYLOAD__", data_json))
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(doc)
    return out_path


# ============================================================================
# 7. TD-THEMED UI
# ============================================================================

TD_GREEN = "#008A00"
TD_GREEN_DARK = "#006A00"
TD_GREEN_LIGHT = "#54B848"
TD_MIST = "#F3F7F3"
CHARCOAL = "#2E2E2E"
GREY = "#6B7280"
WHITE = "#FFFFFF"
AMBER = "#B45309"
RED = "#B91C1C"


def launch_ui():
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox

    root = tk.Tk()
    root.title("TD  |  DSCIT Tier-1 Report Consolidator & Executive Analytics")
    root.geometry("1160x790")
    root.configure(bg=TD_MIST)
    root.minsize(980, 660)

    ui_queue = queue.Queue()
    state = {"running": False, "folder": tk.StringVar(value=""),
             "excel": None, "html": None,
             "results": None, "combined": None, "dedup": None,
             "keep_latest": tk.BooleanVar(value=False)}

    style = ttk.Style(root)
    style.theme_use("clam")
    style.configure("TFrame", background=TD_MIST)
    style.configure("Header.TFrame", background=TD_GREEN)
    style.configure("TLabel", background=TD_MIST, foreground=CHARCOAL,
                    font=("Segoe UI", 10))
    style.configure("StatValue.TLabel", background=WHITE, foreground=TD_GREEN,
                    font=("Segoe UI Semibold", 22))
    style.configure("StatCap.TLabel", background=WHITE, foreground=GREY,
                    font=("Segoe UI", 9))
    style.configure("HeaderTitle.TLabel", background=TD_GREEN, foreground=WHITE,
                    font=("Segoe UI Semibold", 17))
    style.configure("HeaderSub.TLabel", background=TD_GREEN, foreground="#D8F0D8",
                    font=("Segoe UI", 10))
    style.configure("TD.TButton", background=TD_GREEN, foreground=WHITE,
                    font=("Segoe UI Semibold", 10), padding=(16, 8),
                    borderwidth=0)
    style.map("TD.TButton",
              background=[("active", TD_GREEN_DARK), ("disabled", "#9CC79C")])
    style.configure("Gold.TButton", background="#0B3D2E", foreground=WHITE,
                    font=("Segoe UI Semibold", 10), padding=(16, 8),
                    borderwidth=0)
    style.map("Gold.TButton",
              background=[("active", "#092E23"), ("disabled", "#7C948A")])
    style.configure("Ghost.TButton", background=WHITE, foreground=TD_GREEN,
                    font=("Segoe UI", 10), padding=(12, 7), borderwidth=1)
    style.map("Ghost.TButton", background=[("active", TD_MIST)])
    style.configure("TD.Horizontal.TProgressbar", troughcolor="#E3EAE3",
                    background=TD_GREEN_LIGHT, thickness=14, borderwidth=0)
    style.configure("Treeview", font=("Segoe UI", 9), rowheight=26,
                    background=WHITE, fieldbackground=WHITE, foreground=CHARCOAL)
    style.configure("Treeview.Heading", font=("Segoe UI Semibold", 9),
                    background=TD_MIST, foreground=CHARCOAL)

    # header
    header = ttk.Frame(root, style="Header.TFrame")
    header.pack(fill="x")
    hin = ttk.Frame(header, style="Header.TFrame")
    hin.pack(fill="x", padx=24, pady=16)
    logo = tk.Canvas(hin, width=46, height=46, bg=TD_GREEN, highlightthickness=0)
    logo.create_rectangle(2, 2, 44, 44, fill=WHITE, outline=WHITE)
    logo.create_text(23, 23, text="TD", fill=TD_GREEN,
                     font=("Segoe UI Black", 16, "bold"))
    logo.pack(side="left", padx=(0, 14))
    tbox = ttk.Frame(hin, style="Header.TFrame")
    tbox.pack(side="left")
    ttk.Label(tbox, text="DSCIT Tier-1 Consolidator & Executive Analytics",
              style="HeaderTitle.TLabel").pack(anchor="w")
    ttk.Label(tbox, text="Enterprise Data Management Office  ·  Tier-1 Sources",
              style="HeaderSub.TLabel").pack(anchor="w")

    body = ttk.Frame(root)
    body.pack(fill="both", expand=True, padx=24, pady=18)

    # folder card
    fcard = tk.Frame(body, bg=WHITE, highlightbackground="#E2E8E2",
                     highlightthickness=1)
    fcard.pack(fill="x")
    frow = tk.Frame(fcard, bg=WHITE)
    frow.pack(fill="x", padx=18, pady=14)
    tk.Label(frow, text="Report Folder", bg=WHITE, fg=GREY,
             font=("Segoe UI Semibold", 9)).pack(anchor="w")
    inner = tk.Frame(frow, bg=WHITE)
    inner.pack(fill="x", pady=(6, 0))
    entry = tk.Entry(inner, textvariable=state["folder"], font=("Segoe UI", 10),
                     bg=TD_MIST, fg=CHARCOAL, relief="flat",
                     highlightthickness=1, highlightbackground="#D5DDD5")
    entry.pack(side="left", fill="x", expand=True, ipady=7, padx=(0, 10))

    stat_vars = {}
    status_var = tk.StringVar(
        value="Select your Tier-1 Report Analysis folder to begin.")

    def browse():
        d = filedialog.askdirectory(
            title="Select the Tier-1 Report Analysis folder")
        if d:
            state["folder"].set(d)
            n = len(discover_files(d))
            stat_vars["found"].set(str(n))
            status_var.set(f"{n} DSCIT file(s) detected — ready to run."
                           if n else "No DSCIT files detected in this folder.")

    ttk.Button(inner, text="Browse…", style="Ghost.TButton",
               command=browse).pack(side="left", padx=(0, 8))
    run_btn = ttk.Button(inner, text="▶  Run Consolidation", style="TD.TButton")
    run_btn.pack(side="left", padx=(0, 8))
    dash_btn = ttk.Button(inner, text="★  Generate Executive Dashboard",
                          style="Gold.TButton", state="disabled")
    dash_btn.pack(side="left")
    dedup_note = tk.Label(
        frow, text="De-duplication is content-based and always on: "
        "byte-identical files are skipped, and rows duplicated across or "
        "within files are removed. Similar file names are NOT treated as "
        "duplicates.", bg=WHITE, fg=GREY, font=("Segoe UI", 9),
        anchor="w", justify="left", wraplength=980)
    dedup_note.pack(anchor="w", pady=(8, 0))
    dedup_chk = tk.Checkbutton(
        frow, text="Optional filename versioning: keep only the newest file "
        "per DSCIT ID (leave OFF unless IDs are unique per report)",
        variable=state["keep_latest"], onvalue=True, offvalue=False,
        bg=WHITE, fg=GREY, activebackground=WHITE,
        font=("Segoe UI", 9), anchor="w", highlightthickness=0)
    dedup_chk.pack(anchor="w", pady=(2, 0))

    # stats row
    stats = tk.Frame(body, bg=TD_MIST)
    stats.pack(fill="x", pady=(16, 0))

    def stat_card(parent, key, caption, last=False):
        c = tk.Frame(parent, bg=WHITE, highlightbackground="#E2E8E2",
                     highlightthickness=1)
        c.pack(side="left", fill="x", expand=True,
               padx=(0, 0 if last else 12))
        v = tk.StringVar(value="0")
        stat_vars[key] = v
        ttk.Label(c, textvariable=v, style="StatValue.TLabel").pack(
            anchor="w", padx=16, pady=(12, 0))
        ttk.Label(c, text=caption, style="StatCap.TLabel").pack(
            anchor="w", padx=16, pady=(0, 12))

    stat_card(stats, "found", "FILES FOUND")
    stat_card(stats, "processed", "FILES PROCESSED")
    stat_card(stats, "rows", "ROWS EXTRACTED")
    stat_card(stats, "dupes", "DUPLICATES REMOVED")
    stat_card(stats, "warn", "WARNINGS")
    stat_card(stats, "err", "ERRORS", last=True)

    # progress
    pcard = tk.Frame(body, bg=TD_MIST)
    pcard.pack(fill="x", pady=(16, 0))
    pbar = ttk.Progressbar(pcard, style="TD.Horizontal.TProgressbar",
                           mode="determinate")
    pbar.pack(fill="x")
    tk.Label(pcard, textvariable=status_var, bg=TD_MIST, fg=GREY,
             font=("Segoe UI", 9)).pack(anchor="w", pady=(6, 0))

    # results table
    tcard = tk.Frame(body, bg=WHITE, highlightbackground="#E2E8E2",
                     highlightthickness=1)
    tcard.pack(fill="both", expand=True, pady=(16, 0))
    cols = ("file", "size", "status", "sheet", "hdr", "rows", "matched",
            "notes")
    tree = ttk.Treeview(tcard, columns=cols, show="headings")
    heads = {"file": ("File", 280), "size": ("Size", 75),
             "status": ("Status", 65), "sheet": ("Sheet", 105),
             "hdr": ("Header Row", 80), "rows": ("Rows", 70),
             "matched": ("Cols Matched", 90), "notes": ("Notes", 280)}
    for c, (txt, w) in heads.items():
        tree.heading(c, text=txt)
        tree.column(c, width=w, anchor="w")
    vsb = ttk.Scrollbar(tcard, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)
    tree.pack(side="left", fill="both", expand=True, padx=(1, 0), pady=1)
    vsb.pack(side="right", fill="y")
    tree.tag_configure("OK", foreground="#006A00")
    tree.tag_configure("WARN", foreground=AMBER)
    tree.tag_configure("ERROR", foreground=RED)
    tree.tag_configure("SKIP", foreground="#8a97a5")

    # footer buttons
    foot = tk.Frame(body, bg=TD_MIST)
    foot.pack(fill="x", pady=(12, 0))

    def _open(path):
        if not path:
            return
        try:
            os.startfile(path)          # Windows
        except AttributeError:
            webbrowser.open(f"file://{os.path.abspath(path)}")

    open_html_btn = ttk.Button(foot, text="Open Dashboard (HTML)",
                               style="Ghost.TButton", state="disabled",
                               command=lambda: _open(state["html"]))
    open_html_btn.pack(side="right")
    open_xl_btn = ttk.Button(foot, text="Open Workbook (Excel)",
                             style="Ghost.TButton", state="disabled",
                             command=lambda: _open(state["excel"]))
    open_xl_btn.pack(side="right", padx=(0, 8))
    open_folder_btn = ttk.Button(
        foot, text="Open Folder", style="Ghost.TButton", state="disabled",
        command=lambda: state["excel"] and _open(os.path.dirname(state["excel"])))
    open_folder_btn.pack(side="right", padx=(0, 8))

    # ---------------- workers ----------------
    def consolidation_worker(folder, keep_latest):
        try:
            def cb(done, total, res):
                ui_queue.put(("progress", done, total, res))
            results, combined, dedup = run_consolidation(
                folder, progress_cb=cb, keep_latest_per_id=keep_latest)
            analytics = build_analytics(combined, results, dedup) \
                if len(combined) else None
            out = write_output(folder, results, combined, analytics, dedup)
            ui_queue.put(("done", results, combined, out, dedup))
        except Exception:
            ui_queue.put(("fatal", traceback.format_exc()))

    def dashboard_worker(folder):
        try:
            combined = state["combined"]
            analytics = build_analytics(combined, state["results"],
                                        state["dedup"])
            html_path = generate_dashboard(folder, analytics, combined,
                                           state["excel"] or "",
                                           state["dedup"])
            ui_queue.put(("dash_done", html_path))
        except Exception:
            ui_queue.put(("fatal", traceback.format_exc()))

    def start():
        folder = state["folder"].get().strip()
        if not os.path.isdir(folder):
            messagebox.showwarning("Folder required",
                                   "Please browse to a valid folder first.")
            return
        files = discover_files(folder)
        if not files:
            messagebox.showinfo(
                "No files", "No DSCIT*.xlsx files were found in that folder.")
            return
        for i in tree.get_children():
            tree.delete(i)
        for kk in ("processed", "rows", "dupes", "warn", "err"):
            stat_vars[kk].set("0")
        stat_vars["found"].set(str(len(files)))
        pbar["value"] = 0
        pbar["maximum"] = len(files)
        state.update(running=True, excel=None, html=None,
                     results=None, combined=None, dedup=None)
        run_btn.state(["disabled"])
        dash_btn.state(["disabled"])
        for b in (open_html_btn, open_xl_btn, open_folder_btn):
            b.state(["disabled"])
        status_var.set("Running consolidation…")
        threading.Thread(target=consolidation_worker,
                         args=(folder, state["keep_latest"].get()),
                         daemon=True).start()

    def start_dashboard():
        if state["combined"] is None or state["combined"].empty:
            messagebox.showinfo("Run first",
                                "Run a consolidation before generating "
                                "the dashboard.")
            return
        dash_btn.state(["disabled"])
        status_var.set("Building executive dashboard…")
        threading.Thread(target=dashboard_worker,
                         args=(state["folder"].get().strip(),),
                         daemon=True).start()

    run_btn.configure(command=start)
    dash_btn.configure(command=start_dashboard)

    def poll():
        try:
            while True:
                msg = ui_queue.get_nowait()
                if msg[0] == "progress":
                    _, done, total, res = msg
                    pbar["value"] = done
                    if res.status != "SKIP":
                        stat_vars["processed"].set(
                            str(int(stat_vars["processed"].get()) + 1))
                    stat_vars["rows"].set(
                        str(int(stat_vars["rows"].get()) + res.rows))
                    if res.status == "WARN":
                        stat_vars["warn"].set(
                            str(int(stat_vars["warn"].get()) + 1))
                    if res.status == "ERROR":
                        stat_vars["err"].set(
                            str(int(stat_vars["err"].get()) + 1))
                    tree.insert("", "end", tags=(res.status,), values=(
                        res.file, human_size(res.size), res.status, res.sheet,
                        res.header_row or "—", res.rows,
                        f"{res.matched}/{len(TARGET_COLUMNS)}",
                        (res.message + ("  |  Missing: " + ", ".join(res.missing)
                                        if res.missing else ""))[:180]))
                    status_var.set(f"Processed {done} of {total}…")
                elif msg[0] == "done":
                    _, results, combined, out, dedup = msg
                    state.update(running=False, excel=out,
                                 results=results, combined=combined,
                                 dedup=dedup)
                    total_dupes = (dedup["duplicate_rows_removed"]
                                   if dedup else 0)
                    stat_vars["dupes"].set(str(total_dupes))
                    run_btn.state(["!disabled"])
                    dash_btn.state(["!disabled"])
                    open_xl_btn.state(["!disabled"])
                    open_folder_btn.state(["!disabled"])
                    skipped_files = (dedup["identical_files_skipped"]
                                     + dedup["superseded_files_skipped"]) \
                        if dedup else 0
                    status_var.set(
                        f"Complete — {len(combined):,} unique rows "
                        f"({total_dupes:,} duplicate rows removed, "
                        f"{skipped_files} duplicate/superseded file(s) "
                        f"skipped). Workbook: {os.path.basename(out)}. "
                        f"Ready for the dashboard ★")
                elif msg[0] == "dash_done":
                    state["html"] = msg[1]
                    dash_btn.state(["!disabled"])
                    open_html_btn.state(["!disabled"])
                    status_var.set("Executive dashboard ready: "
                                   f"{os.path.basename(msg[1])}")
                    _open(msg[1])
                elif msg[0] == "fatal":
                    state["running"] = False
                    run_btn.state(["!disabled"])
                    dash_btn.state(["!disabled"])
                    status_var.set("Run failed — see details.")
                    messagebox.showerror("Run failed", msg[1][-1500:])
        except queue.Empty:
            pass
        root.after(100, poll)

    poll()
    root.mainloop()


# ============================================================================
# 8. ENTRY
# ============================================================================

if __name__ == "__main__":
    if "--headless" in sys.argv:
        folder = sys.argv[sys.argv.index("--headless") + 1]
        results, combined, dedup = run_consolidation(
            folder, progress_cb=lambda d, t, r: print(
                f"[{d}/{t}] {r.file}: {r.status} rows={r.rows} {r.message}"))
        analytics = build_analytics(combined, results, dedup)
        out = write_output(folder, results, combined, analytics, dedup)
        dash = generate_dashboard(folder, analytics, combined, out, dedup)
        print(f"\nDedup: {dedup['duplicate_rows_removed']:,} duplicate rows "
              f"removed; {dedup['identical_files_skipped']} identical + "
              f"{dedup['superseded_files_skipped']} superseded file(s) "
              f"skipped")
        print(f"Workbook:  {out}  ({len(combined):,} unique rows)")
        print(f"Dashboard: {dash}")
    else:
        launch_ui()
