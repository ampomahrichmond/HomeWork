"""
Excel Data Quality Analysis Tool - Enhanced Version
A GUI application for analyzing EDMO Data Quality Control Evaluation Workbooks
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
from pathlib import Path


class SheetSelectorDialog:
    """Dialog for selecting sheets to analyze"""
    def __init__(self, parent, file_path):
        self.result = None
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(f"Select Sheets - {os.path.basename(file_path)}")
        self.dialog.geometry("500x400")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Load workbook to get sheet names
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
            self.sheet_names = wb.sheetnames
            wb.close()
        except Exception as e:
            messagebox.showerror("Error", f"Could not load file: {str(e)}")
            self.dialog.destroy()
            return
            
        self.setup_ui()
        
    def setup_ui(self):
        """Setup the dialog UI"""
        # Header
        header = tk.Label(
            self.dialog,
            text="Select sheets to analyze:",
            font=("Helvetica", 12, "bold"),
            pady=10
        )
        header.pack()
        
        # Sheet listbox with checkboxes
        frame = tk.Frame(self.dialog)
        frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # Scrollbar
        scrollbar = tk.Scrollbar(frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Listbox
        self.listbox = tk.Listbox(
            frame,
            selectmode=tk.MULTIPLE,
            yscrollcommand=scrollbar.set,
            font=("Courier", 10),
            height=15
        )
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.listbox.yview)
        
        # Add sheet names
        for sheet in self.sheet_names:
            self.listbox.insert(tk.END, sheet)
            
        # Pre-select required sheets
        for idx, sheet in enumerate(self.sheet_names):
            if "DQ Control" in sheet or "Evaluation" in sheet or "Scoping" in sheet:
                self.listbox.selection_set(idx)
                
        # Info label
        info = tk.Label(
            self.dialog,
            text="Tip: Required sheets are pre-selected",
            font=("Helvetica", 9, "italic"),
            fg="gray"
        )
        info.pack()
        
        # Buttons
        btn_frame = tk.Frame(self.dialog)
        btn_frame.pack(pady=10)
        
        tk.Button(
            btn_frame,
            text="Select All",
            command=self.select_all,
            padx=20,
            pady=5
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            btn_frame,
            text="Clear All",
            command=self.clear_all,
            padx=20,
            pady=5
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            btn_frame,
            text="OK",
            command=self.ok,
            bg="#27ae60",
            fg="white",
            padx=30,
            pady=5
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            btn_frame,
            text="Cancel",
            command=self.cancel,
            bg="#e74c3c",
            fg="white",
            padx=30,
            pady=5
        ).pack(side=tk.LEFT, padx=5)
        
    def select_all(self):
        self.listbox.selection_set(0, tk.END)
        
    def clear_all(self):
        self.listbox.selection_clear(0, tk.END)
        
    def ok(self):
        selected_indices = self.listbox.curselection()
        self.result = [self.sheet_names[i] for i in selected_indices]
        self.dialog.destroy()
        
    def cancel(self):
        self.result = None
        self.dialog.destroy()


class ExcelDQAnalyzer:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Data Quality Analyzer - Enhanced")
        
        # Make window resizable and adaptive
        self.root.geometry("1200x800")
        self.root.minsize(900, 600)  # Minimum size for usability
        self.root.configure(bg="#f0f0f0")
        
        # Make window resizable
        self.root.rowconfigure(0, weight=1)
        self.root.columnconfigure(0, weight=1)
        
        # Variables
        self.file_configs = []  # List of {path, sheets} dictionaries
        self.results = []
        self.log_data = []
        self.scoping_formula_results = []
        self.cross_reference_results = []
        self.error_cells = {}  # Track error cells for highlighting
        self.highlighted_files = []  # List of highlighted error files created
        
        # Setup UI
        self.setup_ui()
        
    def setup_ui(self):
        """Setup the main user interface"""
        # Header
        header_frame = tk.Frame(self.root, bg="#2c3e50", height=80)
        header_frame.pack(fill=tk.X, padx=0, pady=0)
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(
            header_frame, 
            text="📊 Excel Data Quality Analyzer - Enhanced",
            font=("Helvetica", 24, "bold"),
            bg="#2c3e50",
            fg="white"
        )
        title_label.pack(pady=20)
        
        # Main content frame
        content_frame = tk.Frame(self.root, bg="#f0f0f0")
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # File selection section
        self.setup_file_section(content_frame)
        
        # Statistics section
        self.setup_stats_section(content_frame)
        
        # Progress section
        self.setup_progress_section(content_frame)
        
        # Results section
        self.setup_results_section(content_frame)
        
        # Action buttons
        self.setup_action_buttons(content_frame)
        
    def setup_file_section(self, parent):
        """Setup file selection section"""
        file_frame = tk.LabelFrame(
            parent, 
            text="📁 File Selection & Configuration",
            font=("Helvetica", 12, "bold"),
            bg="#f0f0f0",
            padx=10,
            pady=10
        )
        file_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Create frame for listbox and scrollbar
        list_frame = tk.Frame(file_frame, bg="white")
        list_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Scrollbar
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # File list
        self.file_listbox = tk.Listbox(
            list_frame,
            height=5,
            font=("Courier", 9),
            bg="white",
            selectmode=tk.EXTENDED,
            yscrollcommand=scrollbar.set
        )
        self.file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.file_listbox.yview)
        
        # Buttons
        btn_frame = tk.Frame(file_frame, bg="#f0f0f0")
        btn_frame.pack(fill=tk.X)
        
        self.add_btn = tk.Button(
            btn_frame,
            text="➕ Add Files & Select Sheets",
            command=self.add_files,
            font=("Helvetica", 10),
            bg="#3498db",
            fg="white",
            padx=20,
            pady=5,
            cursor="hand2"
        )
        self.add_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        self.remove_btn = tk.Button(
            btn_frame,
            text="➖ Remove Selected",
            command=self.remove_files,
            font=("Helvetica", 10),
            bg="#e74c3c",
            fg="white",
            padx=20,
            pady=5,
            cursor="hand2"
        )
        self.remove_btn.pack(side=tk.LEFT, padx=5)
        
        self.clear_btn = tk.Button(
            btn_frame,
            text="🗑️ Clear All",
            command=self.clear_files,
            font=("Helvetica", 10),
            bg="#95a5a6",
            fg="white",
            padx=20,
            pady=5,
            cursor="hand2"
        )
        self.clear_btn.pack(side=tk.LEFT, padx=5)
        
    def setup_stats_section(self, parent):
        """Setup statistics section"""
        stats_frame = tk.LabelFrame(
            parent,
            text="📈 Analysis Statistics",
            font=("Helvetica", 12, "bold"),
            bg="#f0f0f0",
            padx=10,
            pady=10
        )
        stats_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Create grid for stats
        stats_grid = tk.Frame(stats_frame, bg="#f0f0f0")
        stats_grid.pack(fill=tk.X)
        
        # Define stat labels
        self.stat_labels = {}
        stats = [
            ("files", "Files Loaded", "#3498db"),
            ("eval_passed", "Eval Checks Passed", "#27ae60"),
            ("eval_failed", "Eval Checks Failed", "#e74c3c"),
            ("cross_matched", "Cross-Ref Matched", "#27ae60"),
            ("cross_unmatched", "Cross-Ref Unmatched", "#e74c3c"),
            ("formula_passed", "Formulas Passed", "#27ae60"),
            ("formula_failed", "Formulas Failed", "#e74c3c")
        ]
        
        for idx, (key, label, color) in enumerate(stats):
            frame = tk.Frame(stats_grid, bg="white", relief=tk.RAISED, borderwidth=1)
            frame.grid(row=0, column=idx, padx=5, pady=5, sticky="ew")
            
            tk.Label(
                frame,
                text=label,
                font=("Helvetica", 9),
                bg="white",
                fg="gray"
            ).pack(pady=(5, 0))
            
            value_label = tk.Label(
                frame,
                text="0",
                font=("Helvetica", 18, "bold"),
                bg="white",
                fg=color
            )
            value_label.pack(pady=(0, 5))
            self.stat_labels[key] = value_label
            
        # Make columns expand equally
        for i in range(len(stats)):
            stats_grid.columnconfigure(i, weight=1)
        
    def setup_progress_section(self, parent):
        """Setup progress section"""
        progress_frame = tk.LabelFrame(
            parent,
            text="⚙️ Analysis Progress",
            font=("Helvetica", 12, "bold"),
            bg="#f0f0f0",
            padx=10,
            pady=10
        )
        progress_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.progress_label = tk.Label(
            progress_frame,
            text="Ready to analyze files...",
            font=("Helvetica", 10),
            bg="#f0f0f0",
            fg="#2c3e50"
        )
        self.progress_label.pack(anchor=tk.W, pady=(0, 5))
        
        self.progress_bar = ttk.Progressbar(
            progress_frame,
            mode='determinate',
            length=100
        )
        self.progress_bar.pack(fill=tk.X)
        
    def setup_results_section(self, parent):
        """Setup results display section"""
        results_frame = tk.LabelFrame(
            parent,
            text="📋 Analysis Results",
            font=("Helvetica", 12, "bold"),
            bg="#f0f0f0",
            padx=10,
            pady=10
        )
        results_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Set minimum height to ensure buttons below are visible
        results_frame.pack_propagate(True)
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(results_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # Summary tab
        summary_frame = tk.Frame(self.notebook, bg="white")
        self.notebook.add(summary_frame, text="📊 Summary")
        
        self.summary_text = scrolledtext.ScrolledText(
            summary_frame,
            wrap=tk.WORD,
            font=("Courier", 10),
            bg="white",
            padx=10,
            pady=10,
            height=15
        )
        self.summary_text.pack(fill=tk.BOTH, expand=True)
        
        # Detailed log tab
        log_frame = tk.Frame(self.notebook, bg="white")
        self.notebook.add(log_frame, text="📝 Detailed Log")
        
        self.log_text = scrolledtext.ScrolledText(
            log_frame,
            wrap=tk.WORD,
            font=("Courier", 9),
            bg="white",
            padx=10,
            pady=10,
            height=15
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
    def setup_action_buttons(self, parent):
        """Setup action buttons"""
        action_frame = tk.Frame(parent, bg="#f0f0f0", relief=tk.RAISED, borderwidth=2)
        action_frame.pack(fill=tk.X, pady=10)
        
        # Add a label to make it more obvious
        button_label = tk.Label(
            action_frame,
            text="🎯 Actions",
            font=("Helvetica", 11, "bold"),
            bg="#f0f0f0",
            fg="#2c3e50"
        )
        button_label.pack(pady=(5, 5))
        
        # Button container
        button_container = tk.Frame(action_frame, bg="#f0f0f0")
        button_container.pack(pady=(0, 10))
        
        self.analyze_btn = tk.Button(
            button_container,
            text="🚀 Start Analysis",
            command=self.start_analysis,
            font=("Helvetica", 13, "bold"),
            bg="#27ae60",
            fg="white",
            padx=40,
            pady=12,
            cursor="hand2",
            relief=tk.RAISED,
            borderwidth=3
        )
        self.analyze_btn.pack(side=tk.LEFT, padx=10)
        
        self.export_btn = tk.Button(
            button_container,
            text="💾 Export Results to Excel",
            command=self.export_results,
            font=("Helvetica", 13, "bold"),
            bg="#f39c12",
            fg="white",
            padx=40,
            pady=12,
            cursor="hand2",
            state=tk.DISABLED,
            relief=tk.RAISED,
            borderwidth=3
        )
        self.export_btn.pack(side=tk.LEFT, padx=10)
        
    def add_files(self):
        """Add files and select sheets"""
        files = filedialog.askopenfilenames(
            title="Select Excel Files",
            filetypes=[("Excel Files", "*.xlsx *.xlsm *.xls"), ("All Files", "*.*")]
        )
        
        for file_path in files:
            # Check if file already added
            if any(config['path'] == file_path for config in self.file_configs):
                continue
                
            # Open sheet selector dialog
            dialog = SheetSelectorDialog(self.root, file_path)
            self.root.wait_window(dialog.dialog)
            
            if dialog.result:
                self.file_configs.append({
                    'path': file_path,
                    'sheets': dialog.result
                })
                
                # Add to listbox
                file_name = os.path.basename(file_path)
                sheets_str = ", ".join(dialog.result[:3])
                if len(dialog.result) > 3:
                    sheets_str += f" +{len(dialog.result) - 3} more"
                    
                display_text = f"{file_name} [{sheets_str}]"
                self.file_listbox.insert(tk.END, display_text)
                
        self.update_stats()
                
    def remove_files(self):
        """Remove selected files"""
        selected = self.file_listbox.curselection()
        for index in reversed(selected):
            self.file_listbox.delete(index)
            self.file_configs.pop(index)
        self.update_stats()
            
    def clear_files(self):
        """Clear all files"""
        self.file_listbox.delete(0, tk.END)
        self.file_configs.clear()
        self.update_stats()
        
    def update_stats(self):
        """Update statistics display"""
        self.stat_labels['files'].config(text=str(len(self.file_configs)))
        
    def start_analysis(self):
        """Start the analysis process"""
        if not self.file_configs:
            messagebox.showwarning("No Files", "Please add at least one file to analyze.")
            return
            
        # Clear previous results
        self.results.clear()
        self.log_data.clear()
        self.scoping_formula_results.clear()
        self.cross_reference_results.clear()
        self.summary_text.delete(1.0, tk.END)
        self.log_text.delete(1.0, tk.END)
        
        # Reset stats
        for key in ['eval_passed', 'eval_failed', 'cross_matched', 'cross_unmatched', 
                    'formula_passed', 'formula_failed']:
            self.stat_labels[key].config(text="0")
        
        # Disable buttons
        self.analyze_btn.config(state=tk.DISABLED)
        self.export_btn.config(state=tk.DISABLED)
        
        # Reset progress
        self.progress_bar['value'] = 0
        self.progress_bar['maximum'] = len(self.file_configs)
        
        # Process each file
        for idx, config in enumerate(self.file_configs):
            file_path = config['path']
            selected_sheets = config['sheets']
            
            self.progress_label.config(
                text=f"Analyzing: {os.path.basename(file_path)} ({len(selected_sheets)} sheets)"
            )
            self.root.update()
            
            try:
                self.analyze_file(file_path, selected_sheets)
            except Exception as e:
                self.log_message(
                    f"ERROR processing {os.path.basename(file_path)}: {str(e)}", 
                    "ERROR"
                )
                import traceback
                self.log_message(traceback.format_exc(), "ERROR")
                
            self.progress_bar['value'] = idx + 1
            self.root.update()
            
        # Analysis complete
        self.progress_label.config(text="Creating highlighted error files...")
        self.root.update()
        
        # Create highlighted files for each analyzed file
        self.highlighted_files.clear()
        for config in self.file_configs:
            file_path = config['path']
            if file_path in self.error_cells and self.error_cells[file_path]:
                highlighted_path = self.create_highlighted_error_file(file_path)
                if highlighted_path:
                    self.highlighted_files.append(highlighted_path)
        
        self.progress_label.config(text="✅ Analysis Complete!")
        self.display_summary()
        self.analyze_btn.config(state=tk.NORMAL)
        self.export_btn.config(state=tk.NORMAL)
        
        # Show completion message with highlighted file info
        msg = "Analysis completed successfully!"
        if self.highlighted_files:
            msg += f"\n\n{len(self.highlighted_files)} file(s) with highlighted errors created in outputs folder."
        
        messagebox.showinfo("Complete", msg)
        
    def analyze_file(self, file_path, selected_sheets):
        """Analyze a single Excel file"""
        self.log_message(
            f"\n{'='*80}\nAnalyzing File: {os.path.basename(file_path)}\n{'='*80}", 
            "INFO"
        )
        self.log_message(f"Selected Sheets: {', '.join(selected_sheets)}", "INFO")
        
        # Load workbook WITHOUT data_only to check formulas
        wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        # Load workbook WITH data_only to get evaluated values
        wb_values = openpyxl.load_workbook(file_path, data_only=True)
        
        # Find evaluation and scoping sheets
        eval_sheet_formulas = None
        eval_sheet_values = None
        scoping_sheet_formulas = None
        scoping_sheet_values = None
        
        for sheet_name in selected_sheets:
            if sheet_name not in wb_formulas.sheetnames:
                self.log_message(f"WARNING: Sheet '{sheet_name}' not found!", "WARNING")
                continue
                
            sheet_lower = sheet_name.lower()
            if "evaluation" in sheet_lower and "control" in sheet_lower:
                eval_sheet_formulas = wb_formulas[sheet_name]
                eval_sheet_values = wb_values[sheet_name]
                self.log_message(f"Found Evaluation Sheet: {sheet_name}", "INFO")
            elif "scoping" in sheet_lower and "control" in sheet_lower:
                scoping_sheet_formulas = wb_formulas[sheet_name]
                scoping_sheet_values = wb_values[sheet_name]
                self.log_message(f"Found Scoping Sheet: {sheet_name}", "INFO")
        
        if not eval_sheet_formulas:
            self.log_message("ERROR: DQ Controls Evaluation Sheet not found in selected sheets!", "ERROR")
        else:
            self.analyze_evaluation_sheet(eval_sheet_formulas, eval_sheet_values, file_path)
            
        if not scoping_sheet_formulas:
            self.log_message("ERROR: DQ Control Scoping sheet not found in selected sheets!", "ERROR")
        else:
            self.analyze_scoping_sheet(scoping_sheet_formulas, file_path)
            
        # Cross-reference check
        if eval_sheet_values and scoping_sheet_values:
            self.cross_reference_check(eval_sheet_values, scoping_sheet_values, file_path)
        
        wb_formulas.close()
        wb_values.close()
        
    def find_header_row(self, sheet, header_keywords):
        """Find the header row based on keywords"""
        for row_idx in range(1, min(50, sheet.max_row + 1)):
            row_values = [str(cell.value).strip() if cell.value else "" for cell in sheet[row_idx]]
            row_text = " ".join(row_values).upper()
            
            # Check if any of the keywords are present
            matches = sum(1 for keyword in header_keywords if keyword.upper() in row_text)
            if matches >= len(header_keywords) - 1:  # Allow for slight variations
                return row_idx
        return None
    
    def extract_components_from_formula(self, formula_str, evaluated_value):
        """
        Extract components based on formula structure from column AO
        The formula shows how components are concatenated
        Returns list of component values
        """
        if not formula_str or not evaluated_value:
            return evaluated_value.split('.') if evaluated_value else []
        
        # Parse formula to find separator pattern
        # Common patterns: CONCATENATE(R,".",V,".",U,".",W,".",X)
        # or R&"."&V&"."&U&"."&W&"."&X
        
        formula_upper = str(formula_str).upper()
        
        # Count how many separators (periods in quotes) are in the formula
        separator_count = formula_upper.count('"."') + formula_upper.count("'.'")
        
        # Expected components = separators + 1
        expected_components = separator_count + 1 if separator_count > 0 else 5
        
        # Split the evaluated value by period
        parts = str(evaluated_value).split('.')
        
        # If we have exactly the expected number of components, return them
        if len(parts) == expected_components:
            return parts
        
        # If formula suggests 5 components but we have more periods
        # Try to intelligently group them
        if expected_components == 5 and len(parts) > 5:
            # This handles cases where components themselves contain periods
            # We'll try to match the formula pattern
            return self.smart_split_components(formula_str, evaluated_value, expected_components)
        
        # Default: split by period
        return parts
    
    def smart_split_components(self, formula_str, value_str, expected_count=5):
        """
        Intelligently split value when components may contain periods
        Uses formula structure to guide the split
        """
        # For now, use simple heuristic: if we have more parts than expected,
        # join middle parts together
        parts = value_str.split('.')
        
        if len(parts) <= expected_count:
            return parts
        
        # Keep first component, last component, and group middle
        result = []
        result.append(parts[0])  # First component (MAL code)
        
        # Calculate how many parts to group in middle
        excess = len(parts) - expected_count
        
        # Second component (Database) - usually simple
        result.append(parts[1])
        
        # Third component (Schema) - might have periods
        if excess > 0:
            # Group some middle parts for schema
            schema_parts = parts[2:2 + excess + 1]
            result.append('.'.join(schema_parts))
        else:
            result.append(parts[2])
        
        # Fourth component (Table)
        result.append(parts[-2])
        
        # Fifth component (Column)
        result.append(parts[-1])
        
        return result[:expected_count]
        
    def analyze_evaluation_sheet(self, sheet_formulas, sheet_values, file_path):
        """Analyze the DQ Controls Evaluation Sheet"""
        self.log_message("\n--- DQ Controls Evaluation Sheet Analysis ---", "INFO")
        
        # Find header row
        header_keywords = ["MAL", "DATABASE", "UNIQUE IDENTIFIER"]
        header_row = self.find_header_row(sheet_formulas, header_keywords)
        
        if not header_row:
            self.log_message("ERROR: Could not find header row in Evaluation Sheet", "ERROR")
            return
            
        self.log_message(f"Header row found at: Row {header_row}", "INFO")
        
        # Get component name from Column B (MAL/EUC Code)
        mal_code_header = None
        for col_idx in range(1, 10):
            header_cell = sheet_formulas.cell(row=header_row, column=col_idx).value
            if header_cell and "MAL" in str(header_cell).upper() and "EUC" in str(header_cell).upper():
                mal_code_header = str(header_cell).strip()
                break
        
        if not mal_code_header:
            mal_code_header = "MAL/EUC Code"
        
        # Find Column D (should be the unique identifier column)
        col_d_idx = 4  # Column D is the 4th column
        
        # Analyze each row
        passed = 0
        failed = 0
        
        # Store rows with errors for highlighting
        error_rows = []
        
        for row_idx in range(header_row + 1, sheet_formulas.max_row + 1):
            # Get formula from formula sheet
            cell_formula = sheet_formulas.cell(row=row_idx, column=col_d_idx)
            cell_formula_value = cell_formula.value
            
            # Get evaluated value from value sheet
            cell_value = sheet_values.cell(row=row_idx, column=col_d_idx)
            evaluated_value = cell_value.value
            
            # Get MAL/EUC code from column B
            mal_code_cell = sheet_values.cell(row=row_idx, column=2)
            mal_code_value = str(mal_code_cell.value or "").strip()
            
            if not evaluated_value or str(evaluated_value).strip() == "":
                continue
                
            # Convert to string for processing
            evaluated_value_str = str(evaluated_value).strip()
            
            if evaluated_value_str == "None" or evaluated_value_str == "":
                continue
            
            # Check 1: Is it a formula? (check the formula sheet)
            is_formula = str(cell_formula_value).startswith('=') if cell_formula_value else False
            formula_check = "PASSED" if is_formula else "FAILED"
            
            # Check 2: Split by period for now (we'll use AO column info in cross-reference)
            # This is a basic check - the proper component check happens in cross-reference
            components = evaluated_value_str.split('.')
            component_count = len(components)
            component_check = "PASSED" if component_count == 5 else "FAILED"
            
            # Track errors for highlighting
            has_error = (formula_check == "FAILED" or component_check == "FAILED")
            if has_error:
                error_rows.append({
                    'row': row_idx,
                    'column': col_d_idx,
                    'reason': f"Formula: {formula_check}, Components: {component_check}"
                })
            
            # Log the result
            result = {
                "File": os.path.basename(file_path),
                "Sheet": sheet_formulas.title,
                "Row": row_idx,
                "Cell": f"D{row_idx}",
                "MAL_EUC_Code": mal_code_value,
                "Formula_Text": str(cell_formula_value)[:100] if cell_formula_value else "N/A",
                "Evaluated_Value": evaluated_value_str,
                "Formula Check": formula_check,
                "Component Check": component_check,
                "Component Count": component_count,
                f"Component 1 ({mal_code_header})": components[0] if len(components) > 0 else "",
                "Component 2 (Database Name)": components[1] if len(components) > 1 else "",
                "Component 3 (Schema)": components[2] if len(components) > 2 else "",
                "Component 4 (Table Name)": components[3] if len(components) > 3 else "",
                "Component 5 (Column Name)": components[4] if len(components) > 4 else "",
                "Has_Error": has_error
            }
            
            self.log_data.append(result)
            
            if formula_check == "PASSED" and component_check == "PASSED":
                passed += 1
                self.log_message(
                    f"Row {row_idx}: ✓ ALL CHECKS PASSED - Value: {evaluated_value_str[:70]}",
                    "INFO"
                )
            else:
                failed += 1
                self.log_message(
                    f"Row {row_idx}: ✗ FAILED - Formula: {formula_check}, Components: {component_check} (found {component_count}) - Value: {evaluated_value_str[:70]}",
                    "WARNING"
                )
                
        self.log_message(f"\nEvaluation Sheet Summary: {passed} PASSED, {failed} FAILED", "INFO")
        
        # Store error rows for highlighting
        if not hasattr(self, 'error_cells'):
            self.error_cells = {}
        self.error_cells[file_path] = {'evaluation': error_rows}
        
        # Update stats
        current_passed = int(self.stat_labels['eval_passed'].cget("text"))
        current_failed = int(self.stat_labels['eval_failed'].cget("text"))
        self.stat_labels['eval_passed'].config(text=str(current_passed + passed))
        self.stat_labels['eval_failed'].config(text=str(current_failed + failed))
        
    def analyze_scoping_sheet(self, sheet, file_path):
        """Analyze the DQ Control Scoping Sheet for formula validation"""
        self.log_message("\n--- DQ Control Scoping Sheet - Formula Analysis ---", "INFO")
        
        # Start from row 10 or find header
        start_row = 10
        header_keywords = ["UNIQUE IDENTIFIERS"]
        header_row = self.find_header_row(sheet, header_keywords)
        
        if header_row:
            start_row = header_row + 1
            self.log_message(f"Header row found at: Row {header_row}, starting analysis at row {start_row}", "INFO")
        else:
            self.log_message(f"Header not found, starting from row {start_row}", "INFO")
        
        # Check columns AF-AV for formulas (columns 32-48)
        af_col = 32  # Column AF
        av_col = 48  # Column AV
        
        passed = 0
        failed = 0
        total_cells = 0
        
        for row_idx in range(start_row, min(start_row + 1000, sheet.max_row + 1)):
            # Check if row has any data in columns Z-AD (reference columns)
            has_reference_data = False
            for ref_col in range(26, 31):  # Z to AD
                if sheet.cell(row=row_idx, column=ref_col).value:
                    has_reference_data = True
                    break
                    
            if not has_reference_data:
                continue
                
            # Now check formulas in AF-AV
            row_results = []
            
            for col_idx in range(af_col, av_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell_value = cell.value
                
                if cell_value is not None and str(cell_value).strip() != "":
                    total_cells += 1
                    is_formula = str(cell_value).startswith('=')
                    formula_check = "PASSED" if is_formula else "FAILED"
                    
                    col_letter = get_column_letter(col_idx)
                    
                    row_results.append({
                        "File": os.path.basename(file_path),
                        "Sheet": sheet.title,
                        "Row": row_idx,
                        "Column": col_letter,
                        "Cell": f"{col_letter}{row_idx}",
                        "Has_Formula": is_formula,
                        "Formula Check": formula_check,
                        "Value_Preview": str(cell_value)[:50]
                    })
                    
                    if formula_check == "PASSED":
                        passed += 1
                    else:
                        failed += 1
                        
            # Store results for this row
            if row_results:
                self.scoping_formula_results.extend(row_results)
                
                # Log summary for this row
                failed_cols = [r['Column'] for r in row_results if r['Formula Check'] == 'FAILED']
                if failed_cols:
                    self.log_message(
                        f"Row {row_idx}: Formula checks - {len(failed_cols)} FAILED ({', '.join(failed_cols[:5])}{'...' if len(failed_cols) > 5 else ''})",
                        "WARNING"
                    )
                    
        self.log_message(
            f"\nScoping Sheet Formula Summary: Total Cells Checked: {total_cells}, {passed} PASSED, {failed} FAILED",
            "INFO"
        )
        
        # Update stats
        current_passed = int(self.stat_labels['formula_passed'].cget("text"))
        current_failed = int(self.stat_labels['formula_failed'].cget("text"))
        self.stat_labels['formula_passed'].config(text=str(current_passed + passed))
        self.stat_labels['formula_failed'].config(text=str(current_failed + failed))
        
    def cross_reference_check(self, eval_sheet, scoping_sheet, file_path):
        """Cross-reference between Evaluation and Scoping sheets"""
        self.log_message("\n--- Cross-Reference Validation ---", "INFO")
        self.log_message("Validating that Column D components match Columns R, V, U, W, X in Scoping sheet", "INFO")
        self.log_message("Using Column AO formula structure to properly extract components", "INFO")
        
        # Find headers
        eval_header_row = self.find_header_row(eval_sheet, ["MAL", "DATABASE"])
        scoping_header_row = self.find_header_row(scoping_sheet, ["UNIQUE IDENTIFIERS"])
        
        if not scoping_header_row:
            scoping_header_row = 10
            
        if not eval_header_row:
            self.log_message("ERROR: Could not find evaluation sheet header", "ERROR")
            return
            
        self.log_message(
            f"Using Scoping data starting from row {scoping_header_row}",
            "INFO"
        )
        
        # Load scoping data using the CORRECT columns: R, V, U, W, X
        # Also load column AO (column 41) which contains the formula structure
        # R=18 (MAL/EUC), V=22 (Database), U=21 (Schema), W=23 (Table), X=24 (Column), AO=41 (Formula)
        scoping_data = []
        for row_idx in range(scoping_header_row, min(scoping_header_row + 2000, scoping_sheet.max_row + 1)):
            row_data = {
                "MAL_EUC": str(scoping_sheet.cell(row=row_idx, column=18).value or "").strip().upper(),     # Column R
                "Database": str(scoping_sheet.cell(row=row_idx, column=22).value or "").strip().upper(),    # Column V
                "Schema": str(scoping_sheet.cell(row=row_idx, column=21).value or "").strip().upper(),      # Column U
                "Table": str(scoping_sheet.cell(row=row_idx, column=23).value or "").strip().upper(),       # Column W
                "Column": str(scoping_sheet.cell(row=row_idx, column=24).value or "").strip().upper(),      # Column X
                "Formula": str(scoping_sheet.cell(row=row_idx, column=41).value or ""),                     # Column AO
                "Row": row_idx
            }
            
            # Only add if at least one field has data
            if any([v for k, v in row_data.items() if k not in ["Row", "Formula"]]):
                scoping_data.append(row_data)
                
        self.log_message(f"Loaded {len(scoping_data)} scoping records from columns R, V, U, W, X (using AO for structure)", "INFO")
        
        # Check each evaluation row
        matched = 0
        unmatched = 0
        error_rows = []
        
        for row_idx in range(eval_header_row + 1, min(eval_header_row + 2000, eval_sheet.max_row + 1)):
            # Get EVALUATED VALUE from Column D (using value sheet)
            cell = eval_sheet.cell(row=row_idx, column=4)  # Column D
            cell_value = str(cell.value or "").strip()
            
            if not cell_value or cell_value == "None":
                continue
            
            # Try to find a scoping row with a formula to understand component structure
            sample_formula = None
            for scoping_row in scoping_data:
                if scoping_row["Formula"]:
                    sample_formula = scoping_row["Formula"]
                    break
            
            # Extract components using formula structure
            components = self.extract_components_from_formula(sample_formula, cell_value)
            
            if len(components) != 5:
                self.log_message(
                    f"Row {row_idx}: Skipping - Invalid component count ({len(components)}) in value: {cell_value[:60]}",
                    "WARNING"
                )
                error_rows.append({
                    'row': row_idx,
                    'column': 4,
                    'reason': f"Expected 5 components, found {len(components)}"
                })
                continue
                
            # Normalize components (case-insensitive)
            comp_mal_euc = components[0].strip().upper()      # Component 1 → Column R
            comp_database = components[1].strip().upper()      # Component 2 → Column V
            comp_schema = components[2].strip().upper()        # Component 3 → Column U
            comp_table = components[3].strip().upper()         # Component 4 → Column W
            comp_column = components[4].strip().upper()        # Component 5 → Column X
            
            # Try to find match in scoping data
            match_found = False
            matched_row = None
            component_matches = {
                "MAL_EUC": False,
                "Database": False,
                "Schema": False,
                "Table": False,
                "Column": False
            }
            
            for scoping_row in scoping_data:
                # Check each component against the CORRECT columns
                mal_match = scoping_row["MAL_EUC"] == comp_mal_euc      # R
                db_match = scoping_row["Database"] == comp_database      # V
                schema_match = scoping_row["Schema"] == comp_schema      # U
                table_match = scoping_row["Table"] == comp_table         # W
                col_match = scoping_row["Column"] == comp_column         # X
                
                # All components must match
                if all([mal_match, db_match, schema_match, table_match, col_match]):
                    match_found = True
                    matched_row = scoping_row["Row"]
                    component_matches = {
                        "MAL_EUC": True,
                        "Database": True,
                        "Schema": True,
                        "Table": True,
                        "Column": True
                    }
                    break
                    
            # Track errors
            if not match_found:
                error_rows.append({
                    'row': row_idx,
                    'column': 4,
                    'reason': 'No match found in scoping data'
                })
            
            # Record result with YES/NO indicators
            result = {
                "File": os.path.basename(file_path),
                "Eval_Sheet": eval_sheet.title,
                "Eval_Row": row_idx,
                "Full_String": cell_value,
                "Component_1_MAL_EUC": comp_mal_euc,
                "MAL_Code_Found_In_DQ_Controls": "YES" if component_matches["MAL_EUC"] else "NO",
                "Component_2_Database": comp_database,
                "Database_Found_In_DQ_Controls": "YES" if component_matches["Database"] else "NO",
                "Component_3_Schema": comp_schema,
                "Schema_Found_In_DQ_Controls": "YES" if component_matches["Schema"] else "NO",
                "Component_4_Table": comp_table,
                "Table_Found_In_DQ_Controls": "YES" if component_matches["Table"] else "NO",
                "Component_5_Column": comp_column,
                "Column_Found_In_DQ_Controls": "YES" if component_matches["Column"] else "NO",
                "Match_Found": "YES" if match_found else "NO",
                "MAL_EUC_Match_Col_R": "PASS" if component_matches["MAL_EUC"] or match_found else "FAIL",
                "Database_Match_Col_V": "PASS" if component_matches["Database"] or match_found else "FAIL",
                "Schema_Match_Col_U": "PASS" if component_matches["Schema"] or match_found else "FAIL",
                "Table_Match_Col_W": "PASS" if component_matches["Table"] or match_found else "FAIL",
                "Column_Match_Col_X": "PASS" if component_matches["Column"] or match_found else "FAIL",
                "Matched_Scoping_Row": matched_row if match_found else "N/A"
            }
            
            self.cross_reference_results.append(result)
            
            if match_found:
                matched += 1
                self.log_message(
                    f"Row {row_idx}: ✓ COMPLETE MATCH - {cell_value[:60]} (matched to Scoping row {matched_row}, columns R,V,U,W,X)",
                    "INFO"
                )
            else:
                unmatched += 1
                self.log_message(
                    f"Row {row_idx}: ✗ NO MATCH in Scoping columns R,V,U,W,X - {cell_value[:60]}",
                    "WARNING"
                )
                self.log_message(
                    f"  Looking for: R={comp_mal_euc} | V={comp_database} | U={comp_schema} | W={comp_table} | X={comp_column}",
                    "WARNING"
                )
                
        self.log_message(
            f"\nCross-Reference Summary: {matched} MATCHED, {unmatched} UNMATCHED (checked against Scoping columns R,V,U,W,X)",
            "INFO"
        )
        
        # Store error rows for highlighting
        if not hasattr(self, 'error_cells'):
            self.error_cells = {}
        if file_path not in self.error_cells:
            self.error_cells[file_path] = {}
        self.error_cells[file_path]['cross_reference'] = error_rows
        
        # Update stats
        current_matched = int(self.stat_labels['cross_matched'].cget("text"))
        current_unmatched = int(self.stat_labels['cross_unmatched'].cget("text"))
        self.stat_labels['cross_matched'].config(text=str(current_matched + matched))
        self.stat_labels['cross_unmatched'].config(text=str(current_unmatched + unmatched))
        
    def create_highlighted_error_file(self, original_file_path):
        """
        Create a copy of the original file with errors highlighted in red
        Returns the path to the highlighted file
        """
        try:
            from openpyxl.styles import PatternFill
            
            # Create output filename
            base_name = os.path.splitext(os.path.basename(original_file_path))[0]
            highlighted_filename = f"{base_name}_ERRORS_HIGHLIGHTED.xlsx"
            output_path = os.path.join(os.path.dirname(original_file_path), highlighted_filename)
            
            # Also save to outputs directory
            outputs_path = "/mnt/user-data/outputs/" + highlighted_filename
            
            self.log_message(f"Creating highlighted error file: {highlighted_filename}", "INFO")
            
            # Load workbook
            wb = openpyxl.load_workbook(original_file_path)
            
            # Red fill for errors
            red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
            
            # Get error cells for this file
            if original_file_path in self.error_cells:
                errors = self.error_cells[original_file_path]
                
                # Highlight evaluation sheet errors
                if 'evaluation' in errors:
                    for sheet_name in wb.sheetnames:
                        if "evaluation" in sheet_name.lower() and "control" in sheet_name.lower():
                            sheet = wb[sheet_name]
                            for error in errors['evaluation']:
                                cell = sheet.cell(row=error['row'], column=error['column'])
                                cell.fill = red_fill
                                # Add comment
                                from openpyxl.comments import Comment
                                cell.comment = Comment(f"ERROR: {error['reason']}", "DQ Analyzer")
                            self.log_message(f"Highlighted {len(errors['evaluation'])} errors in {sheet_name}", "INFO")
                            break
                
                # Highlight cross-reference errors
                if 'cross_reference' in errors:
                    for sheet_name in wb.sheetnames:
                        if "evaluation" in sheet_name.lower() and "control" in sheet_name.lower():
                            sheet = wb[sheet_name]
                            for error in errors['cross_reference']:
                                cell = sheet.cell(row=error['row'], column=error['column'])
                                cell.fill = red_fill
                                # Add comment
                                from openpyxl.comments import Comment
                                cell.comment = Comment(f"ERROR: {error['reason']}", "DQ Analyzer")
                            self.log_message(f"Highlighted {len(errors['cross_reference'])} cross-ref errors in {sheet_name}", "INFO")
                            break
            
            # Save the file
            wb.save(outputs_path)
            wb.close()
            
            self.log_message(f"✓ Highlighted file saved to: {outputs_path}", "INFO")
            return outputs_path
            
        except Exception as e:
            self.log_message(f"ERROR creating highlighted file: {str(e)}", "ERROR")
            return None
    
    def log_message(self, message, level="INFO"):
        """Add a message to the log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        formatted_msg = f"[{timestamp}] [{level}] {message}\n"
        
        self.log_text.insert(tk.END, formatted_msg)
        
        # Color coding
        if level == "ERROR":
            self.log_text.tag_add("error", "end-2l", "end-1l")
            self.log_text.tag_config("error", foreground="red")
        elif level == "WARNING":
            self.log_text.tag_add("warning", "end-2l", "end-1l")
            self.log_text.tag_config("warning", foreground="orange")
        elif level == "INFO":
            self.log_text.tag_add("info", "end-2l", "end-1l")
            self.log_text.tag_config("info", foreground="blue")
            
        self.log_text.see(tk.END)
        self.root.update()
        
    def display_summary(self):
        """Display summary of results"""
        highlighted_info = ""
        if self.highlighted_files:
            highlighted_info = f"""
🔴 ERROR HIGHLIGHTING:
   • Files with Errors Highlighted: {len(self.highlighted_files)}
   • Highlighted files saved in outputs folder
   • Look for: *_ERRORS_HIGHLIGHTED.xlsx
"""
        
        summary = f"""
╔═══════════════════════════════════════════════════════════════╗
║              ANALYSIS SUMMARY - ENHANCED REPORT               ║
╚═══════════════════════════════════════════════════════════════╝

📊 FILES ANALYZED: {len(self.file_configs)}

📋 EVALUATION SHEET CHECKS:
   • Total Records Analyzed: {len(self.log_data)}
   • Checks Passed: {self.stat_labels['eval_passed'].cget('text')}
   • Checks Failed: {self.stat_labels['eval_failed'].cget('text')}

🔗 CROSS-REFERENCE VALIDATION:
   • Total Cross-References: {len(self.cross_reference_results)}
   • Matched: {self.stat_labels['cross_matched'].cget('text')}
   • Unmatched: {self.stat_labels['cross_unmatched'].cget('text')}

📝 FORMULA CHECKS (Scoping Sheet):
   • Total Cells Checked: {len(self.scoping_formula_results)}
   • Formulas Found: {self.stat_labels['formula_passed'].cget('text')}
   • Non-Formulas: {self.stat_labels['formula_failed'].cget('text')}
{highlighted_info}
📅 Analysis Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

✅ All results are ready for export in three separate sheets:
   1. Evaluation Sheet Validation Results
   2. Cross-Reference Validation Results  
   3. Scoping Sheet Formula Check Results

"""
        self.summary_text.insert(tk.END, summary)
        
    def export_results(self):
        """Export results to Excel with multiple sheets"""
        if not self.log_data and not self.cross_reference_results and not self.scoping_formula_results:
            messagebox.showwarning("No Data", "No analysis data to export.")
            return
            
        # Ask for save location
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=f"DQ_Analysis_Results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not file_path:
            return
            
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Sheet 1: Evaluation Sheet Results
                if self.log_data:
                    df_eval = pd.DataFrame(self.log_data)
                    df_eval.to_excel(writer, sheet_name='Evaluation Results', index=False)
                    
                # Sheet 2: Cross-Reference Results
                if self.cross_reference_results:
                    df_cross = pd.DataFrame(self.cross_reference_results)
                    df_cross.to_excel(writer, sheet_name='Cross-Reference Results', index=False)
                    
                # Sheet 3: Scoping Formula Check Results
                if self.scoping_formula_results:
                    df_scoping = pd.DataFrame(self.scoping_formula_results)
                    df_scoping.to_excel(writer, sheet_name='Scoping Formula Results', index=False)
                    
                # Sheet 4: Summary Statistics
                summary_data = {
                    'Metric': [
                        'Files Analyzed',
                        'Evaluation Checks Passed',
                        'Evaluation Checks Failed',
                        'Cross-References Matched',
                        'Cross-References Unmatched',
                        'Formulas Found',
                        'Non-Formulas Found',
                        'Analysis Date'
                    ],
                    'Value': [
                        len(self.file_configs),
                        self.stat_labels['eval_passed'].cget('text'),
                        self.stat_labels['eval_failed'].cget('text'),
                        self.stat_labels['cross_matched'].cget('text'),
                        self.stat_labels['cross_unmatched'].cget('text'),
                        self.stat_labels['formula_passed'].cget('text'),
                        self.stat_labels['formula_failed'].cget('text'),
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    ]
                }
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Summary', index=False)
                
                # Sheet 5: Detailed Log
                log_content = self.log_text.get(1.0, tk.END)
                log_df = pd.DataFrame({'Execution Log': [log_content]})
                log_df.to_excel(writer, sheet_name='Detailed Log', index=False)
                
            messagebox.showinfo(
                "Success", 
                f"Results exported successfully!\n\n"
                f"Location: {file_path}\n\n"
                f"Sheets created:\n"
                f"• Evaluation Results\n"
                f"• Cross-Reference Results\n"
                f"• Scoping Formula Results\n"
                f"• Summary\n"
                f"• Detailed Log"
            )
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export results:\n{str(e)}")


def main():
    """Main function to run the application"""
    root = tk.Tk()
    app = ExcelDQAnalyzer(root)
    root.mainloop()


if __name__ == "__main__":
    main()
