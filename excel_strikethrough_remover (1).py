import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import re

def remove_strikethrough_rows(file_path, output_path):
    """
    Remove rows where Column A has strikethrough formatting and unstack multi-value cells
    """
    try:
        print("Step 1: Loading workbook and checking for strikethrough formatting...")
        
        # Load the workbook with openpyxl to check formatting
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Find rows to remove (store row numbers)
        rows_to_remove = []
        
        # Check each row in column A for strikethrough formatting
        for row_num in range(1, ws.max_row + 1):
            cell = ws[f'A{row_num}']
            if cell.font and cell.font.strike:
                rows_to_remove.append(row_num)
                print(f"Found strikethrough in row {row_num}: {cell.value}")
        
        # Remove strikethrough rows in reverse order
        for row_num in reversed(rows_to_remove):
            ws.delete_rows(row_num)
            print(f"Removed strikethrough row {row_num}")
        
        # Save intermediate file
        temp_file = file_path.replace('.xlsx', '_temp.xlsx')
        wb.save(temp_file)
        
        print("Step 2: Loading data for unstacking multi-value cells...")
        
        # Load the cleaned data into pandas for unstacking
        df = pd.read_excel(temp_file)
        
        print("Step 3: Unstacking multi-value cells...")
        
        # Function to split multi-value cells
        def split_multivalue_cell(value):
            if pd.isna(value) or value == '':
                return [value]
            
            # Convert to string and split by common delimiters
            str_value = str(value).strip()
            
            # Split by newlines first (most common in your data)
            if '\n' in str_value:
                parts = [part.strip() for part in str_value.split('\n') if part.strip()]
                return parts if parts else [str_value]
            
            # If no newlines, check for other delimiters
            for delimiter in [';', '|', ',']:
                if delimiter in str_value:
                    parts = [part.strip() for part in str_value.split(delimiter) if part.strip()]
                    return parts if parts else [str_value]
            
            return [str_value]
        
        # Find columns that contain multi-value cells
        expanded_rows = []
        
        for index, row in df.iterrows():
            # Check each cell in the row for multiple values
            max_splits = 1
            split_data = {}
            
            # Analyze each column for multi-value content
            for col in df.columns:
                cell_value = row[col]
                split_values = split_multivalue_cell(cell_value)
                split_data[col] = split_values
                max_splits = max(max_splits, len(split_values))
            
            # Create rows for each split
            for i in range(max_splits):
                new_row = {}
                for col in df.columns:
                    # Use the i-th value if it exists, otherwise use the last available value
                    split_values = split_data[col]
                    if i < len(split_values):
                        new_row[col] = split_values[i]
                    else:
                        # Use the last value or empty if all are empty
                        new_row[col] = split_values[-1] if split_values else ''
                
                expanded_rows.append(new_row)
        
        # Create new dataframe with expanded rows
        expanded_df = pd.DataFrame(expanded_rows)
        
        # Remove duplicate rows that might have been created
        print("Step 4: Removing duplicate rows...")
        initial_count = len(expanded_df)
        expanded_df = expanded_df.drop_duplicates().reset_index(drop=True)
        final_count = len(expanded_df)
        
        print(f"Removed {initial_count - final_count} duplicate rows")
        
        # Save the final result
        expanded_df.to_excel(output_path, index=False)
        
        # Clean up temp file
        if os.path.exists(temp_file):
            os.remove(temp_file)
        
        print(f"Successfully created unstacked file: {output_path}")
        print(f"Original rows: {len(df)}")
        print(f"Final rows after unstacking: {len(expanded_df)}")
        
        return True
        
    except FileNotFoundError:
        print(f"Error: File not found at {file_path}")
        return False
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        return False

def analyze_multivalue_cells(file_path):
    """
    Analyze the file to show which cells contain multiple values
    """
    try:
        df = pd.read_excel(file_path)
        print("Analyzing multi-value cells...")
        
        multivalue_found = False
        for col in df.columns:
            for index, value in enumerate(df[col]):
                if pd.notna(value) and isinstance(value, str):
                    if '\n' in str(value) or ';' in str(value) or '|' in str(value):
                        if not multivalue_found:
                            print("\nMulti-value cells found:")
                            multivalue_found = True
                        print(f"Row {index + 2}, Column '{col}': {len(str(value).split())} parts")
        
        if not multivalue_found:
            print("No obvious multi-value cells detected")
            
    except Exception as e:
        print(f"Error analyzing file: {str(e)}")

def main():
    # File paths
    input_file = r"C:\users\taf\Commercial_Lineage.xlsx"
    output_file = r"C:\users\taf\Commercial_Lineage_cleaned_unstacked.xlsx"
    
    # Check if input file exists
    if not os.path.exists(input_file):
        print(f"Input file not found: {input_file}")
        return
    
    print(f"Processing file: {input_file}")
    print("=" * 60)
    
    # First, analyze the file
    analyze_multivalue_cells(input_file)
    print("=" * 60)
    
    # Process the file
    success = remove_strikethrough_rows(input_file, output_file)
    
    if success:
        print("=" * 60)
        print("Processing completed successfully!")
        print(f"Check your new file: {output_file}")
    else:
        print("Processing failed. Please check the error messages above.")

if __name__ == "__main__":
    main()