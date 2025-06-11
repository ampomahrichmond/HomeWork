import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import os

def remove_strikethrough_rows(file_path, output_path):
    """
    Remove rows where Column A (Version) has strikethrough formatting
    """
    try:
        # Load the workbook with openpyxl to check formatting
        wb = load_workbook(file_path)
        ws = wb.active  # Get the active worksheet
        
        # Find rows to remove (store row numbers)
        rows_to_remove = []
        
        # Check each row in column A for strikethrough formatting
        for row_num in range(1, ws.max_row + 1):
            cell = ws[f'A{row_num}']
            
            # Check if the cell has strikethrough formatting
            if cell.font and cell.font.strike:
                rows_to_remove.append(row_num)
                print(f"Found strikethrough in row {row_num}: {cell.value}")
        
        # If no strikethrough rows found
        if not rows_to_remove:
            print("No strikethrough formatting found in Column A")
            # Still create a copy of the original file
            wb.save(output_path)
            print(f"Original file copied to: {output_path}")
            return
        
        # Remove rows in reverse order (from bottom to top) to avoid index shifting
        for row_num in reversed(rows_to_remove):
            ws.delete_rows(row_num)
            print(f"Removed row {row_num}")
        
        # Save the modified workbook
        wb.save(output_path)
        print(f"Successfully created new file: {output_path}")
        print(f"Removed {len(rows_to_remove)} rows with strikethrough formatting")
        
    except FileNotFoundError:
        print(f"Error: File not found at {file_path}")
    except Exception as e:
        print(f"Error processing file: {str(e)}")

def main():
    # File paths
    input_file = r"C:\users\taf\Commercial_Lineage.xlsx"
    output_file = r"C:\users\taf\Commercial_Lineage_cleaned.xlsx"
    
    # Check if input file exists
    if not os.path.exists(input_file):
        print(f"Input file not found: {input_file}")
        return
    
    print(f"Processing file: {input_file}")
    print("Scanning for strikethrough formatting in Column A (Version)...")
    
    # Process the file
    remove_strikethrough_rows(input_file, output_file)

if __name__ == "__main__":
    main()