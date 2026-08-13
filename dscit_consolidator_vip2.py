"""
DSCIT Tier-1 Report Consolidator + Executive Analytics
-------------------------------------------------------
v2 enhancements:
  * Database Type normalization (Oracle/ORACLE/Oracle 19c/Essbase -> Oracle, etc.)
  * Analytics engine (EUC/MAL, outputs, sources, segments, owners, DB landscape,
    gap analysis, completeness scoring)
  * Ultra-premium TD-themed executive HTML dashboard (self-contained, no
    internet needed - safe to email to VP/SVP)
  * Analytics tabs added to the Excel output
  * New "Generate Executive Dashboard" button in the UI

Run:  python dscit_consolidator.py
Requires:  pip install openpyxl pandas
"""

import os
import re
import sys
import glob
import json
import html
import queue
import threading
import traceback
import webbrowser
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

# ============================================================================
# 1. EXTRACTION ENGINE
# ============================================================================

TARGET_COLUMNS = [
    "Output Name",
    "Output Owner",
    "Data Element",
    "Data Element Description",
    "Source Name",
    "Source Type",
    "Source Application EUC / MAL Code",
    "Database Type",
    "Database Name",
    "Database",
    "Business Segment / Corporate Function",
    "Asset Owner Name",
    "Technology Asset Owner Name",
    "Data Owner Name",
]

COLUMN_ALIASES = {
    "Output Name": ["outputname", "reportname", "outputreportname"],
    "Output Owner": ["outputowner", "reportowner"],
    "Data Element": ["dataelement", "dataelementname", "criticaldataelement", "cde"],
    "Data Element Description": [
        "dataelementdescription", "dataelementdiscription",
        "dataelementdesc", "elementdescription",
    ],
    "Source Name": ["sourcename", "sourcesystemname", "datasourcename"],
    "Source Type": ["sourcetype", "sourcesystemtype"],
    "Source Application EUC / MAL Code": [
        "sourceapplicationeucmalcode", "sourceapplicationeucormalcode",
        "applicationeucmalcode", "eucmalcode", "sourceappeucmalcode",
        "sourceapplicationmalcode", "sourceapplicationeuccode", "malcode",
    ],
    "Database Type": ["databasetype", "dbtype"],
    "Database Name": ["databasename", "dbname"],
    "Database": ["database", "db"],
    "Business Segment / Corporate Function": [
        "businesssegmentcorporatefunction", "businesssegmentorcorporatefunction",
        "businesssegment", "corporatefunction", "segmentfunction",
    ],
    "Asset Owner Name": ["assetownername", "assetowner", "dataassetownername"],
    "Technology Asset Owner Name": [
        "technologyassetownername", "technologyassetowner",
        "techassetownername", "techassetowner", "taoname",
    ],
    "Data Owner Name": ["dataownername", "dataowner"],
}

HEADER_SCAN_ROWS = 40
HEADER_SCAN_COLS = 5
SHEET_KEY = "dscit"


def _norm(value) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(value).lower())


def find_dscit_sheet(sheetnames):
    candidates = [s for s in sheetnames if SHEET_KEY in _norm(s)]
    if not candidates:
        return None
    candidates.sort(key=lambda s: (_norm(s) != SHEET_KEY, len(s)))
    return candidates[0]


def find_header_row(rows):
    for idx, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        for cell in row[:HEADER_SCAN_COLS]:
            n = _norm(cell)
            if n == "objectid":
                return idx, row
            if n:
                break
    return None, None


def map_columns(header_row):
    headers = {i: _norm(h) for i, h in enumerate(header_row) if _norm(h)}
    mapping, claimed = {}, set()
    order = sorted(TARGET_COLUMNS, key=lambda t: -len(_norm(t)))
    for target in order:
        aliases = set(COLUMN_ALIASES[target]) | {_norm(target)}
        for idx, h in headers.items():
            if idx not in claimed and h in aliases:
                mapping[target] = idx
                claimed.add(idx)
                break
    for target in order:
        if target in mapping:
            continue
        aliases = sorted(set(COLUMN_ALIASES[target]) | {_norm(target)},
                         key=len, reverse=True)
        for alias in aliases:
            if len(alias) < 6:
                continue
            hit = None
            for idx, h in headers.items():
                if idx not in claimed and alias in h:
                    hit = idx
                    break
            if hit is not None:
                mapping[target] = hit
                claimed.add(hit)
                break
    return mapping


# ============================================================================
# 2. DATABASE TYPE NORMALIZATION
# ============================================================================
# Rules are checked top-to-bottom against the normalized value. First hit wins.

DB_TYPE_RULES = [
    (r"essbase|oracle|exadata|oradb", "Oracle"),
    (r"excel|xls|spreadsheet|workbook", "Excel"),
    (r"sqlserver|mssql|microsoftsql|sqlsrv", "SQL Server"),
    (r"db2|udb", "DB2"),
    (r"teradata", "Teradata"),
    (r"snowflake", "Snowflake"),
    (r"databricks|deltalake|unitycatalog", "Databricks"),
    (r"azure|synapse|adls|blob|awss3|amazons3|gcs|bigquery|redshift|cloud",
     "Cloud"),
    (r"hive|hadoop|impala|hdfs", "Hadoop"),
    (r"postgres", "PostgreSQL"),
    (r"mysql|mariadb", "MySQL"),
    (r"mongo", "MongoDB"),
    (r"msaccess|accessdb|access", "MS Access"),
    (r"sharepoint", "SharePoint"),
    (r"csvfile|csv|flatfile|textfile|txtfile|delimited", "Flat File"),
    (r"sasdataset|sas", "SAS"),
    (r"mainframe|vsam|imsdb|ims|cobol", "Mainframe"),
    (r"sybase", "Sybase"),
    (r"informix", "Informix"),
    (r"netezza", "Netezza"),
]

_DB_RULES_COMPILED = [(re.compile(p), canon) for p, canon in DB_TYPE_RULES]


def normalize_db_type(raw):
    """Map any raw Database Type variant to a canonical name."""
    if raw is None or (isinstance(raw, float) and raw != raw) \
            or str(raw).strip() == "":
        return "Not Specified"
    n = _norm(raw)
    if not n or n in ("na", "nan", "none", "null", "tbd", "unknown"):
        return "Not Specified"
    for rx, canon in _DB_RULES_COMPILED:
        if rx.search(n):
            return canon
    return str(raw).strip().title()


# ============================================================================
# 3. FILE PROCESSING
# ============================================================================

@dataclass
class FileResult:
    file: str
    status: str = "OK"
    sheet: str = ""
    header_row: int | None = None
    rows: int = 0
    matched: int = 0
    missing: list = field(default_factory=list)
    message: str = ""
    data: pd.DataFrame | None = None


def process_file(path: str) -> FileResult:
    fname = os.path.basename(path)
    res = FileResult(file=fname)
    try:
        if fname.lower().endswith(".xls"):
            xls = pd.ExcelFile(path)
            sheet_name = find_dscit_sheet(xls.sheet_names)
            if not sheet_name:
                res.status, res.message = "ERROR", "No DSCIT sheet found"
                return res
            res.sheet = sheet_name
            df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
            rows = [tuple(r) for r in df.itertuples(index=False, name=None)]
            return _extract(rows, res)

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet_name = find_dscit_sheet(wb.sheetnames)
            if not sheet_name:
                res.status, res.message = "ERROR", "No DSCIT sheet found"
                return res
            res.sheet = sheet_name
            ws = wb[sheet_name]
            rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
        return _extract(rows, res)
    except Exception as e:
        res.status = "ERROR"
        res.message = f"{type(e).__name__}: {e}"
        return res


def _extract(rows, res: FileResult) -> FileResult:
    hdr_idx, header = find_header_row(rows)
    if hdr_idx is None:
        res.status, res.message = "ERROR", "Header row with 'Object ID' not found"
        return res
    res.header_row = hdr_idx + 1

    mapping = map_columns(header)
    res.matched = len(mapping)
    res.missing = [t for t in TARGET_COLUMNS if t not in mapping]
    if not mapping:
        res.status, res.message = "ERROR", "No target columns matched"
        return res
    if res.missing:
        res.status = "WARN"
        res.message = f"{len(res.missing)} column(s) not found"

    records = []
    for row in rows[hdr_idx + 1:]:
        rec, blank = {}, True
        for target in TARGET_COLUMNS:
            idx = mapping.get(target)
            val = row[idx] if (idx is not None and idx < len(row)) else None
            if val is not None and str(val).strip() != "":
                blank = False
                rec[target] = str(val).strip() if isinstance(val, str) else val
            else:
                rec[target] = None
        if not blank:
            records.append(rec)

    df = pd.DataFrame(records, columns=TARGET_COLUMNS)
    # --- Database Type normalization (keep original for audit) ---
    df["Database Type (Original)"] = df["Database Type"]
    df["Database Type"] = df["Database Type"].map(normalize_db_type)
    df.insert(0, "DSCIT File Name", res.file)
    res.rows = len(df)
    res.data = df
    return res


def discover_files(folder: str):
    pats = ["DSCIT*.xlsx", "DSCIT*.xlsm", "DSCIT*.xls",
            "dscit*.xlsx", "dscit*.xlsm", "dscit*.xls"]
    files = set()
    for p in pats:
        files.update(glob.glob(os.path.join(folder, p)))
    return sorted(f for f in files
                  if not os.path.basename(f).startswith("~$")
                  and "consolidated" not in os.path.basename(f).lower())


def run_consolidation(folder, progress_cb=None, workers=8):
    files = discover_files(folder)
    results = []
    with ThreadPoolExecutor(max_workers=min(workers, max(1, len(files)))) as ex:
        futures = {ex.submit(process_file, f): f for f in files}
        done = 0
        for fut in as_completed(futures):
            results.append(fut.result())
            done += 1
            if progress_cb:
                progress_cb(done, len(files), results[-1])
    results.sort(key=lambda r: r.file.lower())
    frames = [r.data for r in results if r.data is not None and not r.data.empty]
    cols = ["DSCIT File Name"] + TARGET_COLUMNS + ["Database Type (Original)"]
    combined = pd.concat(frames, ignore_index=True) if frames else \
        pd.DataFrame(columns=cols)
    return results, combined


# ============================================================================
# 4. ANALYTICS ENGINE
# ============================================================================

def _clean(series):
    """Series of stripped strings with blanks/placeholder values as NA."""
    s = series.astype("string").str.strip()
    s = s.mask(s.str.lower().isin(["", "na", "n/a", "none", "null", "tbd", "-"]))
    return s


def _nunique(df, col):
    return int(_clean(df[col]).dropna().nunique()) if col in df else 0


def build_analytics(combined: pd.DataFrame, results=None):
    df = combined.copy()
    for c in TARGET_COLUMNS:
        if c in df:
            df[c] = _clean(df[c])

    a = {}

    # ---- KPIs ----
    kpi = {
        "Files Consolidated": df["DSCIT File Name"].nunique(),
        "Total Data Element Rows": len(df),
        "Distinct Data Elements": _nunique(df, "Data Element"),
        "Output Names": _nunique(df, "Output Name"),
        "Output Owners": _nunique(df, "Output Owner"),
        "Source Systems (Source Names)": _nunique(df, "Source Name"),
        "Source Types": _nunique(df, "Source Type"),
        "EUC / MAL Codes": _nunique(df, "Source Application EUC / MAL Code"),
        "Database Names": _nunique(df, "Database Name"),
        "Database Types": _nunique(df, "Database Type"),
        "Business Segments / Corp Functions":
            _nunique(df, "Business Segment / Corporate Function"),
        "Asset Owners": _nunique(df, "Asset Owner Name"),
        "Technology Asset Owners": _nunique(df, "Technology Asset Owner Name"),
        "Data Owners": _nunique(df, "Data Owner Name"),
    }
    a["kpis"] = kpi

    # ---- By Database Type (normalized) ----
    g = (df.groupby("Database Type", dropna=False)
           .agg(Databases=("Database Name", lambda s: s.dropna().nunique()),
                Data_Element_Rows=("Database Type", "size"),
                Outputs=("Output Name", lambda s: s.dropna().nunique()),
                Sources=("Source Name", lambda s: s.dropna().nunique()))
           .reset_index()
           .sort_values("Data_Element_Rows", ascending=False))
    g.columns = ["Database Type", "Distinct Databases", "Data Element Rows",
                 "Outputs", "Sources"]
    a["db_type"] = g

    # ---- Database Name x Type inventory ----
    inv = (df.dropna(subset=["Database Name"])
             .groupby(["Database Type", "Database Name"])
             .agg(Rows=("Database Name", "size"),
                  Outputs=("Output Name", lambda s: s.dropna().nunique()))
             .reset_index()
             .sort_values(["Database Type", "Rows"],
                          ascending=[True, False]))
    a["db_inventory"] = inv

    # ---- Outputs by Business Segment ----
    seg = (df.groupby("Business Segment / Corporate Function", dropna=False)
             .agg(Outputs=("Output Name", lambda s: s.dropna().nunique()),
                  Output_Owners=("Output Owner", lambda s: s.dropna().nunique()),
                  Data_Elements=("Data Element", lambda s: s.dropna().nunique()),
                  Rows=("Output Name", "size"))
             .reset_index()
             .sort_values("Outputs", ascending=False))
    seg["Business Segment / Corporate Function"] = \
        seg["Business Segment / Corporate Function"].fillna("Not Specified")
    seg.columns = ["Business Segment / Corporate Function", "Outputs",
                   "Output Owners", "Data Elements", "Rows"]
    a["segment"] = seg

    # ---- Databases by Asset Owner ----
    ao = (df.groupby("Asset Owner Name", dropna=False)
            .agg(Databases=("Database Name", lambda s: s.dropna().nunique()),
                 DB_Types=("Database Type", lambda s: s.dropna().nunique()),
                 Outputs=("Output Name", lambda s: s.dropna().nunique()))
            .reset_index()
            .sort_values("Databases", ascending=False))
    ao["Asset Owner Name"] = ao["Asset Owner Name"].fillna("Not Specified")
    ao.columns = ["Asset Owner Name", "Databases", "Database Types", "Outputs"]
    a["asset_owner"] = ao

    # ---- Databases by Output Name / Output Owner / Source Type ----
    def dbs_by(col):
        t = (df.groupby(col, dropna=False)
               .agg(Databases=("Database Name", lambda s: s.dropna().nunique()),
                    DB_Types=("Database Type",
                              lambda s: ", ".join(sorted(s.dropna().unique())[:6])))
               .reset_index()
               .sort_values("Databases", ascending=False))
        t[col] = t[col].fillna("Not Specified")
        t.columns = [col, "Databases", "Database Types"]
        return t
    a["dbs_by_output"] = dbs_by("Output Name")
    a["dbs_by_owner"] = dbs_by("Output Owner")
    a["dbs_by_srctype"] = dbs_by("Source Type")

    # ---- Source inventory + gap flags ----
    src = (df.groupby(["Source Name", "Source Application EUC / MAL Code",
                       "Source Type"], dropna=False)
             .agg(Databases=("Database Name", lambda s: s.dropna().nunique()),
                  Rows=("Source Name", "size"),
                  Missing_DB_Name=("Database Name",
                                   lambda s: int(s.isna().sum())),
                  Missing_Data_Owner=("Data Owner Name",
                                      lambda s: int(s.isna().sum())),
                  Missing_Asset_Owner=("Asset Owner Name",
                                       lambda s: int(s.isna().sum())))
             .reset_index())
    src["Source Name"] = src["Source Name"].fillna("«Source Missing»")
    src["Source Application EUC / MAL Code"] = \
        src["Source Application EUC / MAL Code"].fillna("«Code Missing»")
    src["Source Type"] = src["Source Type"].fillna("«Type Missing»")
    src["Gap Flags"] = src.apply(
        lambda r: "; ".join(f for f, c in [
            ("Source missing", r["Source Name"] == "«Source Missing»"),
            ("EUC/MAL missing", r["Source Application EUC / MAL Code"] == "«Code Missing»"),
            ("DB name gaps", r["Missing_DB_Name"] > 0),
            ("Data owner gaps", r["Missing_Data_Owner"] > 0),
            ("Asset owner gaps", r["Missing_Asset_Owner"] > 0),
        ] if c) or "Complete", axis=1)
    src = src.sort_values("Rows", ascending=False)
    a["source_inventory"] = src

    # ---- Data quality / completeness ----
    dq_rows = []
    for c in TARGET_COLUMNS:
        if c not in df:
            continue
        missing = int(df[c].isna().sum())
        pct = round(100 * (1 - missing / len(df)), 1) if len(df) else 0.0
        dq_rows.append({"Field": c, "Populated %": pct,
                        "Missing Rows": missing, "Total Rows": len(df)})
    dq = pd.DataFrame(dq_rows).sort_values("Populated %")
    a["dq"] = dq
    key_fields = ["Output Name", "Source Name", "Database Name",
                  "Data Owner Name", "Asset Owner Name",
                  "Source Application EUC / MAL Code"]
    kdq = dq[dq["Field"].isin(key_fields)]
    a["completeness_score"] = round(float(kdq["Populated %"].mean()), 1) \
        if len(kdq) else 0.0

    # ---- Top outputs by data-element footprint ----
    top = (df.groupby("Output Name", dropna=False)
             .agg(Data_Elements=("Data Element", lambda s: s.dropna().nunique()),
                  Sources=("Source Name", lambda s: s.dropna().nunique()),
                  Databases=("Database Name", lambda s: s.dropna().nunique()),
                  Owner=("Output Owner",
                         lambda s: s.dropna().iloc[0] if s.dropna().size else "—"))
             .reset_index()
             .sort_values("Data_Elements", ascending=False))
    top["Output Name"] = top["Output Name"].fillna("Not Specified")
    top.columns = ["Output Name", "Data Elements", "Sources", "Databases",
                   "Output Owner"]
    a["top_outputs"] = top

    # ---- DB type normalization audit map ----
    if "Database Type (Original)" in df:
        m = (combined.groupby(["Database Type (Original)", "Database Type"],
                              dropna=False).size().reset_index(name="Rows")
             .sort_values(["Database Type", "Rows"], ascending=[True, False]))
        m.columns = ["Raw Value", "Normalized To", "Rows"]
        a["norm_map"] = m

    return a


# ============================================================================
# 5. EXCEL OUTPUT (data + analytics tabs)
# ============================================================================

_ILLEGAL_XML = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _excel_safe(v):
    """Make any cell value safe for openpyxl: strip illegal XML control
    characters from strings, drop timezones from datetimes, stringify
    exotic objects."""
    if v is None:
        return v
    if isinstance(v, str):
        return _ILLEGAL_XML.sub("", v)
    if isinstance(v, datetime):
        return v.replace(tzinfo=None) if v.tzinfo else v
    tz = getattr(v, "tzinfo", None)          # pandas Timestamp etc.
    if tz is not None:
        try:
            return v.tz_localize(None)
        except Exception:
            return str(v)
    if isinstance(v, (int, float, bool)):
        return v
    if hasattr(v, "isoformat") or isinstance(v, (bytes, complex)):
        return _ILLEGAL_XML.sub("", str(v))
    return v


def sanitize_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for c in out.columns:
        if out[c].dtype == object or str(out[c].dtype) in ("string", "str"):
            out[c] = out[c].map(_excel_safe)
        elif "datetime" in str(out[c].dtype) and getattr(
                out[c].dtype, "tz", None) is not None:
            out[c] = out[c].dt.tz_localize(None)
    return out


def write_output(folder, results, combined, analytics=None) -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(folder, f"DSCIT_Consolidated_{stamp}.xlsx")
    summary = pd.DataFrame([{
        "File": r.file, "Status": r.status, "Sheet": r.sheet,
        "Header Row": r.header_row, "Rows Extracted": r.rows,
        "Columns Matched": f"{r.matched}/{len(TARGET_COLUMNS)}",
        "Missing Columns": "; ".join(r.missing), "Notes": r.message,
    } for r in results])

    xw = pd.ExcelWriter(out_path, engine="openpyxl")
    try:
        sanitize_df(combined).to_excel(
            xw, sheet_name="Consolidated Data", index=False)
        sanitize_df(summary).to_excel(
            xw, sheet_name="Run Summary", index=False)
        if analytics:
            pd.DataFrame(list(analytics["kpis"].items()),
                         columns=["Metric", "Value"]).to_excel(
                xw, sheet_name="Executive KPIs", index=False)
            sanitize_df(analytics["db_type"]).to_excel(xw, sheet_name="By Database Type", index=False)
            sanitize_df(analytics["db_inventory"]).to_excel(xw, sheet_name="DB Inventory", index=False)
            sanitize_df(analytics["segment"]).to_excel(xw, sheet_name="By Business Segment", index=False)
            sanitize_df(analytics["asset_owner"]).to_excel(xw, sheet_name="By Asset Owner", index=False)
            sanitize_df(analytics["dbs_by_output"]).to_excel(xw, sheet_name="DBs by Output", index=False)
            sanitize_df(analytics["dbs_by_owner"]).to_excel(xw, sheet_name="DBs by Output Owner",
                                               index=False)
            sanitize_df(analytics["dbs_by_srctype"]).to_excel(xw, sheet_name="DBs by Source Type",
                                                 index=False)
            sanitize_df(analytics["source_inventory"]).to_excel(xw, sheet_name="Source Inventory & Gaps",
                                                   index=False)
            sanitize_df(analytics["dq"]).to_excel(xw, sheet_name="Data Quality", index=False)
            if "norm_map" in analytics:
                sanitize_df(analytics["norm_map"]).to_excel(xw, sheet_name="DB Type Mapping",
                                               index=False)
        for ws in xw.book.worksheets:
            for col in ws.columns:
                width = max((len(str(c.value)) for c in col[:200] if c.value),
                            default=10)
                ws.column_dimensions[col[0].column_letter].width = \
                    min(width + 2, 48)
        xw.close()
    except Exception:
        # Ensure the save-on-close can't mask the real error with
        # "At least one sheet must be visible".
        try:
            if not xw.book.worksheets:
                xw.book.create_sheet("Sheet1")
            xw.close()
        except Exception:
            pass
        raise
    return out_path


# ============================================================================
# 6. EXECUTIVE HTML DASHBOARD (self-contained, SVG charts, TD theme)
# ============================================================================

def _dash_rows(combined: pd.DataFrame):
    """Serialize row-level data for the interactive dashboard (compact arrays)."""
    cols = ["DSCIT File Name", "Output Name", "Output Owner", "Data Element",
            "Source Name", "Source Type", "Source Application EUC / MAL Code",
            "Database Type", "Database Name", "Database",
            "Business Segment / Corporate Function", "Asset Owner Name",
            "Technology Asset Owner Name", "Data Owner Name"]
    df = combined.reindex(columns=cols)
    rows = []
    for tup in df.itertuples(index=False, name=None):
        rec = []
        for v in tup:
            if v is None or (isinstance(v, float) and v != v):
                rec.append(None)
            else:
                s = str(v).strip()
                rec.append(s if s and s.lower() not in
                           ("na", "n/a", "none", "null", "nan", "tbd", "-")
                           else None)
        rows.append(rec)
    return rows


DASH_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>DSCIT Tier-1 Executive Dashboard</title>
<style>
:root{
 --td:#008A00; --td-d:#006A00; --td-xd:#0B3D2E; --td-l:#54B848;
 --mist:#F3F7F3; --ink:#243024; --grey:#6b7f6b; --line:#E2EBE2;
 --amber:#B45309; --card:#ffffff;
}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Arial,sans-serif;background:var(--mist);
 color:var(--ink);overflow:hidden;height:100vh;display:flex;
 flex-direction:column}
/* ---------- top bar ---------- */
.topbar{background:linear-gradient(120deg,var(--td-xd),var(--td) 65%,
 var(--td-l) 130%);color:#fff;padding:14px 26px;display:flex;
 align-items:center;gap:16px;flex-shrink:0;z-index:5;
 box-shadow:0 4px 16px rgba(11,61,46,.25)}
.tdmark{background:#fff;color:var(--td);font-weight:900;font-size:17px;
 width:40px;height:40px;display:flex;align-items:center;
 justify-content:center;border-radius:8px;flex-shrink:0}
.topbar h1{font-size:18px;font-weight:650;line-height:1.15}
.topbar .sub{font-size:11.5px;color:#CBE9CB}
.topbar .meta{margin-left:auto;text-align:right;font-size:11px;
 color:#B7E0B7;line-height:1.5}
/* ---------- filter bar ---------- */
.filters{background:#fff;border-bottom:1px solid var(--line);
 padding:10px 26px;display:flex;gap:10px;align-items:center;
 flex-wrap:wrap;flex-shrink:0}
.filters label{font-size:10px;font-weight:700;color:var(--grey);
 text-transform:uppercase;letter-spacing:.6px;display:block;
 margin-bottom:3px}
.fgroup{min-width:150px;flex:1;max-width:220px}
.filters select,.filters input[type=text]{width:100%;padding:7px 9px;
 border:1px solid #D5DDD5;border-radius:8px;font-size:12.5px;
 background:var(--mist);color:var(--ink);outline:none}
.filters select:focus,.filters input:focus{border-color:var(--td)}
.fbtn{background:var(--td);color:#fff;border:none;border-radius:8px;
 padding:9px 16px;font-size:12px;font-weight:650;cursor:pointer;
 align-self:flex-end}
.fbtn.ghost{background:#fff;color:var(--td);border:1px solid var(--td)}
.fbtn:hover{filter:brightness(.95)}
.chip{background:#EAF5EA;color:var(--td-d);border-radius:99px;
 padding:4px 12px;font-size:11px;font-weight:650;align-self:flex-end;
 margin-bottom:2px}
/* ---------- layout ---------- */
.main{display:flex;flex:1;min-height:0}
.side{width:230px;background:#fff;border-right:1px solid var(--line);
 padding:16px 12px;flex-shrink:0;overflow-y:auto}
.side .nav{display:flex;flex-direction:column;gap:4px}
.navbtn{display:flex;align-items:center;gap:10px;padding:11px 13px;
 border:none;background:none;border-radius:10px;font-size:13px;
 font-weight:600;color:#415441;cursor:pointer;text-align:left;width:100%;
 font-family:inherit}
.navbtn:hover{background:var(--mist)}
.navbtn.active{background:linear-gradient(120deg,var(--td),var(--td-l));
 color:#fff;box-shadow:0 4px 12px rgba(0,138,0,.28)}
.navbtn .ic{width:20px;text-align:center;font-size:14px}
.side .foot{margin-top:18px;font-size:10px;color:#9ab09a;
 padding:0 10px;line-height:1.6}
.content{flex:1;overflow-y:auto;padding:22px 28px 48px}
/* ---------- components ---------- */
.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(158px,1fr));
 gap:12px;margin-bottom:18px}
.kpi{background:var(--card);border-radius:13px;padding:14px 16px;
 box-shadow:0 4px 14px rgba(11,61,46,.08);border-top:4px solid var(--td)}
.kpi b{display:block;font-size:24px;color:var(--td);font-weight:700}
.kpi span{font-size:10px;color:var(--grey);text-transform:uppercase;
 letter-spacing:.6px;font-weight:650}
.sec{background:var(--card);border-radius:15px;padding:20px 24px;
 margin-bottom:18px;box-shadow:0 4px 14px rgba(11,61,46,.07)}
.sec h2{font-size:15.5px;color:var(--td-xd);font-weight:650;display:flex;
 align-items:center;gap:9px;margin-bottom:4px}
.sec h2:before{content:'';width:7px;height:19px;background:var(--td);
 border-radius:4px}
.sec .note{font-size:12px;color:var(--grey);margin-bottom:14px}
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:22px}
@media(max-width:1100px){.grid2{grid-template-columns:1fr}}
.tbl{width:100%;border-collapse:collapse;font-size:12px}
.tbl th{background:#F0F6F0;color:var(--td-xd);text-align:left;
 padding:8px 9px;font-size:10px;text-transform:uppercase;
 letter-spacing:.5px;border-bottom:2px solid #DCE8DC;white-space:nowrap;
 position:sticky;top:0}
.tbl td{padding:7px 9px;border-bottom:1px solid #EDF3ED;vertical-align:top}
.tbl tr:hover td{background:#F7FBF7}
.tbl .num{text-align:right;font-weight:650;color:var(--td-xd)}
.tbl .gap{color:var(--amber);font-weight:650}
.tbl .ok{color:var(--td);font-weight:650}
.scrolltbl{max-height:430px;overflow-y:auto;border:1px solid var(--line);
 border-radius:10px}
.story{background:linear-gradient(135deg,#F0F8F0,#fff);
 border-left:5px solid var(--td);border-radius:12px;padding:16px 20px;
 margin-bottom:18px}
.story h3{color:var(--td-xd);font-size:13.5px;margin-bottom:7px}
.story li{font-size:13px;margin:5px 0 5px 18px;color:#37473a}
.donutwrap{display:flex;align-items:center;gap:20px;flex-wrap:wrap}
.legend{display:flex;flex-direction:column;gap:5px;font-size:12px}
.legend .lg{cursor:pointer}
.legend .lg:hover{text-decoration:underline}
.legend .lg i{display:inline-block;width:10px;height:10px;border-radius:3px;
 margin-right:6px}
.meter{background:#E7EFE7;border-radius:99px;height:11px;overflow:hidden;
 margin:5px 0 2px}
.meter i{display:block;height:100%;border-radius:99px;
 background:linear-gradient(90deg,var(--td),var(--td-l))}
.meter i.warn{background:linear-gradient(90deg,#D97706,#F59E0B)}
.meter i.bad{background:linear-gradient(90deg,#B91C1C,#EF4444)}
.dqrow{margin-bottom:11px;font-size:12px}
.dqrow .lbl{display:flex;justify-content:space-between;color:#37473a}
.empty{color:#9ab09a;font-size:13px;padding:22px;text-align:center}
.bar-click{cursor:pointer}
.hint{font-size:10.5px;color:#9ab09a;margin-top:6px}
</style></head><body>

<div class="topbar">
  <div class="tdmark">TD</div>
  <div><h1>DSCIT Tier-1 Source Landscape</h1>
    <div class="sub">Interactive Executive Analytics · Enterprise Data
    Management Office</div></div>
  <div class="meta">Generated __GENERATED__<br>
    Companion workbook: __EXCEL__</div>
</div>

<div class="filters">
  <div class="fgroup"><label>Business Segment</label>
    <select id="f_seg"></select></div>
  <div class="fgroup"><label>Database Type</label>
    <select id="f_dbt"></select></div>
  <div class="fgroup"><label>Output Owner</label>
    <select id="f_own"></select></div>
  <div class="fgroup"><label>Source Type</label>
    <select id="f_sty"></select></div>
  <div class="fgroup"><label>Asset Owner</label>
    <select id="f_ao"></select></div>
  <div class="fgroup"><label>Search (output / source / element / MAL)</label>
    <input type="text" id="f_q" placeholder="Type to search…"></div>
  <button class="fbtn ghost" id="f_reset">Reset</button>
  <span class="chip" id="f_chip"></span>
</div>

<div class="main">
  <div class="side">
    <div class="nav" id="nav"></div>
    <div class="foot">TD Bank Group · EDMO<br>Tier-1 Report Analysis<br>
      Internal use only · Self-contained</div>
  </div>
  <div class="content" id="content"></div>
</div>

<script>
/* ==================== DATA ==================== */
const COLS = {file:0,out:1,owner:2,de:3,src:4,sty:5,mal:6,dbt:7,dbn:8,
              db:9,seg:10,ao:11,tao:12,downer:13};
const FIELD_LABELS = [
 ["out","Output Name"],["owner","Output Owner"],["de","Data Element"],
 ["src","Source Name"],["sty","Source Type"],
 ["mal","Source Application EUC / MAL Code"],["dbt","Database Type"],
 ["dbn","Database Name"],["db","Database"],
 ["seg","Business Segment / Corporate Function"],
 ["ao","Asset Owner Name"],["tao","Technology Asset Owner Name"],
 ["downer","Data Owner Name"]];
const DATA = __DATA__;
const PALETTE = ["#008A00","#54B848","#1E5E3A","#8CC63F","#2E7D52",
 "#A7D28D","#0B3D2E","#66A182","#C7E5B5","#94b894"];

/* ==================== STATE ==================== */
const state = {tab:"overview", seg:"ALL", dbt:"ALL", own:"ALL",
               sty:"ALL", ao:"ALL", q:""};
const TABS = [
 ["overview","⌂","Executive Overview"],
 ["segment","▦","Business Segment"],
 ["mal","⌗","MAL Code Analysis"],
 ["owners","👤","Asset Owners"],
 ["tech","🗄","Technology Landscape"],
 ["quality","✓","Data Quality"]];

/* ==================== HELPERS ==================== */
const esc = s => s==null ? "—" :
  String(s).replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;");
const fmt = n => (n==null?0:n).toLocaleString("en-US");
function filtered(){
  const q = state.q.toLowerCase();
  return DATA.filter(r =>
    (state.seg==="ALL" || (r[COLS.seg]||"Not Specified")===state.seg) &&
    (state.dbt==="ALL" || (r[COLS.dbt]||"Not Specified")===state.dbt) &&
    (state.own==="ALL" || (r[COLS.owner]||"Not Specified")===state.own) &&
    (state.sty==="ALL" || (r[COLS.sty]||"Not Specified")===state.sty) &&
    (state.ao==="ALL"  || (r[COLS.ao]||"Not Specified")===state.ao) &&
    (!q || [COLS.out,COLS.src,COLS.de,COLS.mal,COLS.dbn]
      .some(i => (r[i]||"").toLowerCase().includes(q))));
}
function uniq(rows, i){
  const s = new Set();
  for (const r of rows) if (r[i]!=null) s.add(r[i]);
  return s.size;
}
function groupBy(rows, i, blank){
  const m = new Map();
  for (const r of rows){
    const k = r[i]==null ? (blank||"Not Specified") : r[i];
    if (!m.has(k)) m.set(k, []);
    m.get(k).push(r);
  }
  return m;
}
const topN = (m, fn, n) => [...m.entries()]
  .map(([k,v]) => [k, fn(v), v]).sort((a,b)=>b[1]-a[1]).slice(0, n||9999);

/* ==================== CHARTS (SVG) ==================== */
function hbar(pairs, opts){
  opts = opts||{};
  pairs = pairs.filter(p=>p[1]>0).slice(0, opts.max||12);
  if(!pairs.length) return "<div class='empty'>No data for current filters</div>";
  const mx = Math.max(...pairs.map(p=>p[1]));
  const W = 560, LW = 195, RH = 28, GAP = 9;
  const H = pairs.length*(RH+GAP);
  let s = `<svg viewBox="0 0 ${W} ${H}" width="100%"
    xmlns="http://www.w3.org/2000/svg">`;
  pairs.forEach((p,i)=>{
    const y = i*(RH+GAP), bw = Math.max(4,(W-LW-70)*p[1]/mx);
    const c = i===0 ? "#0B3D2E" : "#008A00";
    const click = opts.click ?
      ` class="bar-click" onclick="${opts.click}('${String(p[0])
        .replace(/'/g,"\\'")}')"` : "";
    s += `<g${click}><text x="${LW-8}" y="${y+RH/2+4}" text-anchor="end"
      font-size="11.5" fill="#3c4a3c"
      font-family="Segoe UI,Arial">${esc(String(p[0]).slice(0,36))}</text>
      <rect x="${LW}" y="${y+4}" rx="6" width="${bw.toFixed(1)}"
      height="${RH-8}" fill="${c}" opacity=".92"/>
      <text x="${(LW+bw+8).toFixed(1)}" y="${y+RH/2+4}" font-size="11.5"
      font-weight="650" fill="#0B3D2E"
      font-family="Segoe UI,Arial">${fmt(p[1])}</text></g>`;
  });
  return s + "</svg>" +
    (opts.click ? "<div class='hint'>Tip: click a bar to filter the whole dashboard.</div>" : "");
}
function donut(pairs, unitLabel, clickFn){
  pairs = pairs.filter(p=>p[1]>0);
  if(!pairs.length) return "<div class='empty'>No data for current filters</div>";
  if(pairs.length>8){
    const head = pairs.slice(0,7);
    head.push(["Other", pairs.slice(7).reduce((a,p)=>a+p[1],0)]);
    pairs = head;
  }
  const total = pairs.reduce((a,p)=>a+p[1],0);
  const SZ=210, R=72, SW=30, CIRC=2*Math.PI*R;
  let off = CIRC*0.25;
  let s = `<svg viewBox="0 0 ${SZ} ${SZ}" width="${SZ}"
    xmlns="http://www.w3.org/2000/svg">`;
  pairs.forEach((p,i)=>{
    const dash = p[1]/total*CIRC;
    s += `<circle cx="${SZ/2}" cy="${SZ/2}" r="${R}" fill="none"
      stroke="${PALETTE[i%PALETTE.length]}" stroke-width="${SW}"
      stroke-dasharray="${dash.toFixed(2)} ${(CIRC-dash).toFixed(2)}"
      stroke-dashoffset="${off.toFixed(2)}"/>`;
    off -= dash;
  });
  s += `<text x="${SZ/2}" y="${SZ/2-3}" text-anchor="middle" font-size="24"
    font-weight="700" fill="#0B3D2E"
    font-family="Segoe UI,Arial">${fmt(total)}</text>
    <text x="${SZ/2}" y="${SZ/2+17}" text-anchor="middle" font-size="10.5"
    fill="#6b7f6b" font-family="Segoe UI,Arial">${unitLabel}</text></svg>`;
  const legend = pairs.map((p,i)=>{
    const click = clickFn && p[0]!=="Other" ?
      ` onclick="${clickFn}('${String(p[0]).replace(/'/g,"\\'")}')"` : "";
    return `<span class="lg"${click}><i style="background:${
      PALETTE[i%PALETTE.length]}"></i>${esc(p[0])} <b>${fmt(p[1])}</b></span>`;
  }).join("");
  return `<div class="donutwrap">${s}<div class="legend">${legend}</div></div>`;
}
function table(headers, rows, opts){
  opts = opts||{};
  if(!rows.length) return "<div class='empty'>No data for current filters</div>";
  const th = headers.map(h=>`<th>${esc(h)}</th>`).join("");
  const body = rows.slice(0, opts.max||15).map(r=>"<tr>"+r.map((v,i)=>{
    if (typeof v==="number") return `<td class="num">${fmt(v)}</td>`;
    if (opts.flagCol===i){
      const cls = v==="Complete" ? "ok" : "gap";
      return `<td class="${cls}">${esc(v)}</td>`;
    }
    return `<td>${esc(v)}</td>`;
  }).join("")+"</tr>").join("");
  const t = `<table class="tbl"><thead><tr>${th}</tr></thead>
    <tbody>${body}</tbody></table>`;
  return opts.scroll ? `<div class="scrolltbl">${t}</div>` : t;
}
const kpiCards = items => `<div class="kpis">${items.map(([l,v])=>
  `<div class="kpi"><b>${fmt(v)}</b><span>${esc(l)}</span></div>`)
  .join("")}</div>`;

/* click-to-filter handlers (referenced by generated SVG) */
function setSeg(v){ state.seg=v; syncSelects(); render(); }
function setDbt(v){ state.dbt=v; syncSelects(); render(); }
function setAo(v){ state.ao=v; syncSelects(); render(); }

/* ==================== TAB RENDERERS ==================== */
function tabOverview(rows){
  const segTop = topN(groupBy(rows, COLS.seg), v=>uniq(v, COLS.out));
  const dbtTop = topN(groupBy(rows, COLS.dbt), v=>v.length);
  const malMissing = rows.filter(r=>r[COLS.mal]==null).length;
  const story = [];
  story.push(`The filtered view spans <b>${fmt(uniq(rows,COLS.out))} outputs</b>
    owned by <b>${fmt(uniq(rows,COLS.owner))} output owners</b>, drawing on
    <b>${fmt(uniq(rows,COLS.src))} source systems</b> across
    <b>${fmt(uniq(rows,COLS.mal))} EUC/MAL-coded applications</b>.`);
  if(dbtTop.length) story.push(`<b>${esc(dbtTop[0][0])}</b> is the dominant
    platform with ${fmt(uniq(dbtTop[0][2],COLS.dbn))} databases behind
    ${fmt(uniq(dbtTop[0][2],COLS.out))} outputs.`);
  if(segTop.length) story.push(`<b>${esc(segTop[0][0])}</b> carries the largest
    reporting footprint (${fmt(segTop[0][1])} outputs).`);
  if(malMissing) story.push(`<b>${fmt(malMissing)} rows</b> are missing an
    EUC/MAL code — see the Data Quality tab for the remediation list.`);
  return `
  ${kpiCards([["Outputs",uniq(rows,COLS.out)],
    ["Output Owners",uniq(rows,COLS.owner)],
    ["Source Systems",uniq(rows,COLS.src)],
    ["Source Types",uniq(rows,COLS.sty)],
    ["EUC / MAL Codes",uniq(rows,COLS.mal)],
    ["Databases",uniq(rows,COLS.dbn)],
    ["Database Types",uniq(rows,COLS.dbt)],
    ["Data Elements",uniq(rows,COLS.de)],
    ["Segments / Functions",uniq(rows,COLS.seg)],
    ["Asset Owners",uniq(rows,COLS.ao)],
    ["Data Owners",uniq(rows,COLS.downer)],
    ["Data Element Rows",rows.length]])}
  <div class="story"><h3>Executive Summary</h3>
    <ul>${story.map(s=>`<li>${s}</li>`).join("")}</ul></div>
  <div class="grid2">
    <div class="sec"><h2>Rows by Database Type</h2>
      <div class="note">Normalized platform view — click a legend item to
      filter.</div>${donut(dbtTop, "rows", "setDbt")}</div>
    <div class="sec"><h2>Outputs by Business Segment</h2>
      <div class="note">Distinct outputs per segment / corporate
      function.</div>${hbar(segTop, {click:"setSeg"})}</div>
  </div>
  <div class="sec"><h2>Top Outputs by Data-Element Footprint</h2>
    <div class="note">Concentration of Tier-1 data elements.</div>
    ${table(["Output Name","Output Owner","Data Elements","Sources","Databases"],
      topN(groupBy(rows,COLS.out,"Not Specified"),v=>uniq(v,COLS.de),12)
      .map(([k,n,v])=>[k, v.find(r=>r[COLS.owner])?.[COLS.owner]||"—",
        n, uniq(v,COLS.src), uniq(v,COLS.dbn)]))}</div>`;
}

function tabSegment(rows){
  const m = groupBy(rows, COLS.seg);
  const t = topN(m, v=>uniq(v,COLS.out));
  return `
  ${kpiCards([["Segments / Functions",m.size],
    ["Outputs",uniq(rows,COLS.out)],
    ["Output Owners",uniq(rows,COLS.owner)],
    ["Data Elements",uniq(rows,COLS.de)]])}
  <div class="grid2">
    <div class="sec"><h2>Outputs by Segment</h2>
      <div class="note">Click a bar to focus the dashboard on one
      segment.</div>${hbar(t,{click:"setSeg"})}</div>
    <div class="sec"><h2>Segment Detail</h2>
      <div class="note">Owners, elements, sources and databases per
      segment.</div>
      ${table(["Segment / Function","Outputs","Owners","Elements","Sources",
        "Databases"], t.map(([k,n,v])=>[k,n,uniq(v,COLS.owner),
        uniq(v,COLS.de),uniq(v,COLS.src),uniq(v,COLS.dbn)]),
        {scroll:true,max:100})}</div>
  </div>
  <div class="sec"><h2>Segment × Database Type Mix</h2>
    <div class="note">Which platforms each segment depends on.</div>
    ${table(["Segment / Function","Database Types Used","Top Platform",
      "Rows"], t.map(([k,n,v])=>{
        const dm = topN(groupBy(v,COLS.dbt),x=>x.length);
        return [k, dm.length, dm.length?dm[0][0]+" ("+fmt(dm[0][1])+")":"—",
          v.length];}), {scroll:true,max:100})}</div>`;
}

function tabMal(rows){
  const m = groupBy(rows, COLS.mal, "«Code Missing»");
  const withCode = rows.filter(r=>r[COLS.mal]!=null).length;
  const pct = rows.length? Math.round(100*withCode/rows.length) : 0;
  const t = topN(m, v=>v.length);
  return `
  ${kpiCards([["Distinct EUC / MAL Codes",uniq(rows,COLS.mal)],
    ["Rows with a Code",withCode],
    ["Rows Missing a Code",rows.length-withCode],
    ["Code Coverage %",pct]])}
  <div class="sec"><h2>MAL / EUC Code Inventory</h2>
    <div class="note">Every application code with its sources, outputs and
    database footprint. «Code Missing» groups the uncoded rows —
    a key remediation queue for application-inventory alignment.</div>
    ${table(["EUC / MAL Code","Sources","Source Types","Outputs","Databases",
      "Rows"], t.map(([k,n,v])=>[k,uniq(v,COLS.src),uniq(v,COLS.sty),
      uniq(v,COLS.out),uniq(v,COLS.dbn),n]), {scroll:true,max:300})}</div>
  <div class="grid2">
    <div class="sec"><h2>Rows per Code (Top 12)</h2>
      ${hbar(t.slice(0,12))}</div>
    <div class="sec"><h2>Sources Missing a Code</h2>
      <div class="note">Source systems appearing without any EUC/MAL
      code.</div>
      ${table(["Source Name","Source Type","Rows"],
        topN(groupBy(rows.filter(r=>r[COLS.mal]==null), COLS.src,
        "«Source Missing»"), v=>v.length)
        .map(([k,n,v])=>[k, v.find(r=>r[COLS.sty])?.[COLS.sty]||"—", n]),
        {scroll:true,max:100})}</div>
  </div>`;
}

function tabOwners(rows){
  const ao = topN(groupBy(rows, COLS.ao), v=>uniq(v,COLS.dbn));
  const dow = topN(groupBy(rows, COLS.downer), v=>v.length);
  return `
  ${kpiCards([["Asset Owners",uniq(rows,COLS.ao)],
    ["Technology Asset Owners",uniq(rows,COLS.tao)],
    ["Data Owners",uniq(rows,COLS.downer)],
    ["Rows Missing Asset Owner",rows.filter(r=>r[COLS.ao]==null).length],
    ["Rows Missing Data Owner",rows.filter(r=>r[COLS.downer]==null).length]])}
  <div class="grid2">
    <div class="sec"><h2>Databases by Asset Owner</h2>
      <div class="note">Custody concentration — click to filter.</div>
      ${hbar(ao,{click:"setAo"})}</div>
    <div class="sec"><h2>Asset Owner Detail</h2>
      ${table(["Asset Owner","Databases","DB Types","Outputs","Tech Owners"],
      ao.map(([k,n,v])=>[k,n,uniq(v,COLS.dbt),uniq(v,COLS.out),
      uniq(v,COLS.tao)]), {scroll:true,max:200})}</div>
  </div>
  <div class="grid2">
    <div class="sec"><h2>Data Owner Coverage</h2>
      <div class="note">Rows accountable to each data owner.</div>
      ${table(["Data Owner","Rows","Outputs","Elements"],
        dow.map(([k,n,v])=>[k,n,uniq(v,COLS.out),uniq(v,COLS.de)]),
        {scroll:true,max:200})}</div>
    <div class="sec"><h2>Technology Asset Owners</h2>
      ${table(["Technology Asset Owner","Databases","DB Types","Rows"],
        topN(groupBy(rows,COLS.tao),v=>uniq(v,COLS.dbn))
        .map(([k,n,v])=>[k,n,uniq(v,COLS.dbt),v.length]),
        {scroll:true,max:200})}</div>
  </div>`;
}

function tabTech(rows){
  const m = groupBy(rows, COLS.dbt);
  const t = topN(m, v=>v.length);
  const inv = [];
  for (const [dt, v] of m){
    for (const [dbn, vv] of groupBy(v, COLS.dbn, "«Name Missing»"))
      inv.push([dt, dbn, uniq(vv,COLS.out), uniq(vv,COLS.src), vv.length]);
  }
  inv.sort((a,b)=> a[0]===b[0] ? b[4]-a[4] : (a[0]<b[0]?-1:1));
  const sty = topN(groupBy(rows, COLS.sty), v=>uniq(v,COLS.dbn));
  return `
  ${kpiCards([["Database Types",m.size],
    ["Databases",uniq(rows,COLS.dbn)],
    ["Source Systems",uniq(rows,COLS.src)],
    ["Source Types",uniq(rows,COLS.sty)],
    ["Rows Missing DB Name",rows.filter(r=>r[COLS.dbn]==null).length]])}
  <div class="grid2">
    <div class="sec"><h2>Platform Mix</h2>
      <div class="note">Normalized Database Types (ORACLE / Oracle 19c /
      Essbase → Oracle, etc.). Click a legend item to filter.</div>
      ${donut(t,"rows","setDbt")}</div>
    <div class="sec"><h2>Distinct Databases per Platform</h2>
      ${hbar(topN(m, v=>uniq(v,COLS.dbn)),{click:"setDbt"})}</div>
  </div>
  <div class="sec"><h2>Database Inventory (Name × Type)</h2>
    <div class="note">Every database with its platform, outputs and
    sources.</div>
    ${table(["Database Type","Database Name","Outputs","Sources","Rows"],
      inv, {scroll:true,max:500})}</div>
  <div class="sec"><h2>Databases by Source Type</h2>
    ${table(["Source Type","Databases","Sources","Rows"],
      sty.map(([k,n,v])=>[k,n,uniq(v,COLS.src),v.length]),
      {scroll:true,max:100})}</div>`;
}

function tabQuality(rows){
  const meters = FIELD_LABELS.map(([key,label])=>{
    const miss = rows.filter(r=>r[COLS[key]]==null).length;
    const pct = rows.length? Math.round(1000*(1-miss/rows.length))/10 : 0;
    const cls = pct>=95?"":(pct>=75?"warn":"bad");
    return `<div class="dqrow"><div class="lbl"><span>${esc(label)}</span>
      <b>${pct}%</b></div><div class="meter"><i class="${cls}"
      style="width:${pct}%"></i></div>
      <div style="font-size:10.5px;color:#9ab09a">${fmt(miss)} missing
      rows</div></div>`;
  }).join("");
  const key = ["out","src","dbn","downer","ao","mal"];
  const score = rows.length ? Math.round(10*key.reduce((a,k)=>
    a+100*(1-rows.filter(r=>r[COLS[k]]==null).length/rows.length),0)
    /key.length)/10 : 0;
  const srcMap = new Map();
  for (const r of rows){
    const k = [r[COLS.src]||"«Source Missing»", r[COLS.mal]||"«Code Missing»",
               r[COLS.sty]||"«Type Missing»"].join("|");
    if(!srcMap.has(k)) srcMap.set(k, []);
    srcMap.get(k).push(r);
  }
  const gapRows = [...srcMap.entries()].map(([k,v])=>{
    const [src,mal,sty] = k.split("|");
    const flags = [];
    if(src==="«Source Missing»") flags.push("Source missing");
    if(mal==="«Code Missing»") flags.push("EUC/MAL missing");
    if(v.some(r=>r[COLS.dbn]==null)) flags.push("DB name gaps");
    if(v.some(r=>r[COLS.downer]==null)) flags.push("Data owner gaps");
    if(v.some(r=>r[COLS.ao]==null)) flags.push("Asset owner gaps");
    return [src,mal,sty,uniq(v,COLS.dbn),v.length,
            flags.length?flags.join("; "):"Complete"];
  }).sort((a,b)=> (a[5]==="Complete") - (b[5]==="Complete") || b[4]-a[4]);
  const gapCt = gapRows.filter(r=>r[5]!=="Complete").length;
  return `
  ${kpiCards([["Key-Field Completeness %",score],
    ["Source Lines with Gaps",gapCt],
    ["Rows Missing DB Name",rows.filter(r=>r[COLS.dbn]==null).length],
    ["Rows Missing Data Owner",rows.filter(r=>r[COLS.downer]==null).length],
    ["Rows Missing MAL Code",rows.filter(r=>r[COLS.mal]==null).length]])}
  <div class="grid2">
    <div class="sec"><h2>Field Completeness</h2>
      <div class="note">Populated % per extracted field, under current
      filters. Green ≥95 · amber ≥75 · red &lt;75.</div>${meters}</div>
    <div class="sec"><h2>Source Gap Register</h2>
      <div class="note">Each source line with its flags — amber rows are the
      remediation queue.</div>
      ${table(["Source Name","EUC / MAL Code","Source Type","Databases",
        "Rows","Gap Flags"], gapRows, {scroll:true,max:500,flagCol:5})}</div>
  </div>`;
}

/* ==================== SHELL ==================== */
const RENDERERS = {overview:tabOverview, segment:tabSegment, mal:tabMal,
                   owners:tabOwners, tech:tabTech, quality:tabQuality};

function buildNav(){
  document.getElementById("nav").innerHTML = TABS.map(([id,ic,label])=>
    `<button class="navbtn${state.tab===id?" active":""}"
     onclick="state.tab='${id}';buildNav();render()">
     <span class="ic">${ic}</span>${label}</button>`).join("");
}
function fillSelect(id, idx, label){
  const vals = [...new Set(DATA.map(r=>r[idx]==null?"Not Specified":r[idx]))]
    .sort((a,b)=>a.localeCompare(b));
  document.getElementById(id).innerHTML =
    `<option value="ALL">All ${label} (${vals.length})</option>` +
    vals.map(v=>`<option>${esc(v)}</option>`).join("");
}
function syncSelects(){
  f_seg.value=state.seg; f_dbt.value=state.dbt; f_own.value=state.own;
  f_sty.value=state.sty; f_ao.value=state.ao; f_q.value=state.q;
}
function render(){
  const rows = filtered();
  const active = ["seg","dbt","own","sty","ao"]
    .filter(k=>state[k]!=="ALL").length + (state.q?1:0);
  document.getElementById("f_chip").textContent =
    `${fmt(rows.length)} of ${fmt(DATA.length)} rows` +
    (active? ` · ${active} filter${active>1?"s":""}` : "");
  document.getElementById("content").innerHTML =
    RENDERERS[state.tab](rows);
  document.querySelector(".content").scrollTop = 0;
}
window.addEventListener("DOMContentLoaded", ()=>{
  fillSelect("f_seg", COLS.seg, "Segments");
  fillSelect("f_dbt", COLS.dbt, "Types");
  fillSelect("f_own", COLS.owner, "Owners");
  fillSelect("f_sty", COLS.sty, "Types");
  fillSelect("f_ao", COLS.ao, "Owners");
  const bind = (id,key)=>document.getElementById(id)
    .addEventListener("change", e=>{state[key]=e.target.value; render();});
  bind("f_seg","seg"); bind("f_dbt","dbt"); bind("f_own","own");
  bind("f_sty","sty"); bind("f_ao","ao");
  let deb;
  document.getElementById("f_q").addEventListener("input", e=>{
    clearTimeout(deb);
    deb = setTimeout(()=>{state.q=e.target.value; render();}, 180);
  });
  document.getElementById("f_reset").addEventListener("click", ()=>{
    Object.assign(state,{seg:"ALL",dbt:"ALL",own:"ALL",sty:"ALL",
      ao:"ALL",q:""});
    syncSelects(); render();
  });
  buildNav(); render();
});
</script></body></html>"""


def generate_dashboard(folder, analytics, combined, excel_path) -> str:
    stamp = datetime.now()
    out_path = os.path.join(
        folder,
        f"DSCIT_Executive_Dashboard_{stamp.strftime('%Y%m%d_%H%M%S')}.html")
    data_json = json.dumps(_dash_rows(combined), ensure_ascii=False,
                           separators=(",", ":")).replace("</", "<\\/")
    doc = (DASH_TEMPLATE
           .replace("__GENERATED__",
                    stamp.strftime("%B %d, %Y at %H:%M"))
           .replace("__EXCEL__",
                    html.escape(os.path.basename(excel_path or "—")))
           .replace("__DATA__", data_json))
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(doc)
    return out_path


# ============================================================================
# 7. TD-THEMED UI
# ============================================================================

TD_GREEN = "#008A00"
TD_GREEN_DARK = "#006A00"
TD_GREEN_LIGHT = "#54B848"
TD_MIST = "#F3F7F3"
CHARCOAL = "#2E2E2E"
GREY = "#6B7280"
WHITE = "#FFFFFF"
AMBER = "#B45309"
RED = "#B91C1C"


def launch_ui():
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox

    root = tk.Tk()
    root.title("TD  |  DSCIT Tier-1 Report Consolidator & Executive Analytics")
    root.geometry("1160x790")
    root.configure(bg=TD_MIST)
    root.minsize(980, 660)

    ui_queue = queue.Queue()
    state = {"running": False, "folder": tk.StringVar(value=""),
             "excel": None, "html": None,
             "results": None, "combined": None}

    style = ttk.Style(root)
    style.theme_use("clam")
    style.configure("TFrame", background=TD_MIST)
    style.configure("Header.TFrame", background=TD_GREEN)
    style.configure("TLabel", background=TD_MIST, foreground=CHARCOAL,
                    font=("Segoe UI", 10))
    style.configure("StatValue.TLabel", background=WHITE, foreground=TD_GREEN,
                    font=("Segoe UI Semibold", 22))
    style.configure("StatCap.TLabel", background=WHITE, foreground=GREY,
                    font=("Segoe UI", 9))
    style.configure("HeaderTitle.TLabel", background=TD_GREEN, foreground=WHITE,
                    font=("Segoe UI Semibold", 17))
    style.configure("HeaderSub.TLabel", background=TD_GREEN, foreground="#D8F0D8",
                    font=("Segoe UI", 10))
    style.configure("TD.TButton", background=TD_GREEN, foreground=WHITE,
                    font=("Segoe UI Semibold", 10), padding=(16, 8),
                    borderwidth=0)
    style.map("TD.TButton",
              background=[("active", TD_GREEN_DARK), ("disabled", "#9CC79C")])
    style.configure("Gold.TButton", background="#0B3D2E", foreground=WHITE,
                    font=("Segoe UI Semibold", 10), padding=(16, 8),
                    borderwidth=0)
    style.map("Gold.TButton",
              background=[("active", "#092E23"), ("disabled", "#7C948A")])
    style.configure("Ghost.TButton", background=WHITE, foreground=TD_GREEN,
                    font=("Segoe UI", 10), padding=(12, 7), borderwidth=1)
    style.map("Ghost.TButton", background=[("active", TD_MIST)])
    style.configure("TD.Horizontal.TProgressbar", troughcolor="#E3EAE3",
                    background=TD_GREEN_LIGHT, thickness=14, borderwidth=0)
    style.configure("Treeview", font=("Segoe UI", 9), rowheight=26,
                    background=WHITE, fieldbackground=WHITE, foreground=CHARCOAL)
    style.configure("Treeview.Heading", font=("Segoe UI Semibold", 9),
                    background=TD_MIST, foreground=CHARCOAL)

    # header
    header = ttk.Frame(root, style="Header.TFrame")
    header.pack(fill="x")
    hin = ttk.Frame(header, style="Header.TFrame")
    hin.pack(fill="x", padx=24, pady=16)
    logo = tk.Canvas(hin, width=46, height=46, bg=TD_GREEN, highlightthickness=0)
    logo.create_rectangle(2, 2, 44, 44, fill=WHITE, outline=WHITE)
    logo.create_text(23, 23, text="TD", fill=TD_GREEN,
                     font=("Segoe UI Black", 16, "bold"))
    logo.pack(side="left", padx=(0, 14))
    tbox = ttk.Frame(hin, style="Header.TFrame")
    tbox.pack(side="left")
    ttk.Label(tbox, text="DSCIT Tier-1 Consolidator & Executive Analytics",
              style="HeaderTitle.TLabel").pack(anchor="w")
    ttk.Label(tbox, text="Enterprise Data Management Office  ·  Tier-1 Sources",
              style="HeaderSub.TLabel").pack(anchor="w")

    body = ttk.Frame(root)
    body.pack(fill="both", expand=True, padx=24, pady=18)

    # folder card
    fcard = tk.Frame(body, bg=WHITE, highlightbackground="#E2E8E2",
                     highlightthickness=1)
    fcard.pack(fill="x")
    frow = tk.Frame(fcard, bg=WHITE)
    frow.pack(fill="x", padx=18, pady=14)
    tk.Label(frow, text="Report Folder", bg=WHITE, fg=GREY,
             font=("Segoe UI Semibold", 9)).pack(anchor="w")
    inner = tk.Frame(frow, bg=WHITE)
    inner.pack(fill="x", pady=(6, 0))
    entry = tk.Entry(inner, textvariable=state["folder"], font=("Segoe UI", 10),
                     bg=TD_MIST, fg=CHARCOAL, relief="flat",
                     highlightthickness=1, highlightbackground="#D5DDD5")
    entry.pack(side="left", fill="x", expand=True, ipady=7, padx=(0, 10))

    stat_vars = {}
    status_var = tk.StringVar(
        value="Select your Tier-1 Report Analysis folder to begin.")

    def browse():
        d = filedialog.askdirectory(
            title="Select the Tier-1 Report Analysis folder")
        if d:
            state["folder"].set(d)
            n = len(discover_files(d))
            stat_vars["found"].set(str(n))
            status_var.set(f"{n} DSCIT file(s) detected — ready to run."
                           if n else "No DSCIT files detected in this folder.")

    ttk.Button(inner, text="Browse…", style="Ghost.TButton",
               command=browse).pack(side="left", padx=(0, 8))
    run_btn = ttk.Button(inner, text="▶  Run Consolidation", style="TD.TButton")
    run_btn.pack(side="left", padx=(0, 8))
    dash_btn = ttk.Button(inner, text="★  Generate Executive Dashboard",
                          style="Gold.TButton", state="disabled")
    dash_btn.pack(side="left")

    # stats row
    stats = tk.Frame(body, bg=TD_MIST)
    stats.pack(fill="x", pady=(16, 0))

    def stat_card(parent, key, caption, last=False):
        c = tk.Frame(parent, bg=WHITE, highlightbackground="#E2E8E2",
                     highlightthickness=1)
        c.pack(side="left", fill="x", expand=True,
               padx=(0, 0 if last else 12))
        v = tk.StringVar(value="0")
        stat_vars[key] = v
        ttk.Label(c, textvariable=v, style="StatValue.TLabel").pack(
            anchor="w", padx=16, pady=(12, 0))
        ttk.Label(c, text=caption, style="StatCap.TLabel").pack(
            anchor="w", padx=16, pady=(0, 12))

    stat_card(stats, "found", "FILES FOUND")
    stat_card(stats, "processed", "FILES PROCESSED")
    stat_card(stats, "rows", "ROWS EXTRACTED")
    stat_card(stats, "warn", "WARNINGS")
    stat_card(stats, "err", "ERRORS", last=True)

    # progress
    pcard = tk.Frame(body, bg=TD_MIST)
    pcard.pack(fill="x", pady=(16, 0))
    pbar = ttk.Progressbar(pcard, style="TD.Horizontal.TProgressbar",
                           mode="determinate")
    pbar.pack(fill="x")
    tk.Label(pcard, textvariable=status_var, bg=TD_MIST, fg=GREY,
             font=("Segoe UI", 9)).pack(anchor="w", pady=(6, 0))

    # results table
    tcard = tk.Frame(body, bg=WHITE, highlightbackground="#E2E8E2",
                     highlightthickness=1)
    tcard.pack(fill="both", expand=True, pady=(16, 0))
    cols = ("file", "status", "sheet", "hdr", "rows", "matched", "notes")
    tree = ttk.Treeview(tcard, columns=cols, show="headings")
    heads = {"file": ("File", 300), "status": ("Status", 70),
             "sheet": ("Sheet", 110), "hdr": ("Header Row", 85),
             "rows": ("Rows", 70), "matched": ("Cols Matched", 95),
             "notes": ("Notes", 300)}
    for c, (txt, w) in heads.items():
        tree.heading(c, text=txt)
        tree.column(c, width=w, anchor="w")
    vsb = ttk.Scrollbar(tcard, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)
    tree.pack(side="left", fill="both", expand=True, padx=(1, 0), pady=1)
    vsb.pack(side="right", fill="y")
    tree.tag_configure("OK", foreground="#006A00")
    tree.tag_configure("WARN", foreground=AMBER)
    tree.tag_configure("ERROR", foreground=RED)

    # footer buttons
    foot = tk.Frame(body, bg=TD_MIST)
    foot.pack(fill="x", pady=(12, 0))

    def _open(path):
        if not path:
            return
        try:
            os.startfile(path)          # Windows
        except AttributeError:
            webbrowser.open(f"file://{os.path.abspath(path)}")

    open_html_btn = ttk.Button(foot, text="Open Dashboard (HTML)",
                               style="Ghost.TButton", state="disabled",
                               command=lambda: _open(state["html"]))
    open_html_btn.pack(side="right")
    open_xl_btn = ttk.Button(foot, text="Open Workbook (Excel)",
                             style="Ghost.TButton", state="disabled",
                             command=lambda: _open(state["excel"]))
    open_xl_btn.pack(side="right", padx=(0, 8))
    open_folder_btn = ttk.Button(
        foot, text="Open Folder", style="Ghost.TButton", state="disabled",
        command=lambda: state["excel"] and _open(os.path.dirname(state["excel"])))
    open_folder_btn.pack(side="right", padx=(0, 8))

    # ---------------- workers ----------------
    def consolidation_worker(folder):
        try:
            def cb(done, total, res):
                ui_queue.put(("progress", done, total, res))
            results, combined = run_consolidation(folder, progress_cb=cb)
            analytics = build_analytics(combined, results) if len(combined) else None
            out = write_output(folder, results, combined, analytics)
            ui_queue.put(("done", results, combined, out))
        except Exception:
            ui_queue.put(("fatal", traceback.format_exc()))

    def dashboard_worker(folder):
        try:
            combined = state["combined"]
            analytics = build_analytics(combined, state["results"])
            html_path = generate_dashboard(folder, analytics, combined,
                                           state["excel"] or "")
            ui_queue.put(("dash_done", html_path))
        except Exception:
            ui_queue.put(("fatal", traceback.format_exc()))

    def start():
        folder = state["folder"].get().strip()
        if not os.path.isdir(folder):
            messagebox.showwarning("Folder required",
                                   "Please browse to a valid folder first.")
            return
        files = discover_files(folder)
        if not files:
            messagebox.showinfo(
                "No files", "No DSCIT*.xlsx files were found in that folder.")
            return
        for i in tree.get_children():
            tree.delete(i)
        for kk in ("processed", "rows", "warn", "err"):
            stat_vars[kk].set("0")
        stat_vars["found"].set(str(len(files)))
        pbar["value"] = 0
        pbar["maximum"] = len(files)
        state.update(running=True, excel=None, html=None,
                     results=None, combined=None)
        run_btn.state(["disabled"])
        dash_btn.state(["disabled"])
        for b in (open_html_btn, open_xl_btn, open_folder_btn):
            b.state(["disabled"])
        status_var.set("Running consolidation…")
        threading.Thread(target=consolidation_worker, args=(folder,),
                         daemon=True).start()

    def start_dashboard():
        if state["combined"] is None or state["combined"].empty:
            messagebox.showinfo("Run first",
                                "Run a consolidation before generating "
                                "the dashboard.")
            return
        dash_btn.state(["disabled"])
        status_var.set("Building executive dashboard…")
        threading.Thread(target=dashboard_worker,
                         args=(state["folder"].get().strip(),),
                         daemon=True).start()

    run_btn.configure(command=start)
    dash_btn.configure(command=start_dashboard)

    def poll():
        try:
            while True:
                msg = ui_queue.get_nowait()
                if msg[0] == "progress":
                    _, done, total, res = msg
                    pbar["value"] = done
                    stat_vars["processed"].set(str(done))
                    stat_vars["rows"].set(
                        str(int(stat_vars["rows"].get()) + res.rows))
                    if res.status == "WARN":
                        stat_vars["warn"].set(
                            str(int(stat_vars["warn"].get()) + 1))
                    if res.status == "ERROR":
                        stat_vars["err"].set(
                            str(int(stat_vars["err"].get()) + 1))
                    tree.insert("", "end", tags=(res.status,), values=(
                        res.file, res.status, res.sheet,
                        res.header_row or "—", res.rows,
                        f"{res.matched}/{len(TARGET_COLUMNS)}",
                        (res.message + ("  |  Missing: " + ", ".join(res.missing)
                                        if res.missing else ""))[:180]))
                    status_var.set(f"Processed {done} of {total}…")
                elif msg[0] == "done":
                    _, results, combined, out = msg
                    state.update(running=False, excel=out,
                                 results=results, combined=combined)
                    run_btn.state(["!disabled"])
                    dash_btn.state(["!disabled"])
                    open_xl_btn.state(["!disabled"])
                    open_folder_btn.state(["!disabled"])
                    status_var.set(
                        f"Complete — {len(combined):,} rows from "
                        f"{sum(1 for r in results if r.status != 'ERROR')} "
                        f"file(s). Workbook saved with analytics tabs: "
                        f"{os.path.basename(out)}. Ready for the dashboard ★")
                elif msg[0] == "dash_done":
                    state["html"] = msg[1]
                    dash_btn.state(["!disabled"])
                    open_html_btn.state(["!disabled"])
                    status_var.set("Executive dashboard ready: "
                                   f"{os.path.basename(msg[1])}")
                    _open(msg[1])
                elif msg[0] == "fatal":
                    state["running"] = False
                    run_btn.state(["!disabled"])
                    dash_btn.state(["!disabled"])
                    status_var.set("Run failed — see details.")
                    messagebox.showerror("Run failed", msg[1][-1500:])
        except queue.Empty:
            pass
        root.after(100, poll)

    poll()
    root.mainloop()


# ============================================================================
# 8. ENTRY
# ============================================================================

if __name__ == "__main__":
    if "--headless" in sys.argv:
        folder = sys.argv[sys.argv.index("--headless") + 1]
        results, combined = run_consolidation(
            folder, progress_cb=lambda d, t, r: print(
                f"[{d}/{t}] {r.file}: {r.status} rows={r.rows} {r.message}"))
        analytics = build_analytics(combined, results)
        out = write_output(folder, results, combined, analytics)
        dash = generate_dashboard(folder, analytics, combined, out)
        print(f"\nWorkbook:  {out}  ({len(combined):,} rows)")
        print(f"Dashboard: {dash}")
    else:
        launch_ui()
